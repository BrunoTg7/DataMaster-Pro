"""
Minerador Pro v4.1 – Motor Profissional de Rastreio de Preços e Dados da Web
100% Offline/Local, Sem Uso de IA.

Funcionalidades:
- Selector Registry embutido para os maiores marketplaces brasileiros
- Seletores CSS/XPath totalmente customizáveis pelo usuário (qualquer site)
- Motor Anti-Bot com emulação orgânica de comportamento humano
- Exportação profissional para Excel com tema visual personalizado

Correções v4.1:
- extract_via_soup usa soup.select_one() para suportar seletores CSS complexos com atributos
- Retry automático (até 2x) antes de acionar o fallback ScraperAPI
- Stealth script expandido para melhor evasão de anti-bot
- Log explícito quando ScraperAPI key está ausente
- Aguarda JS renderizar após scroll antes de extrair dados
- Fallback de título via page.title() quando todos os seletores falham
"""
import asyncio
import logging
import re
import random
import os
import sys
import time
import pandas as pd
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Any, Callable
from urllib.parse import urlparse
import config

log = logging.getLogger(__name__)
from src.utils.user_agents import UserAgentProvider
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── URL validation ──────────────────────────────────────────────────────────
ALLOWED_URL_SCHEMES = {"http", "https"}
BANNED_SCHEMES = {"javascript", "data", "file", "ftp", "ftps", "sftp", "blob", "about", "chrome", "edge", "vbscript"}

def validate_url(url: str) -> Optional[str]:
    """Valida e normaliza uma URL. Retorna URL limpa ou None se inválida."""
    if not url or not isinstance(url, str):
        return None
    url = url.strip().strip('"\'')
    if not url:
        return None
    try:
        parsed = urlparse(url)
        if parsed.scheme and parsed.scheme.lower() in BANNED_SCHEMES:
            return None
        if not parsed.scheme:
            clean = url.lstrip("/")
            url = "https://" + clean
            parsed = urlparse(url)
        if parsed.scheme.lower() not in ALLOWED_URL_SCHEMES:
            return None
        if not parsed.netloc:
            return None
        return url
    except Exception:
        return None

# ─────────────────────────────────────────────────────────────────────────────
# SELECTOR REGISTRY: seletores CSS pré-calibrados por marketplace
# ─────────────────────────────────────────────────────────────────────────────
SELECTOR_REGISTRY: Dict[str, Dict[str, List[str]]] = {
    "mercadolivre": {
        "title":        ["h1", ".ui-pdp-title", "[class*='product-title']", "h1.ui-pdp-title"],
        "price":        [
            "[class*='price']", "[class*='andes-money-amount']",
            "span.andes-money-amount__fraction",
            ".price-tag-fraction", ".ui-pdp-price__second-line span",
            "[itemprop='price']",
        ],
        "availability": ["[class*='stock']", ".ui-pdp-buybox__quantity", ".ui-pdp-stock-information"],
        "rating":       ["[class*='rating']", ".ui-pdp-review__rating", ".ui-pdp-stars__rating--lg"],
        "seller":       ["[class*='seller']", ".ui-pdp-seller__header-info", ".ui-seller-data-header__title"],
    },
    "amazon": {
        "title":        ["h1", "#productTitle", "[class*='product-title']", "h1.a-size-large"],
        "price":        [
            "[class*='price']", "[class*='a-price']",
            "span.a-price-whole", "#priceblock_ourprice",
            ".a-price .a-offscreen", "[itemprop='price']",
        ],
        "availability": ["#availability span", "#outOfStock", "[class*='availability']"],
        "rating":       ["[class*='rating']", "span.a-icon-alt", "#acrPopover span.a-size-base"],
        "seller":       ["#bylineInfo", "#merchant-info", "[class*='seller']"],
    },
    "shopee": {
        "title":        [
            "h1", "[class*='product-title']", "[class*='product-name']",
            "[data-sqe='name']", "._44qnta",
        ],
        "price":        [
            "[class*='price']", "[class*='preco']",
            "[itemprop='price']", "[data-price]",
            "[class*='price-box']",
        ],
        "availability": ["[class*='stock']"],
        "rating":       ["[class*='rating']", "[class*='stars']"],
        "seller":       ["[class*='seller']", "[class*='shop-name']"],
    },
    "magalu": {
        "title":        [
            "h1", ".header-product__title", "[class*='product-title']",
            "[data-testid='heading-product-title']",
        ],
        "price":        [
            "[class*='price']", "[class*='preco']", "[class*='current-price']",
            "[class*='final-price']",
            "p[data-testid='price-value']", ".price-template__text",
            "[itemprop='price']",
        ],
        "availability": ["[class*='stock']", "[data-testid='stock-label']", ".availability-badge"],
        "rating":       ["[class*='rating']", "[data-testid='review-score']", ".rating-score"],
        "seller":       ["[class*='seller']", "[data-testid='seller-name']", ".seller-label"],
    },
    "generico": {
        "title":        [
            "h1", "h2",
            "[class*='title']", "[class*='product-name']", "[class*='product-title']",
            "[class*='nome-produto']", "[class*='nome_produto']",
            "[itemprop='name']", "[data-product-name]",
        ],
        "price":        [
            "[class*='price']", "[class*='preco']", "[class*='valor']",
            "[class*='current-price']", "[class*='final-price']",
            "[class*='finalPrice']", "[class*='preco_desconto']",
            "[data-price]", "[data-product-price]",
            "[itemprop='price']",
        ],
        "availability": ["[class*='stock']", "[class*='estoque']", "[class*='availability']", "[class*='disponibilidade']"],
        "rating":       ["[class*='rating']", "[class*='avaliacao']", "[itemprop='ratingValue']"],
        "seller":       ["[class*='seller']", "[class*='vendedor']", "[class*='loja']", "[class*='shop-name']"],
    },
    "pichau": {
        "title":        ["h1", ".title", ".product-title", "h1[itemprop='name']"],
        "price":        [
            ".current-price",
            ".price",
            "[class*='preco']",
            "[data-product-price]",
            "[itemprop='price']",
        ],
        "availability": ["[class*='stock']", ".out-of-stock"],
        "rating":       ["[class*='rating']"],
        "seller":       [],
    },
    "kabum": {
        "title":        ["h1", ".titulo", ".product-title", "[class*='title']"],
        "price":        [
            ".preco_desconto",
            ".preco_normal",
            ".finalPrice",
            "[class*='preco']",
            "[itemprop='price']",
        ],
        "availability": ["[class*='stock']", ".disponibilidade"],
        "rating":       ["[class*='rating']", "[class*='avaliacao']"],
        "seller":       [],
    },
    "terabyteshop": {
        "title":        ["h1", ".tit-prod", ".product-name", "[class*='title']"],
        "price":        [
            ".preco-prod",
            ".preco_desconto",
            ".final-price",
            "[class*='preco']",
            "[itemprop='price']",
        ],
        "availability": ["[class*='stock']", ".alert-estoque"],
        "rating":       ["[class*='rating']"],
        "seller":       [],
    },
    "aliexpress": {
        "title":        ["h1", "[class*='title']", "[class*='product-title']", "[class*='name']"],
        "price":        [
            "[class*='price']", "[class*='preco']", "[class*='current-price']",
            "[itemprop='price']", "[data-price]",
        ],
        "availability": [],
        "rating":       ["[class*='rating']"],
        "seller":       ["[class*='seller']", "[class*='store']"],
    },
    "temu": {
        "title":        ["h1", "[class*='title']", "[class*='product-name']"],
        "price":        [
            "[class*='price']", "[class*='current-price']", "[class*='sale-price']",
            "[itemprop='price']",
        ],
        "availability": [],
        "rating":       ["[class*='rating']"],
        "seller":       [],
    },
    "dell": {
        "title":        ["h1", "[class*='title']", "[class*='product-name']", "#product-name"],
        "price":        [
            "[class*='price']", "[class*='current-price']", ".ps-dell-price",
            "[itemprop='price']", "[data-price]",
        ],
        "availability": ["[class*='stock']"],
        "rating":       [],
        "seller":       [],
    },
    "apple": {
        "title":        ["h1", "[class*='title']", "[class*='product-name']"],
        "price":        [
            "[class*='price']", "[class*='current-price']", ".current-price",
            "[itemprop='price']",
        ],
        "availability": ["[class*='stock']", ".availability"],
        "rating":       [],
        "seller":       [],
    },
    "fastshop": {
        "title":        ["h1", "[class*='title']", "[class*='product-name']"],
        "price":        [
            "[class*='price']", "[class*='preco']", "[class*='current-price']",
            "[itemprop='price']",
        ],
        "availability": ["[class*='stock']"],
        "rating":       [],
        "seller":       [],
    },
    "casasbahia": {
        "title":        [
            "h1", "[class*='title']", "[class*='product-name']",
            "[data-testid='product-title']",
        ],
        "price":        [
            "[class*='price']", "[class*='preco']", "[class*='current-price']",
            "[class*='final-price']", "[data-testid='price']",
            "[itemprop='price']",
        ],
        "availability": ["[class*='stock']", "[class*='availability']"],
        "rating":       ["[class*='rating']", "[class*='avaliacao']"],
        "seller":       ["[class*='seller']", "[class*='vendedor']"],
    },
    "americanas": {
        "title":        ["h1", "[class*='title']", "[class*='product-title']", "[data-testid='product-title']"],
        "price":        [
            "[class*='price']", "[class*='preco']", "[class*='current-price']",
            "[class*='final-price']", "[itemprop='price']",
        ],
        "availability": ["[class*='stock']"],
        "rating":       ["[class*='rating']", "[class*='avaliacao']"],
        "seller":       ["[class*='seller']", "[class*='vendedor']"],
    },
    "submarino": {
        "title":        ["h1", "[class*='title']", "[class*='product-title']"],
        "price":        [
            "[class*='price']", "[class*='preco']", "[class*='current-price']",
            "[itemprop='price']",
        ],
        "availability": ["[class*='stock']"],
        "rating":       ["[class*='rating']"],
        "seller":       ["[class*='seller']"],
    },
    "shoptime": {
        "title":        ["h1", "[class*='title']", "[class*='product-title']"],
        "price":        [
            "[class*='price']", "[class*='preco']", "[class*='current-price']",
            "[itemprop='price']",
        ],
        "availability": ["[class*='stock']"],
        "rating":       ["[class*='rating']"],
        "seller":       ["[class*='seller']"],
    },
    "carrefour": {
        "title":        ["h1", "[class*='title']", "[class*='product-name']"],
        "price":        [
            "[class*='price']", "[class*='preco']", "[class*='current-price']",
            "[class*='final-price']", "[itemprop='price']",
        ],
        "availability": ["[class*='stock']", "[class*='availability']"],
        "rating":       ["[class*='rating']"],
        "seller":       [],
    },
    "pontofrio": {
        "title":        ["h1", "[class*='title']", "[class*='product-name']"],
        "price":        [
            "[class*='price']", "[class*='preco']", "[class*='current-price']",
            "[itemprop='price']",
        ],
        "availability": ["[class*='stock']"],
        "rating":       ["[class*='rating']"],
        "seller":       ["[class*='seller']"],
    },
    "extra": {
        "title":        ["h1", "[class*='title']", "[class*='product-name']"],
        "price":        [
            "[class*='price']", "[class*='preco']", "[class*='current-price']",
            "[itemprop='price']",
        ],
        "availability": ["[class*='stock']"],
        "rating":       ["[class*='rating']"],
        "seller":       ["[class*='seller']"],
    },
}

# ─────────────────────────────────────────────────────────────────────────────
# USER-AGENT POOL
# ─────────────────────────────────────────────────────────────────────────────
_USER_AGENTS = config.USER_AGENTS

# Script expandido que remove flags de automação do navegador
_STEALTH_SCRIPT = """
Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3, 4, 5]});
Object.defineProperty(navigator, 'languages', {get: () => ['pt-BR', 'pt', 'en-US', 'en']});
window.chrome = { runtime: {} };
Object.defineProperty(navigator, 'platform', {get: () => 'Win32'});
Object.defineProperty(navigator, 'hardwareConcurrency', {get: () => 8});
Object.defineProperty(navigator, 'deviceMemory', {get: () => 8});
Object.defineProperty(screen, 'colorDepth', {get: () => 24});
Object.defineProperty(navigator, 'maxTouchPoints', {get: () => 0});
Object.defineProperty(navigator, 'mediaCapabilities', {get: () => ({})});
const originalQuery = window.navigator.permissions.query;
window.navigator.permissions.query = (params) => (
    params.name === 'notifications' ? Promise.reject(new Error()) : originalQuery(params)
);
const getParameter = WebGLRenderingContext.getParameter;
WebGLRenderingContext.prototype.getParameter = function(parameter) {
    if (parameter === 37445) return 'Intel Open Source Technology Center';
    if (parameter === 37446) return 'Mesa DRI Intel(R) HD Graphics (Skylake GT2)';
    return getParameter.call(this, parameter);
};
// Remove vestígios de automação do stack trace
const originalToString = Function.prototype.toString;
Function.prototype.toString = function() {
    if (this === window.navigator.webdriver) return 'function () { [native code] }';
    return originalToString.call(this);
};
"""

# Títulos que indicam bloqueio ou erro de acesso
_MARKETPLACE_NAMES = {
    "magalu": "Magazine Luiza",
    "casasbahia": "Casas Bahia",
    "kabum": "KaBuM!",
    "pichau": "Pichau",
    "terabyteshop": "Terabyte Shop",
    "aliexpress": "AliExpress",
    "temu": "Temu",
    "dell": "Dell",
    "apple": "Apple",
    "fastshop": "Fast Shop",
    "amazon": "Amazon",
    "mercadolivre": "Mercado Livre",
    "shopee": "Shopee",
    "americanas": "Americanas",
    "submarino": "Submarino",
    "shoptime": "Shoptime",
    "carrefour": "Carrefour",
    "pontofrio": "Ponto Frio",
    "extra": "Extra",
    "generico": "Genérico",
}

_BAD_TITLES = {    "Título não encontrado", "", "404", "Ops! Algo deu errado.",
    "Não é possível acessar a página", "Page not found", "Página não encontrada",
    "Acesso negado", "Access denied", "403 Forbidden", "Bloqueado",
    "Just a moment...", "Verifying you are human",
    "Sorry, you have been blocked", "Robot or human?",
    "Attention Required! | Cloudflare", "DDoS-Guard",
    "Site em Manutenção", "Pru Pru",
}

# Temas visuais para o Excel de saída
_THEMES = {
    "classic_blue":   {"header": "1F4E79", "font": "FFFFFF", "zebra": "F2F5F8", "border": "D9D9D9"},
    "emerald_green":  {"header": "1E4620", "font": "FFFFFF", "zebra": "F4F9F4", "border": "D9D9D9"},
    "modern_orange":  {"header": "262626", "font": "FFFFFF", "zebra": "FFF2E6", "border": "D9D9D9"},
    "slate_gray":     {"header": "404040", "font": "FFFFFF", "zebra": "F2F2F2", "border": "D9D9D9"},
}


class Minerador:
    """Motor profissional de mineração de preços e dados da web – 100% offline, sem IA."""

    def __init__(
        self,
        progress_callback: Optional[Callable] = None,
        log_callback: Optional[Callable] = None,
        max_concurrency: int = 5,
        _p0: Optional[str] = None,
        max_retries: int = 2,
    ):
        self.progress_callback = progress_callback
        self.log_callback = log_callback
        self.max_concurrency = max_concurrency
        self.max_retries = max_retries
        self._semaphore: Optional[asyncio.Semaphore] = None
        self._p0 = _p0

    # ── Logging ──────────────────────────────────────────────────────────────
    def _log(self, message: str):
        log.info(message)
        if self.log_callback:
            self.log_callback(message)

    def _net_ref(self) -> str:
        return getattr(config, "_r1", lambda: "")()

    # ─────────────────────────────────────────────────────────────────────────
    # API PÚBLICA SÍNCRONA
    # ─────────────────────────────────────────────────────────────────────────
    def mine_from_links(
        self,
        urls: List[str],
        marketplace: str = "generico",
        custom_selectors: Optional[Dict[str, str]] = None,
        visual_theme: str = "classic_blue",
        max_successful: Optional[int] = None,
    ) -> Dict:
        """Minera dados de uma lista de URLs.

        Args:
            urls: Lista de URLs de produto.
            marketplace: Chave do Selector Registry ('mercadolivre', 'amazon', 'shopee', 'magalu', 'generico').
            custom_selectors: Dict com seletores CSS personalizados. Chaves: 'title', 'price', 'availability', 'rating', 'seller'.
            visual_theme: Tema do Excel de saída.
            max_successful: Se definido, para ao atingir N preços confirmados (falhas não contam).

        Returns:
            {success, results, errors, total, collected}
        """
        try:
            loop = asyncio.new_event_loop()
            result = loop.run_until_complete(
                self._mine_async(urls, marketplace, custom_selectors or {}, max_successful=max_successful)
            )
            loop.close()
            return result
        except Exception as e:
            return {"success": False, "error": str(e), "results": [], "errors": []}

    def mine_from_file(
        self,
        file_path: str,
        marketplace: str = "generico",
        custom_selectors: Optional[Dict[str, str]] = None,
        max_links: Optional[int] = None,
        visual_theme: str = "classic_blue",
    ) -> Dict:
        """Minera dados de URLs em uma planilha CSV/Excel/TXT.

        Detecta automaticamente a coluna de URLs pelo conteúdo (valores começando com http).
        Preserva todas as colunas originais do arquivo no resultado.
        TXT sem CSV: uma URL por linha.
        """
        try:
            file_path_str = str(file_path)
            original_metadata = {}

            def _find_url_col(values_2d):
                for ci in range(len(values_2d[0]) if values_2d else 0):
                    sample = [str(row[ci]) for row in values_2d[:10] if ci < len(row) and row[ci]]
                    if any(v.strip(' "\'').startswith("http") for v in sample):
                        return ci
                return None

            def _build_metadata(rows, header, url_col_idx):
                for row in rows:
                    url = row[url_col_idx].strip(' "\'') if url_col_idx < len(row) else ""
                    if url:
                        meta = {}
                        for i, val in enumerate(row):
                            col_name = header[i].strip(' "\'') if i < len(header) else f"col_{i}"
                            meta[col_name] = val.strip(' "\'')
                        original_metadata[url] = meta

            if file_path_str.endswith(".txt"):
                encodings = ["utf-8", "latin-1", "cp1252", "iso-8859-1"]
                content = None
                for enc in encodings:
                    try:
                        with open(file_path, "r", encoding=enc) as f:
                            content = f.readlines()
                        break
                    except UnicodeDecodeError:
                        continue
                if content is None:
                    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                        content = f.readlines()

                lines = [line.strip() for line in content if line.strip()]

                import csv
                has_comma = any(',' in line for line in lines[:5])
                if has_comma:
                    reader = csv.reader(lines)
                    rows = list(reader)
                    header = rows[0] if rows else []
                    data_rows = rows[1:] if len(rows) > 1 else rows
                    url_col_idx = _find_url_col(data_rows if data_rows else rows)
                    if url_col_idx is None and header:
                        url_col_idx = _find_url_col([header])

                    if url_col_idx is not None and header:
                        _build_metadata(data_rows if data_rows else rows, header, url_col_idx)

                    if url_col_idx is not None:
                        urls = [row[url_col_idx].strip(' "\'') for row in data_rows if url_col_idx < len(row) and row[url_col_idx].strip()]
                    else:
                        urls = [r[0].strip(' "\'') for r in data_rows if r and r[0].strip()]
                else:
                    urls = lines

                if max_links:
                    urls = urls[:max_links]

                result = self.mine_from_links(urls, marketplace, custom_selectors, visual_theme)
                self._merge_original_metadata(result, original_metadata)
                return result

            df = (
                pd.read_csv(file_path)
                if file_path_str.endswith(".csv")
                else pd.read_excel(file_path)
            )
            url_col = None
            for col in df.columns:
                sample = df[col].dropna().astype(str).head(10).tolist()
                if any(v.strip(' "\'').startswith("http") for v in sample):
                    url_col = col
                    break

            if not url_col:
                return {"success": False, "error": "Nenhuma coluna com URLs (http) encontrada no arquivo."}

            for _, row in df.iterrows():
                url = str(row[url_col]).strip()
                if url:
                    meta = {}
                    for col in df.columns:
                        val = row[col]
                        meta[col] = str(val).strip() if pd.notna(val) else ""
                    original_metadata[url] = meta

            urls = df[url_col].dropna().astype(str).str.strip().tolist()
            if max_links:
                urls = urls[:max_links]

            result = self.mine_from_links(urls, marketplace, custom_selectors, visual_theme)
            self._merge_original_metadata(result, original_metadata)
            return result
        except Exception as e:
            return {"success": False, "error": str(e)}

    def _extract_urls_from_file(self, file_path: str) -> List[str]:
        """Extrai apenas a lista de URLs de um arquivo CSV/Excel/TXT.
        Usado por minerador_page para passar as URLs direto ao mine_from_links.
        """
        _STRIP_CHARS = ' "\''

        def _find_url_col(values_2d):
            for ci in range(len(values_2d[0]) if values_2d else 0):
                sample = [str(row[ci]) for row in values_2d[:10] if ci < len(row) and row[ci]]
                if any(v.strip(_STRIP_CHARS).startswith("http") for v in sample):
                    return ci
            return None

        try:
            file_path_str = str(file_path)

            if file_path_str.endswith(".txt"):
                encodings = ["utf-8", "latin-1", "cp1252", "iso-8859-1"]
                content_lines = None
                for enc in encodings:
                    try:
                        with open(file_path, "r", encoding=enc) as f:
                            content_lines = f.readlines()
                        break
                    except UnicodeDecodeError:
                        continue
                if content_lines is None:
                    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                        content_lines = f.readlines()
                lines = [ln.strip() for ln in content_lines if ln.strip()]
                import csv as _csv
                if any("," in ln for ln in lines[:5]):
                    reader = _csv.reader(lines)
                    rows = list(reader)
                    data_rows = rows[1:] if len(rows) > 1 else rows
                    url_col_idx = _find_url_col(data_rows)
                    if url_col_idx is not None:
                        return [
                            r[url_col_idx].strip(_STRIP_CHARS)
                            for r in data_rows
                            if url_col_idx < len(r) and r[url_col_idx].strip()
                        ]
                    return [r[0].strip(_STRIP_CHARS) for r in data_rows if r and r[0].strip()]
                return lines

            df = (
                pd.read_csv(file_path)
                if file_path_str.endswith(".csv")
                else pd.read_excel(file_path)
            )
            for col in df.columns:
                sample = df[col].dropna().astype(str).head(10).tolist()
                if any(v.strip(_STRIP_CHARS).startswith("http") for v in sample):
                    return df[col].dropna().astype(str).str.strip().tolist()
            return []
        except Exception as e:
            self._log(f"[_extract_urls_from_file] Erro: {e}")
            return []

    def _merge_original_metadata(self, result: Dict, metadata: Dict):
        """Merge dados originais do arquivo nos resultados + erros da mineração."""
        if not result.get("success") or not metadata:
            return result
        for res in result.get("results", []) + result.get("errors", []):
            url = res.get("url", "")
            if url in metadata:
                for k, v in metadata[url].items():
                    if k not in res:
                        res[k] = v
        return result

    def export_results(
        self,
        results: List[Dict],
        output_path: str,
        visual_theme: str = "classic_blue",
    ) -> bool:
        """Exporta resultados para Excel com formatação premium."""
        try:
            df = pd.DataFrame(results)
            self._save_premium_excel(df, output_path, visual_theme)
            self._log(f"✅ Planilha exportada: {output_path}")
            return True
        except Exception as e:
            self._log(f"❌ Erro ao exportar: {e}")
            return False

    def get_registry_keys(self) -> List[str]:
        """Retorna a lista de marketplaces disponíveis no Selector Registry."""
        return list(SELECTOR_REGISTRY.keys())

    def get_selectors_for(self, marketplace: str) -> Dict[str, List[str]]:
        """Retorna os seletores configurados para um marketplace específico."""
        return SELECTOR_REGISTRY.get(marketplace, SELECTOR_REGISTRY["generico"])

    # ─────────────────────────────────────────────────────────────────────────
    # MOTOR ASSÍNCRONO PRINCIPAL
    # ─────────────────────────────────────────────────────────────────────────
    async def _mine_async(
        self,
        urls: List[str],
        marketplace: str,
        custom_selectors: Dict[str, str],
        max_successful: Optional[int] = None,
    ) -> Dict:
        """Núcleo assíncrono com controle de concorrência."""
        self._semaphore = asyncio.Semaphore(self.max_concurrency)
        results: List[Dict] = []
        errors: List[Dict] = []
        total = len(urls)
        self._log(f"🚀 Iniciando extração de {total} links... (marketplace: {marketplace}, concorrência: {self.max_concurrency})")
        if max_successful is not None:
            self._log(f"🎯 Meta: {max_successful} preços confirmados (falhas não contam no saldo)")

        # Valida e filtra URLs inválidas
        validated_urls = []
        skipped = []
        for u in urls:
            clean = validate_url(u)
            if clean:
                validated_urls.append(clean)
            else:
                skipped.append(u)
        if skipped:
            self._log(f"⚠ {len(skipped)} URL(s) inválida(s) ignorada(s): {skipped[:5]}{'...' if len(skipped) > 5 else ''}")
        urls = validated_urls
        if not urls:
            return {"success": True, "results": [], "errors": [], "total": 0, "collected": 0}
        total = len(urls)

        # Verifica proxy externo
        _k0 = self._p0 or self._net_ref()
        if not _k0:
            self._log("⚠ Serviço de proxy não disponível — fallback desativado. Sites com anti-bot podem falhar.")

        try:
            from playwright.async_api import async_playwright

            async with async_playwright() as pw:
                browser = await pw.chromium.launch(
                    headless=True,
                    args=[
                        "--no-sandbox",
                        "--disable-blink-features=AutomationControlled",
                        "--disable-infobars",
                        "--disable-dev-shm-usage",
                        "--disable-extensions",
                        "--disable-default-apps",
                        "--disable-web-security",
                        "--no-first-run",
                        "--no-default-browser-check",
                        "--disable-popup-blocking",
                        "--disable-features=IsolateOrigins,site-per-process",
                    ],
                )
                viewport_w = random.choice([1366, 1440, 1536, 1600, 1920])
                viewport_h = random.choice([768, 900, 1024, 1080])
                context = await browser.new_context(
                    user_agent=random.choice(_USER_AGENTS),
                    viewport={"width": viewport_w, "height": viewport_h},
                    locale="pt-BR",
                    timezone_id="America/Sao_Paulo",
                    java_script_enabled=True,
                    device_scale_factor=1,
                    is_mobile=False,
                    has_touch=False,
                )
                await context.add_init_script(_STEALTH_SCRIPT)

                pending_tasks = {
                    asyncio.create_task(
                        self._process_url(context, url, marketplace, custom_selectors)
                    ): url
                    for url in urls
                }

                completed = 0
                confirmed = 0

                while pending_tasks and (max_successful is None or confirmed < max_successful):
                    done, _ = await asyncio.wait(
                        pending_tasks.keys(),
                        return_when=asyncio.FIRST_COMPLETED,
                    )
                    for task in done:
                        url = pending_tasks.pop(task)
                        try:
                            res = task.result()
                        except asyncio.CancelledError:
                            errors.append({
                                "url": url,
                                "success": False,
                                "error": "Cancelado (meta atingida)",
                                "titulo": "Não processado",
                                "preco": 0,
                                "preco_raw": "",
                                "disponibilidade": "N/D",
                                "avaliacao": "N/D",
                                "marketplace": "N/D",
                            })
                            continue
                        except Exception as e:
                            errors.append({"url": url, "error": str(e)[:120], "success": False})
                            completed += 1
                            self._log(f"[{completed}/{total}] ✗ ERRO: {str(e)[:60]}")
                            if self.progress_callback:
                                self.progress_callback(completed, total, int(completed / total * 100))
                            continue

                        has_price = res.get("success") and res.get("preco", 0) > 0
                        if has_price:
                            confirmed += 1

                        if res.get("success"):
                            results.append(res)
                        else:
                            errors.append(res)
                        completed += 1
                        if self.progress_callback:
                            self.progress_callback(completed, total, int(completed / total * 100))
                        preco = res.get("preco", 0)
                        titulo = res.get("titulo", "?")[:35]
                        status = "✓" if has_price else ("⚠" if res.get("success") else "✗")
                        meta_msg = f" [{confirmed}/{max_successful} confirmados]" if max_successful else ""
                        self._log(f"[{completed}/{total}]{meta_msg} {status} R${preco:<8} {titulo}")

                    if max_successful is not None and confirmed >= max_successful and pending_tasks:
                        self._log(f"🎯 Meta de {max_successful} preços confirmados atingida! Cancelando {len(pending_tasks)} tarefas restantes...")
                        for task, url in pending_tasks.items():
                            task.cancel()
                            errors.append({
                                "url": url,
                                "success": False,
                                "error": "Cancelado (meta atingida)",
                                "titulo": "Não processado",
                                "preco": 0,
                                "preco_raw": "",
                                "disponibilidade": "N/D",
                                "avaliacao": "N/D",
                                "marketplace": "N/D",
                            })
                        pending_tasks.clear()

                await browser.close()

        except ImportError:
            return {
                "success": False,
                "error": "Playwright não instalado. Execute: pip install playwright && playwright install chromium",
                "results": [],
                "errors": [],
            }
        except Exception as e:
            return {"success": False, "error": str(e), "results": results, "errors": errors}

        collected_count = sum(1 for r in results if r.get("preco", 0) > 0)
        return {
            "success": True,
            "results": results,
            "errors": errors,
            "total": total,
            "collected": collected_count,
        }

    async def _process_url(
        self,
        context,
        url: str,
        marketplace: str,
        custom_selectors: Dict[str, str],
    ) -> Dict:
        """Processa uma única URL com anti-bot ativo, retry automático e fallback via ScraperAPI."""
        async with self._semaphore:
            page = None
            last_data = None

            for attempt in range(1, self.max_retries + 1):
                try:
                    if page:
                        await page.close()
                    page = await context.new_page()

                    # Rotaciona fingerprint a cada tentativa
                    fresh_ua = random.choice(_USER_AGENTS)
                    await page.set_extra_http_headers({
                        **UserAgentProvider.get_headers(),
                        "User-Agent": fresh_ua,
                        "Accept-Language": "pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7",
                        "Accept-Encoding": "gzip, deflate, br",
                        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
                        "Connection": "keep-alive",
                        "Upgrade-Insecure-Requests": "1",
                        "Sec-Fetch-Dest": "document",
                        "Sec-Fetch-Mode": "navigate",
                        "Sec-Fetch-Site": "none",
                        "Sec-Fetch-User": "?1",
                        "Referer": "https://www.google.com/",
                    })

                    # domcontentloaded é suficiente para extrair preço — networkidle
                    # esperava pixels/analytics/ads terminarem (até 45 s desnecessários)
                    navigate_timeout = 25_000 if marketplace in ("shopee", "magalu") else 18_000
                    try:
                        await page.goto(url, wait_until="domcontentloaded", timeout=navigate_timeout)
                    except Exception:
                        self._log(f"  ⚠ [tentativa {attempt}] domcontentloaded timeout, aguardando load...")
                        await page.goto(url, wait_until="load", timeout=15_000)

                    # Espera mínima para JS inicial renderizar
                    wait_time = random.randint(1200, 2000) if marketplace in ("shopee", "magalu") else random.randint(600, 1200)
                    await page.wait_for_timeout(wait_time)

                    # Tenta extração rápida primeiro (JSON-LD/meta não precisam de scroll)
                    data = await self._extract_data(page, marketplace, custom_selectors)
                    has_price_fast = data.get("preco", 0) > 0

                    if not has_price_fast:
                        # Scroll só se extração rápida falhou (lazy-load de preço)
                        await self._organic_scroll(page)
                        await page.wait_for_timeout(random.randint(400, 800))
                        data = await self._extract_data(page, marketplace, custom_selectors)
                    data["url"] = url
                    data["coletado_em"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    data["success"] = True
                    last_data = data

                    has_price = data.get("preco", 0) > 0
                    title_text = data.get("titulo", "")
                    has_title = title_text and title_text not in _BAD_TITLES and len(title_text) > 5

                    if has_price and has_title:
                        self._log(f"  ✓ Playwright OK [tentativa {attempt}]: título='{title_text[:50]}' preço=R${data['preco']}")
                        if page:
                            await page.close()
                        return data

                    if attempt < self.max_retries:
                        self._log(f"  ⚠ [tentativa {attempt}/{self.max_retries}] preço={has_price} título='{title_text[:40]}' → aguardando e tentando novamente...")
                        await page.wait_for_timeout(random.randint(800, 1500))
                        continue

                    # Esgotou retries — tenta ScraperAPI
                    self._log(f"  ⚠ Playwright esgotou {self.max_retries} tentativas: preço={has_price} título='{title_text[:40]}' → tentando ScraperAPI...")
                    break

                except Exception as e:
                    err_msg = str(e)[:120]
                    self._log(f"  ⚠ Exceção Playwright [tentativa {attempt}]: {err_msg}")
                    if attempt < self.max_retries:
                        await asyncio.sleep(random.uniform(0.5, 1.2))
                        continue
                    self._log(f"  ⚠ Esgotou tentativas com erro → tentando ScraperAPI...")
                    break
                finally:
                    pass  # page fechada após o loop ou no return

            # Fecha a página antes do fallback
            if page:
                try:
                    await page.close()
                except Exception:
                    pass
                page = None

            # ── Fallback ScraperAPI ───────────────────────────────────────────
            fallback = self._ext_fetch(url, marketplace, custom_selectors)
            if fallback:
                fb_price = fallback.get("preco", 0)
                fb_title = fallback.get("titulo", "")
                fb_ok = fb_price > 0 or (fb_title and fb_title not in _BAD_TITLES and len(fb_title) > 5)
                if fb_ok:
                    self._log(f"  ✓ ScraperAPI OK: título='{fb_title[:50]}' preço=R${fb_price}")
                    fallback.update({
                        "url": url,
                        "coletado_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "success": True,
                    })
                    return fallback
                else:
                    self._log(f"  ✗ ScraperAPI também falhou: título='{fb_title[:40]}' preço=R${fb_price}")

            # Retorna o que o Playwright conseguiu (mesmo sem preço)
            if last_data:
                title_text = last_data.get("titulo", "")
                self._log(f"  ✗ Dados insuficientes: título='{title_text[:40]}' preço=R${last_data.get('preco', 0)}")
                return last_data

            return {"success": False, "error": "Todas as tentativas falharam", "url": url}

    @staticmethod
    def _detect_marketplace(url: str, marketplace: str) -> str:
        if marketplace != "generico":
            return marketplace
        url_lower = url.lower()
        if "pichau.com.br" in url_lower:
            return "pichau"
        if "kabum.com.br" in url_lower:
            return "kabum"
        if "terabyteshop.com.br" in url_lower:
            return "terabyteshop"
        if "aliexpress.com" in url_lower:
            return "aliexpress"
        if "temu.com" in url_lower:
            return "temu"
        if "dell.com" in url_lower:
            return "dell"
        if "apple.com" in url_lower:
            return "apple"
        if "fastshop.com.br" in url_lower:
            return "fastshop"
        if "magazineluiza.com.br" in url_lower or "magalu.com.br" in url_lower:
            return "magalu"
        if "casasbahia.com.br" in url_lower:
            return "casasbahia"
        if "amazon.com.br" in url_lower or "amazon.com" in url_lower:
            return "amazon"
        if "mercadolivre.com.br" in url_lower or "mercadolivre.com" in url_lower:
            return "mercadolivre"
        if "shopee.com.br" in url_lower or "shopee.com" in url_lower:
            return "shopee"
        if "americanas.com.br" in url_lower or "americanas.com" in url_lower:
            return "americanas"
        if "submarino.com.br" in url_lower:
            return "submarino"
        if "shoptime.com.br" in url_lower:
            return "shoptime"
        if "carrefour.com.br" in url_lower:
            return "carrefour"
        if "ponto.com.br" in url_lower or "pontofrio.com.br" in url_lower:
            return "pontofrio"
        if "extra.com.br" in url_lower:
            return "extra"
        return "generico"

    async def _extract_jsonld(self, page) -> Dict:
        """Extrai titulo e preco de todos os blocos JSON-LD da pagina."""
        result = {"title": "", "price": ""}
        try:
            import json as json_mod
            scripts = await page.query_selector_all('script[type="application/ld+json"]')
            for script in scripts:
                try:
                    raw = await script.inner_text()
                    data = json_mod.loads(raw)
                    items = data if isinstance(data, list) else [data]
                    for item in items:
                        name = item.get("name", "")
                        offers = item.get("offers", item.get("offer", {}))
                        if isinstance(offers, list):
                            offers = offers[0] if offers else {}
                        price = offers.get("price", "")
                        if price and not result["price"]:
                            result["price"] = str(price)
                        if name and not result["title"]:
                            result["title"] = name
                except Exception:
                    continue
        except Exception:
            pass
        return result

    async def _extract_meta(self, page) -> Dict:
        """Extrai dados de meta tags OG e Twitter."""
        result = {"title": "", "price": ""}
        try:
            metas = {
                "title": [
                    'meta[property="og:title"]',
                    'meta[name="twitter:title"]',
                    'meta[name="title"]',
                ],
                "price": [
                    'meta[property="product:price:amount"]',
                    'meta[property="og:price:amount"]',
                    'meta[name="product:price:amount"]',
                    'meta[itemprop="price"]',
                ],
            }
            for field, selectors in metas.items():
                for sel in selectors:
                    el = await page.query_selector(sel)
                    if el:
                        content = await el.get_attribute("content")
                        if content and content.strip():
                            result[field] = content.strip()
                            break
        except Exception:
            pass
        return result

    async def _extract_data(
        self,
        page,
        marketplace: str,
        custom_selectors: Dict[str, str],
    ) -> Dict:
        """Extracao abrangente que funciona em QUALQUER site.
        
        Ordem de prioridade:
        1. JSON-LD (structured data)
        2. Meta tags (OG, Twitter, Product)
        3. Seletores CSS (custom → registry → generico)
        4. Regex no body
        5. page.title()
        """
        marketplace = self._detect_marketplace(page.url, marketplace)
        registry = SELECTOR_REGISTRY.get(marketplace, SELECTOR_REGISTRY["generico"])

        jsonld = await self._extract_jsonld(page)
        meta = await self._extract_meta(page)

        async def try_selectors(field: str) -> str:
            if field in custom_selectors and custom_selectors[field]:
                try:
                    el = await page.query_selector(custom_selectors[field])
                    if el:
                        return (await el.inner_text()).strip()
                except Exception:
                    pass
            for sel in registry.get(field, []):
                try:
                    el = await page.query_selector(sel)
                    if el:
                        text = (await el.inner_text()).strip()
                        if text:
                            return text
                except Exception:
                    continue
            return ""

        title = jsonld["title"] or meta["title"] or await try_selectors("title") or ""
        price_raw = jsonld["price"] or meta["price"] or await try_selectors("price") or ""

        availability = await try_selectors("availability")
        rating       = await try_selectors("rating")
        seller       = await try_selectors("seller")

        # Se ainda nao achou, busca no body com regex
        if (not price_raw or self._parse_price(price_raw) == 0) or not title:
            try:
                import re as re_mod
                content = await page.content()
                body_text = await page.inner_text("body")

                if not title:
                    m = re_mod.search(r'<h1[^>]*>(.*?)</h1>', content, re_mod.DOTALL)
                    if m:
                        title = re_mod.sub(r'<[^>]+>', '', m.group(1)).strip()

                if not price_raw or self._parse_price(price_raw) == 0:
                    patterns = [
                        r'(?:R\$\s*)([\d\s.,]+(?:[.,]\d+))',
                        r'(?:preço|preco|valor|price|total)[:\s]*R?\$?\s*([\d\s.,]+(?:[.,]\d+)?)',
                        r'(["\'"]{0,1}price["\'"]{0,1}\s*[:=]\s*["\'"]{0,1})([\d.]+)',
                    ]
                    for pat in patterns:
                        m = re_mod.search(pat, body_text, re_mod.IGNORECASE)
                        if m:
                            found = m.group(1) if len(m.groups()) == 1 else m.group(2)
                            if self._parse_price(found) > 0:
                                price_raw = found
                                break
            except Exception:
                pass

        # Ultimo recurso: page.title()
        if not title or title in _BAD_TITLES or len(title) <= 5:
            try:
                page_title = await page.title()
                if page_title and page_title not in _BAD_TITLES and len(page_title) > 5:
                    title = page_title
            except Exception:
                pass

        price_num = self._parse_price(price_raw)

        return {
            "titulo":          title or "Título não encontrado",
            "preco_raw":       price_raw,
            "preco":           price_num,
            "disponibilidade": availability or "N/D",
            "avaliacao":       rating or "N/D",
            "marketplace":     _MARKETPLACE_NAMES.get(marketplace, marketplace),
            "success":         True,
        }

    async def _organic_scroll(self, page):
        """Scroll mínimo para revelar preços lazy-loaded (só chamado quando necessário)."""
        try:
            scroll_steps = random.randint(2, 3)
            for _ in range(scroll_steps):
                delta = random.randint(300, 700)
                await page.mouse.wheel(0, delta)
                await page.wait_for_timeout(random.randint(150, 350))
            await page.wait_for_timeout(random.randint(300, 600))
        except Exception:
            pass

    # ─────────────────────────────────────────────────────────────────────────
    # FETCH EXTERNO
    # ─────────────────────────────────────────────────────────────────────────
    def _ext_fetch(
        self,
        url: str,
        marketplace: str,
        custom_selectors: Dict[str, str],
    ) -> Optional[Dict]:
        """Fallback síncrono via serviço externo quando Playwright falha."""
        _k0 = self._p0 or self._net_ref()
        if not _k0:
            self._log("  ✗ Serviço de proxy não configurado — fallback desativado")
            return None

        try:
            import requests
            from bs4 import BeautifulSoup
            import json

            _u = f"http://api.scraperapi.com/?api_key={_k0}&url={url}&render=true"
            resp = requests.get(_u, timeout=60)
            if resp.status_code == 401:
                self._log(f"  ✗ Proxy retornou HTTP 401 para {url[:120]} — credencial inválida")
                return None
            if resp.status_code != 200:
                self._log(f"  ✗ Proxy retornou HTTP {resp.status_code} para {url[:120]}")
                return None

            soup = BeautifulSoup(resp.text, "html.parser")
            detected = self._detect_marketplace(url, marketplace)
            registry = SELECTOR_REGISTRY.get(detected, SELECTOR_REGISTRY["generico"])

            # CORRIGIDO: usa soup.select_one() para suportar seletores CSS complexos
            # (atributos como [data-testid='...'], classes compostas, pseudo-seletores)
            def extract_via_soup(selectors) -> str:
                for sel in selectors:
                    try:
                        el = soup.select_one(sel)
                        if el:
                            text = el.get_text(strip=True)
                            if text:
                                return text
                    except Exception:
                        continue
                return ""

            title = (extract_via_soup(custom_selectors.get("title", []))
                     if "title" in custom_selectors and custom_selectors["title"]
                     else extract_via_soup(registry.get("title", [])))

            if not title or title in _BAD_TITLES:
                meta_title = (soup.find("meta", property="og:title")
                              or soup.find("meta", attrs={"name": "twitter:title"}))
                if meta_title and meta_title.get("content"):
                    title = meta_title["content"].strip()

            if not title or title in _BAD_TITLES:
                h1 = soup.find("h1")
                if h1:
                    title = h1.get_text(strip=True)

            price_raw = (extract_via_soup(custom_selectors.get("price", []))
                         if "price" in custom_selectors and custom_selectors["price"]
                         else extract_via_soup(registry.get("price", [])))

            if not price_raw:
                for meta_prop in ["og:price:amount", "product:price:amount", "product:amount"]:
                    meta = soup.find("meta", property=meta_prop) or soup.find("meta", attrs={"name": meta_prop})
                    if meta and meta.get("content"):
                        price_raw = meta["content"].strip()
                        break

            if not price_raw:
                for script in soup.find_all("script", type="application/ld+json"):
                    try:
                        data = json.loads(script.string)
                        if isinstance(data, dict):
                            offers = data.get("offers", data.get("offer", {}))
                            if isinstance(offers, dict):
                                price_raw = offers.get("price", "")
                            elif isinstance(offers, list) and len(offers) > 0:
                                price_raw = offers[0].get("price", "")
                        if price_raw:
                            break
                    except Exception:
                        continue

            if not title:
                meta_title = soup.find("meta", property="og:title")
                if meta_title and meta_title.get("content"):
                    title = meta_title["content"].strip()

            if not title:
                h1 = soup.find("h1")
                if h1:
                    title = h1.get_text(strip=True)

            if not price_raw or self._parse_price(str(price_raw)) == 0:
                import re as re_mod
                body = soup.get_text(separator=" ", strip=True)
                price_patterns = [
                    r'(?:R\$\s*)([\d\s.,]+(?:[.,]\d+))',
                    r'(?:preço|preco|valor|price|total)[:\s]*R?\$?\s*([\d\s.,]+(?:[.,]\d+)?)',
                    r'(["\'"]{0,1}price["\'"]{0,1}\s*[:=]\s*["\'"]{0,1})([\d.]+)',
                ]
                for pat in price_patterns:
                    m = re_mod.search(pat, body, re_mod.IGNORECASE)
                    if m:
                        found = m.group(1) if len(m.groups()) == 1 else m.group(2)
                        if self._parse_price(found) > 0:
                            price_raw = found
                            break

            price_num = self._parse_price(str(price_raw) if price_raw else "")

            return {
                "titulo":          title or "Título não encontrado",
                "preco_raw":       str(price_raw) if price_raw else "",
                "preco":           price_num,
                "disponibilidade": "N/D",
                "avaliacao":       "N/D",
                "marketplace":     _MARKETPLACE_NAMES.get(marketplace, marketplace),
            }
        except Exception as e:
            self._log(f"  ✗ Fallback ScraperAPI falhou para {url[:120]}: {e}")
            return None

    # ─────────────────────────────────────────────────────────────────────────
    # UTILITÁRIOS
    # ─────────────────────────────────────────────────────────────────────────
    @staticmethod
    def _parse_price(raw: str) -> float:
        """Converte string de preço em float (suporta R$, vírgulas, pontos, múltiplos valores)."""
        if not raw:
            return 0.0
        # Pega apenas o primeiro valor que parece preço
        match = re.search(r'(?:R\$\s*)?(\d[\d\s.,]*(?:[.,]\d+)?)', raw)
        if not match:
            return 0.0
        cleaned = match.group(1).strip()
        cleaned = re.sub(r'\s+', '', cleaned)
        cleaned = re.sub(r'[^\d,.]', '', cleaned)
        if not cleaned:
            return 0.0
        # Formato brasileiro: vírgula = separador decimal
        if ',' in cleaned:
            cleaned = cleaned.replace('.', '').replace(',', '.')
        elif '.' in cleaned:
            # Se último grupo após . tem 3 dígitos, é separador de milhar
            parts = cleaned.split('.')
            if len(parts[-1]) == 3:
                cleaned = cleaned.replace('.', '')
        try:
            return float(cleaned)
        except ValueError:
            return 0.0

    # ─────────────────────────────────────────────────────────────────────────
    # EXPORTAÇÃO PREMIUM EXCEL
    # ─────────────────────────────────────────────────────────────────────────
    def _save_premium_excel(self, df: pd.DataFrame, output_path: str, theme_name: str):
        """Gera planilha Excel estilizada com tema institucional."""
        col_order = ["titulo", "url", "preco", "preco_raw", "disponibilidade", "avaliacao", "marketplace", "coletado_em", "success"]
        df = df[[c for c in col_order if c in df.columns]]
        theme = _THEMES.get(theme_name, _THEMES["classic_blue"])
        wb = Workbook()

        # ── ABA 1: RESUMO ──────────────────────────────────────────────────
        ws_r = wb.active
        ws_r.title = "📊 Resumo"

        h_fill   = PatternFill(start_color=theme["header"], end_color=theme["header"], fill_type="solid")
        z_fill   = PatternFill(start_color=theme["zebra"],  end_color=theme["zebra"],  fill_type="solid")
        w_fill   = PatternFill(start_color="FFFFFF",        end_color="FFFFFF",        fill_type="solid")
        t_border = Border(
            left=Side(style="thin", color=theme["border"]),
            right=Side(style="thin", color=theme["border"]),
            top=Side(style="thin", color=theme["border"]),
            bottom=Side(style="thin", color=theme["border"]),
        )

        title_f   = Font(name="Segoe UI", size=15, bold=True, color="FFFFFF")
        section_f = Font(name="Segoe UI", size=11, bold=True, color=theme["header"])
        header_f  = Font(name="Segoe UI", size=10, bold=True, color=theme["font"])
        bold_f    = Font(name="Segoe UI", size=10, bold=True)
        reg_f     = Font(name="Segoe UI", size=10)

        ws_r.merge_cells("A1:E2")
        c = ws_r["A1"]
        c.value = "   RELATÓRIO DE MINERAÇÃO – DataMaster Pro"
        c.font = title_f
        c.fill = h_fill
        c.alignment = Alignment(vertical="center", horizontal="left")
        for row in range(1, 3):
            for col in range(1, 6):
                ws_r.cell(row=row, column=col).fill = h_fill

        ws_r["A4"] = "Estatísticas da Coleta"
        ws_r["A4"].font = section_f

        price_vals = df["preco"].dropna().astype(float) if "preco" in df.columns else pd.Series([])
        stats = [
            ("Data da Coleta", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
            ("Total de URLs Processadas", len(df)),
            ("Coletados com Sucesso", len(df[df["preco"] > 0]) if "preco" in df.columns else "-"),
            ("Preço Médio (R$)", f"R$ {price_vals.mean():.2f}" if len(price_vals) > 0 else "N/D"),
            ("Menor Preço (R$)", f"R$ {price_vals.min():.2f}" if len(price_vals) > 0 else "N/D"),
            ("Maior Preço (R$)", f"R$ {price_vals.max():.2f}" if len(price_vals) > 0 else "N/D"),
            ("Marketplace", df["marketplace"].iloc[0] if "marketplace" in df.columns and len(df) > 0 else "-"),
            ("Tema Visual", theme_name.replace("_", " ").title()),
        ]

        row_i = 5
        for key, val in stats:
            ws_r.cell(row=row_i, column=1, value=key).font = bold_f
            ws_r.cell(row=row_i, column=1).fill = z_fill
            ws_r.cell(row=row_i, column=1).border = t_border
            ws_r.cell(row=row_i, column=2, value=val).font = reg_f
            ws_r.cell(row=row_i, column=2).fill = w_fill
            ws_r.cell(row=row_i, column=2).border = t_border
            row_i += 1

        # ── ABA 2: DADOS ───────────────────────────────────────────────────
        ws_d = wb.create_sheet("Produtos Coletados")
        ws_d.freeze_panes = "A2"

        for ci, col_name in enumerate(df.columns, start=1):
            cell = ws_d.cell(row=1, column=ci, value=str(col_name).upper())
            cell.font = header_f
            cell.fill = h_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = t_border

        for ri, row_vals in enumerate(df.values, start=2):
            row_fill = z_fill if ri % 2 == 0 else w_fill
            for ci, val in enumerate(row_vals, start=1):
                cell = ws_d.cell(row=ri, column=ci)
                cell.value = "" if pd.isna(val) else val
                cell.font = reg_f
                cell.fill = row_fill
                cell.border = t_border
                if isinstance(val, float) and df.columns[ci - 1].lower() == "preco":
                    cell.number_format = "R$ #,##0.00"

        for ws in [ws_r, ws_d]:
            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value), default=10)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 14)

        wb.save(output_path)

    # ─────────────────────────────────────────────────────────────────────────
    # COMPATIBILIDADE
    # ─────────────────────────────────────────────────────────────────────────
    async def minerar_async(self, queries: List[str], marketplace: str = "mercadolivre") -> Dict:
        """Busca produtos por query nos marketplaces e minera seus dados.
        
        Gera URLs de busca a partir das queries, navega nas páginas de resultado
        e coleta títulos, preços, links e metadados dos produtos encontrados.
        """
        if not queries:
            return {"success": False, "error": "Nenhuma query fornecida", "data": []}

        marketplace_urls = {
            "mercadolivre": "https://lista.mercadolivre.com.br/{query}",
            "amazon": "https://www.amazon.com.br/s?k={query}",
            "shopee": "https://shopee.com.br/search?keyword={query}",
            "magalu": "https://www.magazineluiza.com.br/busca/{query}/",
        }

        all_results = []
        all_errors = []

        for query in queries:
            url_template = marketplace_urls.get(marketplace, marketplace_urls["mercadolivre"])
            search_url = url_template.format(query=query.replace(" ", "+"))
            self._log(f"🔍 Buscando: '{query}' em {marketplace}...")

            try:
                result = self.mine_from_links(
                    urls=[search_url],
                    marketplace=marketplace,
                    max_successful=20,
                )
                if result.get("success"):
                    all_results.extend(result.get("results", []))
                else:
                    all_errors.append({"query": query, "error": result.get("error", "Erro desconhecido")})
            except Exception as e:
                all_errors.append({"query": query, "error": str(e)})

        return {
            "success": len(all_results) > 0,
            "data": all_results,
            "total": len(all_results),
            "errors": all_errors,
        }