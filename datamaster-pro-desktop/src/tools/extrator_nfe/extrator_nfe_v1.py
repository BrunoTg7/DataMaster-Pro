"""
Extrator NF-e / XML + Conciliador de Pedidos v1.0

Fluxo:
  1. Varre uma pasta de XMLs (padrão SEFAZ) ou um único arquivo XML
  2. Extrai: número da NF-e, valor (vNF), CPF/CNPJ destinatário, data emissão, infAdic
  3. Cruza com planilha de vendas do marketplace (ML, Shopee, etc.)
     - Chave primária: Número do Pedido (via infAdic ou coluna dedicada)
     - Chave de fallback: CPF/CNPJ do destinatário
  4. Gera coluna Status: ✅ OK | ⚠️ VALOR DIVERGENTE | ❌ NOTA FALTANDO
  5. Exporta Excel premium com 3 abas: Conciliado, Divergências, Notas Faltando
"""

import logging
import pandas as pd
import xml.etree.ElementTree as ET
import re
import os
from typing import Dict, List, Optional, Callable
from pathlib import Path
from datetime import datetime

log = logging.getLogger(__name__)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


SEFAZ_NS = "http://www.portalfiscal.inf.br/nfe"


class ExtratorNFe:
    """
    Motor de extração e cruzamento de NF-e com planilhas de marketplace.
    Tolerância padrão de ±R$ 0,01 para divergência de valores.
    """

    def __init__(self, log_callback: Callable = None, progress_callback: Callable = None):
        self.log_callback = log_callback
        self.progress_callback = progress_callback

    def _log(self, msg: str):
        log.info(msg)
        if self.log_callback:
            self.log_callback(msg)

    def _progress(self, pct: int):
        if self.progress_callback:
            self.progress_callback(pct)

    # ──────────────────────────────────────────────────────────────────
    # PARSING DE XML
    # ──────────────────────────────────────────────────────────────────
    def _parse_xml(self, xml_path: str) -> Optional[Dict]:
        """Extrai campos-chave de um XML de NF-e (padrão SEFAZ 4.00)."""
        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()
            ns = {"nfe": SEFAZ_NS}

            def find(node, path):
                """Tenta com namespace e sem namespace."""
                result = node.find(path, ns)
                if result is None:
                    result = node.find(path.replace("nfe:", ""))
                return result

            infNFe = root.find(".//nfe:infNFe", ns) or root.find(".//infNFe")
            if infNFe is None:
                return None

            ide = find(infNFe, "nfe:ide")
            dest = find(infNFe, "nfe:dest")
            total = find(infNFe, ".//nfe:ICMSTot") or find(infNFe, "nfe:total/nfe:ICMSTot")
            infAdic = find(infNFe, "nfe:infAdic")

            # Número NF-e
            n_nf = ""
            if ide is not None:
                el = find(ide, "nfe:nNF")
                n_nf = el.text.strip() if el is not None and el.text else ""

            # Série
            serie = ""
            if ide is not None:
                el = find(ide, "nfe:serie")
                serie = el.text.strip() if el is not None and el.text else ""

            # Data de emissão
            data_emissao = ""
            if ide is not None:
                el = find(ide, "nfe:dhEmi") or find(ide, "nfe:dEmi")
                if el is not None and el.text:
                    data_emissao = el.text[:10]  # YYYY-MM-DD

            # Valor total da NF-e
            valor_nf = 0.0
            if total is not None:
                el = find(total, "nfe:vNF")
                if el is not None and el.text:
                    try:
                        valor_nf = float(el.text)
                    except ValueError:
                        valor_nf = 0.0

            # CPF / CNPJ destinatário
            cpf_cnpj = ""
            nome_dest = ""
            if dest is not None:
                el_cpf = find(dest, "nfe:CPF")
                el_cnpj = find(dest, "nfe:CNPJ")
                el_nome = find(dest, "nfe:xNome")
                if el_cpf is not None and el_cpf.text:
                    cpf_cnpj = re.sub(r"\D", "", el_cpf.text)
                elif el_cnpj is not None and el_cnpj.text:
                    cpf_cnpj = re.sub(r"\D", "", el_cnpj.text)
                if el_nome is not None and el_nome.text:
                    nome_dest = el_nome.text.strip()

            # Informações adicionais (onde costuma estar o número do pedido)
            inf_adic_texto = ""
            if infAdic is not None:
                el = find(infAdic, "nfe:infCpl")
                if el is not None and el.text:
                    inf_adic_texto = el.text.strip()

            # Tenta extrair número do pedido do infAdic (padrão ML/Shopee)
            numero_pedido_xml = self._extrair_numero_pedido(inf_adic_texto)

            return {
                "nfe_numero": n_nf,
                "nfe_serie": serie,
                "data_emissao": data_emissao,
                "valor_nf": valor_nf,
                "cpf_cnpj": cpf_cnpj,
                "nome_dest": nome_dest,
                "inf_adic": inf_adic_texto,
                "numero_pedido_xml": numero_pedido_xml,
                "arquivo_xml": os.path.basename(xml_path),
            }
        except Exception as e:
            self._log(f"⚠️ Erro ao processar {os.path.basename(xml_path)}: {e}")
            return None

    def _extrair_numero_pedido(self, texto: str) -> str:
        """
        Tenta extrair o número do pedido de texto livre do infAdic.
        Padrões comuns: 'Pedido: 12345', '#12345', 'Pedido #12345', etc.
        """
        if not texto:
            return ""
        patterns = [
            r"pedido[:\s#]+([A-Z0-9\-]+)",
            r"order[:\s#]+([A-Z0-9\-]+)",
            r"#([A-Z0-9\-]{5,})",
            r"\b([0-9]{8,})\b",          # número longo (ex: código ML)
        ]
        for pattern in patterns:
            match = re.search(pattern, texto, re.IGNORECASE)
            if match:
                return match.group(1).strip().upper()
        return ""

    # ──────────────────────────────────────────────────────────────────
    # CARREGAMENTO DE XMLs
    # ──────────────────────────────────────────────────────────────────
    def _load_xmls(self, source: str) -> List[Dict]:
        """Carrega todos os XMLs de uma pasta ou arquivo único."""
        files = []
        if os.path.isdir(source):
            files = [
                os.path.join(source, f)
                for f in os.listdir(source)
                if f.lower().endswith(".xml")
            ]
        elif os.path.isfile(source) and source.lower().endswith(".xml"):
            files = [source]

        self._log(f"{len(files)} arquivo(s) XML encontrado(s)")
        data = []
        for i, f in enumerate(files):
            parsed = self._parse_xml(f)
            if parsed:
                data.append(parsed)
            self._progress(int(((i + 1) / max(len(files), 1)) * 30))
        self._log(f"{len(data)} NF-e(s) extraída(s) com sucesso")
        return data

    # ──────────────────────────────────────────────────────────────────
    # NORMALIZAÇÃO DA PLANILHA DE VENDAS
    # ──────────────────────────────────────────────────────────────────
    def _normalizar_planilha(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Mapeia os nomes de coluna da planilha do marketplace para padrão interno.
        Suporta exportações do ML e Shopee.
        """
        col_map = {}
        for col in df.columns:
            c = col.lower().strip().replace(" ", "_")
            if any(k in c for k in ["pedido", "order", "numero_pedido", "order_id", "id_pedido"]):
                col_map[col] = "numero_pedido"
            elif any(k in c for k in ["cpf", "cnpj", "documento", "document"]):
                col_map[col] = "cpf_cnpj"
            elif any(k in c for k in ["valor", "total", "preco", "amount", "price"]):
                if "valor" not in col_map.values() and "total" not in col_map.values():
                    col_map[col] = "valor_venda"
            elif any(k in c for k in ["comprador", "cliente", "buyer", "nome"]):
                col_map[col] = "nome_comprador"

        df_norm = df.rename(columns=col_map)

        # Normalizar CPF (remover formatação)
        if "cpf_cnpj" in df_norm.columns:
            df_norm["cpf_cnpj"] = df_norm["cpf_cnpj"].astype(str).apply(
                lambda x: re.sub(r"\D", "", x)
            )

        # Normalizar número do pedido (maiúsculas, sem espaços)
        if "numero_pedido" in df_norm.columns:
            df_norm["numero_pedido"] = df_norm["numero_pedido"].astype(str).str.strip().str.upper()

        # Normalizar valor
        if "valor_venda" in df_norm.columns:
            df_norm["valor_venda"] = pd.to_numeric(
                df_norm["valor_venda"].astype(str).str.replace(",", ".").str.replace(r"[^\d.]", "", regex=True),
                errors="coerce",
            )

        return df_norm

    # ──────────────────────────────────────────────────────────────────
    # CRUZAMENTO PRINCIPAL
    # ──────────────────────────────────────────────────────────────────
    def cruzar_com_planilha(
        self,
        xml_source: str,
        planilha_path: str,
        chave: str = "auto",        # "auto" | "numero_pedido" | "cpf_cnpj"
        tolerancia_valor: float = 0.01,
        output_path: str = None,
    ) -> Dict:
        """
        Passo-a-passo:
        1. Carrega e parseia XMLs
        2. Carrega planilha do marketplace
        3. Detecta a melhor chave de cruzamento
        4. Faz o merge e classifica cada linha
        5. Gera Excel com 3 abas
        """
        try:
            # 1. Carregar XMLs
            self._log("Iniciando extração dos XMLs...")
            nfe_data = self._load_xmls(xml_source)
            if not nfe_data:
                return {"success": False, "error": "Nenhum XML de NF-e válido encontrado."}

            nfe_df = pd.DataFrame(nfe_data)

            # 2. Carregar planilha do marketplace
            self._log("Carregando planilha do marketplace...")
            ext = Path(planilha_path).suffix.lower()
            if ext == ".csv":
                vendas_df = pd.read_csv(planilha_path, sep=None, engine="python", dtype=str)
            else:
                vendas_df = pd.read_excel(planilha_path, dtype=str)

            vendas_df = self._normalizar_planilha(vendas_df)
            self._log(f"Planilha carregada: {len(vendas_df)} pedido(s)")
            self._progress(40)

            # 3. Detectar chave de cruzamento
            chave_efetiva = self._detectar_chave(nfe_df, vendas_df, chave)
            self._log(f"Chave de cruzamento utilizada: {chave_efetiva}")

            # 4. Cruzar
            resultados = self._cruzar(nfe_df, vendas_df, chave_efetiva, tolerancia_valor)
            self._progress(80)

            # 5. Separar por status
            ok = [r for r in resultados if r["status"] == "✅ OK"]
            divergentes = [r for r in resultados if r["status"] == "⚠️ VALOR DIVERGENTE"]
            faltando = [r for r in resultados if r["status"] == "❌ NOTA FALTANDO"]
            # Pedidos sem nota (presentes na planilha mas sem XML)
            sem_nota = self._pedidos_sem_nota(vendas_df, resultados, chave_efetiva)

            self._log(f"✅ OK: {len(ok)} | ⚠️ Divergentes: {len(divergentes)} | ❌ Faltando: {len(faltando)} | Pedidos s/ Nota: {len(sem_nota)}")

            # 6. Gerar output
            if not output_path:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                import sys
                sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..")))
                import config as cfg
                output_path = os.path.join(cfg.OUTPUT_DIR, f"extrator_nfe_{ts}.xlsx")

            self._save_excel(ok, divergentes, faltando, sem_nota, output_path)
            self._progress(100)

            self._log(f"Relatório gerado: {output_path}")
            return {
                "success": True,
                "total_xmls": len(nfe_data),
                "total_vendas": len(vendas_df),
                "ok": len(ok),
                "divergentes": len(divergentes),
                "faltando": len(faltando),
                "sem_nota": len(sem_nota),
                "chave": chave_efetiva,
                "output_path": output_path,
            }

        except Exception as e:
            import traceback
            self._log(f"❌ Erro: {e}\n{traceback.format_exc()}")
            return {"success": False, "error": str(e)}

    def _detectar_chave(self, nfe_df, vendas_df, chave_solicitada: str) -> str:
        """
        Detecta automaticamente a melhor chave de cruzamento.
        Prioridade: número do pedido > CPF/CNPJ
        """
        if chave_solicitada in ("numero_pedido", "cpf_cnpj"):
            return chave_solicitada

        # Testa se número do pedido aparece em ambos os lados
        tem_pedido_nfe = nfe_df["numero_pedido_xml"].astype(str).str.strip().ne("").any()
        tem_pedido_vendas = "numero_pedido" in vendas_df.columns and vendas_df["numero_pedido"].astype(str).str.strip().ne("").any()

        if tem_pedido_nfe and tem_pedido_vendas:
            return "numero_pedido"

        # Fallback: CPF/CNPJ
        tem_cpf_nfe = nfe_df["cpf_cnpj"].astype(str).str.strip().ne("").any()
        tem_cpf_vendas = "cpf_cnpj" in vendas_df.columns and vendas_df["cpf_cnpj"].astype(str).str.strip().ne("").any()

        if tem_cpf_nfe and tem_cpf_vendas:
            return "cpf_cnpj"

        return "numero_pedido"  # padrão

    def _cruzar(self, nfe_df, vendas_df, chave: str, tolerancia: float) -> List[Dict]:
        """Realiza o cruzamento NF-e ↔ planilha."""
        resultados = []

        # Chave da NF-e
        if chave == "numero_pedido":
            nfe_key_col = "numero_pedido_xml"
            vendas_key_col = "numero_pedido"
        else:
            nfe_key_col = "cpf_cnpj"
            vendas_key_col = "cpf_cnpj"

        # Construir lookup de vendas {chave: row}
        vendas_lookup = {}
        if vendas_key_col in vendas_df.columns:
            for _, row in vendas_df.iterrows():
                k = str(row.get(vendas_key_col, "")).strip().upper()
                if k and k not in ("", "NAN", "NONE"):
                    vendas_lookup[k] = row

        for _, nfe_row in nfe_df.iterrows():
            chave_nfe = str(nfe_row.get(nfe_key_col, "")).strip().upper()
            valor_nf = float(nfe_row.get("valor_nf", 0))

            venda_row = vendas_lookup.get(chave_nfe)

            if venda_row is None:
                # NF-e sem pedido correspondente
                resultados.append({
                    "status": "❌ NOTA FALTANDO",
                    "nfe_numero": nfe_row.get("nfe_numero", ""),
                    "data_emissao": nfe_row.get("data_emissao", ""),
                    "valor_nf": valor_nf,
                    "cpf_cnpj_nf": nfe_row.get("cpf_cnpj", ""),
                    "nome_dest": nfe_row.get("nome_dest", ""),
                    "numero_pedido": nfe_row.get("numero_pedido_xml", ""),
                    "valor_venda": "",
                    "diferenca": "",
                    "arquivo_xml": nfe_row.get("arquivo_xml", ""),
                    "obs": "Pedido não encontrado na planilha do marketplace",
                })
            else:
                valor_venda = 0.0
                try:
                    valor_venda = float(str(venda_row.get("valor_venda", 0)).replace(",", ".") or 0)
                except (ValueError, TypeError):
                    valor_venda = 0.0

                diferenca = round(abs(valor_nf - valor_venda), 2)

                if diferenca <= tolerancia:
                    status = "✅ OK"
                    obs = ""
                else:
                    status = "⚠️ VALOR DIVERGENTE"
                    obs = f"NF-e: R$ {valor_nf:.2f} | Venda: R$ {valor_venda:.2f} | Diff: R$ {diferenca:.2f}"

                resultados.append({
                    "status": status,
                    "nfe_numero": nfe_row.get("nfe_numero", ""),
                    "data_emissao": nfe_row.get("data_emissao", ""),
                    "valor_nf": valor_nf,
                    "cpf_cnpj_nf": nfe_row.get("cpf_cnpj", ""),
                    "nome_dest": nfe_row.get("nome_dest", ""),
                    "numero_pedido": chave_nfe,
                    "valor_venda": valor_venda,
                    "diferenca": diferenca,
                    "arquivo_xml": nfe_row.get("arquivo_xml", ""),
                    "obs": obs,
                })

        return resultados

    def _pedidos_sem_nota(self, vendas_df, resultados, chave: str) -> List[Dict]:
        """Identifica pedidos na planilha do marketplace que não têm NF-e correspondente."""
        chaves_conciliadas = {
            str(r.get("numero_pedido", "")).strip().upper()
            for r in resultados
            if r["status"] != "❌ NOTA FALTANDO"
        }

        sem_nota = []
        col = "numero_pedido" if chave == "numero_pedido" else "cpf_cnpj"
        if col not in vendas_df.columns:
            return sem_nota

        for _, row in vendas_df.iterrows():
            k = str(row.get(col, "")).strip().upper()
            if k and k not in chaves_conciliadas:
                sem_nota.append({
                    "status": "❌ SEM NOTA FISCAL",
                    "numero_pedido": row.get("numero_pedido", row.get(col, "")),
                    "cpf_cnpj_comprador": row.get("cpf_cnpj", ""),
                    "nome_comprador": row.get("nome_comprador", ""),
                    "valor_venda": row.get("valor_venda", ""),
                    "obs": "Pedido encontrado na planilha sem NF-e correspondente",
                })

        return sem_nota

    # ──────────────────────────────────────────────────────────────────
    # EXPORTAÇÃO EXCEL PREMIUM
    # ──────────────────────────────────────────────────────────────────
    def _save_excel(self, ok, divergentes, faltando, sem_nota, output_path):
        wb = Workbook()

        # Paleta
        COLOR_OK = "D4EDDA"
        COLOR_WARN = "FFF3CD"
        COLOR_ERR = "F8D7DA"
        HEADER_BG = "1C1C1E"
        HEADER_FG = "FFFFFF"
        ACCENT = "F5A623"

        thin = Side(style="thin", color="D1D1D6")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def make_sheet(ws, data, title, bg_color):
            # Título
            cols = list(data[0].keys()) if data else ["status", "obs"]
            n_cols = len(cols)
            ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
            cell = ws["A1"]
            cell.value = title
            cell.font = Font(name="Calibri", bold=True, size=13, color=HEADER_FG)
            cell.fill = PatternFill("solid", fgColor=HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 28

            # Cabeçalho das colunas
            for c_idx, col in enumerate(cols, 1):
                cell = ws.cell(row=2, column=c_idx, value=col.replace("_", " ").title())
                cell.font = Font(name="Calibri", bold=True, size=10, color=ACCENT)
                cell.fill = PatternFill("solid", fgColor="2C2C2E")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            ws.row_dimensions[2].height = 35

            # Dados
            for r_idx, row in enumerate(data, 3):
                for c_idx, col in enumerate(cols, 1):
                    val = row.get(col, "")
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.font = Font(name="Calibri", size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border
                    cell.fill = PatternFill("solid", fgColor=bg_color if r_idx % 2 == 0 else "FFFFFF")

            # Largura das colunas
            for c_idx, col in enumerate(cols, 1):
                ws.column_dimensions[get_column_letter(c_idx)].width = max(len(col) + 4, 14)

        # Aba 1: Conciliado (OK)
        ws1 = wb.active
        ws1.title = "✅ Conciliados"
        make_sheet(ws1, ok if ok else [{"status": "Nenhum resultado"}], f"✅ Pedidos OK — {len(ok)} conciliados", COLOR_OK)

        # Aba 2: Divergências
        ws2 = wb.create_sheet("⚠️ Divergências")
        make_sheet(ws2, divergentes if divergentes else [{"status": "Sem divergências"}], f"⚠️ Valor Divergente — {len(divergentes)} pedidos", COLOR_WARN)

        # Aba 3: Notas Faltando (NF sem pedido + Pedidos sem NF)
        todos_problemas = faltando + sem_nota
        ws3 = wb.create_sheet("❌ Problemas")
        make_sheet(ws3, todos_problemas if todos_problemas else [{"status": "Sem problemas"}], f"❌ Problemas — {len(todos_problemas)} itens", COLOR_ERR)

        wb.save(output_path)
