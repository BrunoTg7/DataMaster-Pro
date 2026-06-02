"""
Conciliador Pro v3.0 - Motor Profissional de Conciliação Comercial
Suporta dois modos 100% offline (Sem uso de IA):
1. Modo Clássico: Extrato bancário ↔ Planilha de vendas (com tolerância de data e fuzzy matching de descrição)
2. Modo NF-e: XML de Notas Fiscais ↔ Extrato bancário (com cruzamento de valores, nomes e datas)
"""
import pandas as pd
import xml.etree.ElementTree as ET
import re
import os
import bisect
from typing import Dict, List, Optional, Callable
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class Conciliador:
    THEMES = {
        "classic_blue": {
            "header_fill": "1F4E79",        # Azul Escuro Clássico
            "header_font_color": "FFFFFF",
            "zebra_fill": "F2F5F8",         # Azul Muito Claro
            "border_color": "D9D9D9",
            "accent_fill": "DDEBF7",
            "summary_accent": "1F4E79"
        },
        "emerald_green": {
            "header_fill": "1E4620",        # Verde Floresta
            "header_font_color": "FFFFFF",
            "zebra_fill": "F4F9F4",         # Verde Muito Claro
            "border_color": "D9D9D9",
            "accent_fill": "E2EFDA",
            "summary_accent": "1E4620"
        },
        "modern_orange": {
            "header_fill": "262626",        # Cinza Escuro Grafite
            "header_font_color": "FFFFFF",
            "zebra_fill": "FFF2E6",         # Laranja Muito Claro
            "border_color": "D9D9D9",
            "accent_fill": "FCE4D6",
            "summary_accent": "E26B0A"      # Laranja Queimado
        },
        "slate_gray": {
            "header_fill": "404040",        # Slate Gray Escuro
            "header_font_color": "FFFFFF",
            "zebra_fill": "F2F2F2",         # Cinza Muito Claro
            "border_color": "D9D9D9",
            "accent_fill": "EAEAEA",
            "summary_accent": "595959"
        }
    }

    def __init__(self, log_callback: Callable = None):
        self.log_callback = log_callback

    def _log(self, message: str):
        if self.log_callback:
            self.log_callback(message)

    # ========================================================================
    # MODO CLÁSSICO: Extrato ↔ Vendas
    # ========================================================================
    def reconcile_classic(
        self,
        extract_file: str,
        sales_file: str,
        output_path: str,
        tolerance: float = 0.01,
        date_tolerance_days: int = 2,
        fuzzy_threshold: int = 75,
        visual_theme: str = "classic_blue"
    ) -> Dict:
        """Concilia extrato bancário com planilha de vendas"""
        try:
            self._log("Iniciando conciliação clássica...")

            if not Path(extract_file).exists():
                return {"success": False, "error": f"Arquivo extrato '{extract_file}' não encontrado."}
            if not Path(sales_file).exists():
                return {"success": False, "error": f"Arquivo vendas '{sales_file}' não encontrado."}

            extract_df = self._load_file(extract_file)
            sales_df = self._load_file(sales_file)

            if extract_df is None or sales_df is None:
                return {"success": False, "error": "Erro ao carregar arquivos"}

            extract_df = self._normalize_columns(extract_df)
            sales_df = self._normalize_columns_sales(sales_df)

            # Garante que os índices sejam sequenciais e únicos para o rastreamento
            extract_df = extract_df.reset_index(drop=True)
            sales_df = sales_df.reset_index(drop=True)

            matched = []
            matched_extract = set()
            matched_sales = set()

            # Pre-conversão vetorizada das datas para garantir datetimes
            extract_df['date'] = pd.to_datetime(extract_df['date'], errors='coerce')
            sales_df['date'] = pd.to_datetime(sales_df['date'], errors='coerce')

            # =================================================================
            # PASSO 1: Casamento Exato (Valor Exato + Data Exata + Fuzzy se houver)
            # =================================================================
            from collections import defaultdict
            sales_exact_map = defaultdict(list)
            for idx_s, row_s in sales_df.iterrows():
                val_s = float(row_s.get("amount", 0))
                date_s = row_s.get("date")
                dt_key = date_s if pd.notna(date_s) else None
                # Arredonda valor para evitar discrepâncias float
                sales_exact_map[(round(val_s, 4), dt_key)].append(idx_s)

            for idx_e, row_e in extract_df.iterrows():
                val_e = float(row_e.get("amount", 0))
                date_e = row_e.get("date")
                dt_key = date_e if pd.notna(date_e) else None

                candidates = sales_exact_map.get((round(val_e, 4), dt_key), [])
                best_idx_s = None

                for idx_s in candidates:
                    if idx_s in matched_sales:
                        continue
                    row_s = sales_df.iloc[idx_s]

                    # Checagem de descrição opcional
                    desc_match = True
                    if fuzzy_threshold > 0:
                        desc_e = str(row_e.get("description", "")).strip().lower()
                        desc_s = str(row_s.get("description", "")).strip().lower()
                        if desc_e and desc_s:
                            try:
                                from fuzzywuzzy import fuzz
                                score = fuzz.token_sort_ratio(desc_e, desc_s)
                                if score < fuzzy_threshold:
                                    desc_match = False
                            except ImportError:
                                # Fallback simples
                                if desc_e not in desc_s and desc_s not in desc_e:
                                    desc_match = False

                    if desc_match:
                        best_idx_s = idx_s
                        break

                if best_idx_s is not None:
                    matched_extract.add(idx_e)
                    matched_sales.add(best_idx_s)
                    row_s = sales_df.iloc[best_idx_s]
                    matched.append({
                        "status": "CONCILIADO",
                        "data": date_e.strftime('%Y-%m-%d') if pd.notna(date_e) else "",
                        "date": date_e.strftime('%Y-%m-%d') if pd.notna(date_e) else "",
                        "valor": val_e,
                        "amount": val_e,
                        "descricao": row_e.get("description", ""),
                        "description": row_e.get("description", ""),
                        "venda_id": row_s.get("id", ""),
                        "origem_match": "EXATO"
                    })

            # =================================================================
            # PASSO 2: Casamento por Tolerância (Valor Tolerância + Janela de Datas + Fuzzy)
            # =================================================================
            unmatched_sales_indices = [idx for idx in range(len(sales_df)) if idx not in matched_sales]
            if unmatched_sales_indices:
                sales_unmatched_sub = sales_df.iloc[unmatched_sales_indices].copy()
                sales_unmatched_sub['original_idx'] = unmatched_sales_indices
                sales_unmatched_sub = sales_unmatched_sub.sort_values('amount')

                sales_amounts = sales_unmatched_sub['amount'].values
                sales_sub_indices = sales_unmatched_sub['original_idx'].values

                for idx_e, row_e in extract_df.iterrows():
                    if idx_e in matched_extract:
                        continue

                    val_e = float(row_e.get("amount", 0))
                    date_e = row_e.get("date")

                    # Busca binária O(log M)
                    left_idx = bisect.bisect_left(sales_amounts, val_e - tolerance)
                    right_idx = bisect.bisect_right(sales_amounts, val_e + tolerance)

                    best_idx_s = None
                    best_date_diff = float('inf')

                    for idx_in_sub in range(left_idx, right_idx):
                        idx_s = sales_sub_indices[idx_in_sub]
                        if idx_s in matched_sales:
                            continue

                        row_s = sales_df.iloc[idx_s]
                        date_s = row_s.get("date")

                        # Checagem de tolerância de data
                        date_match = False
                        date_diff = float('inf')
                        if pd.notna(date_e) and pd.notna(date_s):
                            date_diff = abs((date_e - date_s).days)
                            if date_diff <= date_tolerance_days:
                                date_match = True

                        if not date_match:
                            continue

                        # Checagem de descrição opcional
                        desc_match = True
                        if fuzzy_threshold > 0:
                            desc_e = str(row_e.get("description", "")).strip().lower()
                            desc_s = str(row_s.get("description", "")).strip().lower()
                            if desc_e and desc_s:
                                try:
                                    from fuzzywuzzy import fuzz
                                    score = fuzz.token_sort_ratio(desc_e, desc_s)
                                    if score < fuzzy_threshold:
                                        desc_match = False
                                except ImportError:
                                    if desc_e not in desc_s and desc_s not in desc_e:
                                        desc_match = False

                        if desc_match:
                            if date_diff < best_date_diff:
                                best_idx_s = idx_s
                                best_date_diff = date_diff

                    if best_idx_s is not None:
                        matched_extract.add(idx_e)
                        matched_sales.add(best_idx_s)
                        row_s = sales_df.iloc[best_idx_s]
                        matched.append({
                            "status": "CONCILIADO",
                            "data": date_e.strftime('%Y-%m-%d') if pd.notna(date_e) else "",
                            "date": date_e.strftime('%Y-%m-%d') if pd.notna(date_e) else "",
                            "valor": val_e,
                            "amount": val_e,
                            "descricao": row_e.get("description", ""),
                            "description": row_e.get("description", ""),
                            "venda_id": row_s.get("id", ""),
                            "origem_match": "TOLERANCIA"
                        })

            # =================================================================
            # PASSO 3: Fallback de Valor (Valor Tolerância apenas + Fuzzy, ignorando data)
            # =================================================================
            unmatched_sales_indices = [idx for idx in range(len(sales_df)) if idx not in matched_sales]
            if unmatched_sales_indices:
                sales_unmatched_sub = sales_df.iloc[unmatched_sales_indices].copy()
                sales_unmatched_sub['original_idx'] = unmatched_sales_indices
                sales_unmatched_sub = sales_unmatched_sub.sort_values('amount')

                sales_amounts = sales_unmatched_sub['amount'].values
                sales_sub_indices = sales_unmatched_sub['original_idx'].values

                for idx_e, row_e in extract_df.iterrows():
                    if idx_e in matched_extract:
                        continue

                    val_e = float(row_e.get("amount", 0))
                    date_e = row_e.get("date")

                    # Busca binária O(log M)
                    left_idx = bisect.bisect_left(sales_amounts, val_e - tolerance)
                    right_idx = bisect.bisect_right(sales_amounts, val_e + tolerance)

                    best_idx_s = None

                    for idx_in_sub in range(left_idx, right_idx):
                        idx_s = sales_sub_indices[idx_in_sub]
                        if idx_s in matched_sales:
                            continue

                        row_s = sales_df.iloc[idx_s]
                        date_s = row_s.get("date")

                        # Se ambas as datas estão preenchidas e a diferença excede a tolerância, não casar!
                        if pd.notna(date_e) and pd.notna(date_s):
                            if abs((date_e - date_s).days) > date_tolerance_days:
                                continue

                        # Checagem de descrição opcional
                        desc_match = True
                        if fuzzy_threshold > 0:
                            desc_e = str(row_e.get("description", "")).strip().lower()
                            desc_s = str(row_s.get("description", "")).strip().lower()
                            if desc_e and desc_s:
                                try:
                                    from fuzzywuzzy import fuzz
                                    score = fuzz.token_sort_ratio(desc_e, desc_s)
                                    if score < fuzzy_threshold:
                                        desc_match = False
                                except ImportError:
                                    if desc_e not in desc_s and desc_s not in desc_e:
                                        desc_match = False

                        if desc_match:
                            best_idx_s = idx_s
                            break

                    if best_idx_s is not None:
                        matched_extract.add(idx_e)
                        matched_sales.add(best_idx_s)
                        row_s = sales_df.iloc[best_idx_s]
                        matched.append({
                            "status": "CONCILIADO",
                            "data": date_e.strftime('%Y-%m-%d') if pd.notna(date_e) else "",
                            "date": date_e.strftime('%Y-%m-%d') if pd.notna(date_e) else "",
                            "valor": val_e,
                            "amount": val_e,
                            "descricao": row_e.get("description", ""),
                            "description": row_e.get("description", ""),
                            "venda_id": row_s.get("id", ""),
                            "origem_match": "FALLBACK_VALOR"
                        })

            # =================================================================
            # CONSTRUÇÃO DOS NÃO CONCILIADOS E GERAÇÃO DO DATAFRAME DE RESULTADO
            # =================================================================
            unmatched_extract = []
            for idx_e, row_e in extract_df.iterrows():
                if idx_e not in matched_extract:
                    date_e = row_e.get("date")
                    val_e = float(row_e.get("amount", 0))
                    unmatched_extract.append({
                        "status": "PENDENTE (sem venda)",
                        "data": date_e.strftime('%Y-%m-%d') if pd.notna(date_e) else "",
                        "date": date_e.strftime('%Y-%m-%d') if pd.notna(date_e) else "",
                        "valor": val_e,
                        "amount": val_e,
                        "descricao": row_e.get("description", ""),
                        "description": row_e.get("description", ""),
                        "venda_id": "-",
                        "origem_match": "-"
                    })

            unmatched_sales = []
            for idx_s, row_s in sales_df.iterrows():
                if idx_s not in matched_sales:
                    date_s = row_s.get("date")
                    val_s = float(row_s.get("amount", 0))
                    unmatched_sales.append({
                        "status": "PENDENTE (sem recebimento)",
                        "data": date_s.strftime('%Y-%m-%d') if pd.notna(date_s) else "",
                        "date": date_s.strftime('%Y-%m-%d') if pd.notna(date_s) else "",
                        "valor": val_s,
                        "amount": val_s,
                        "descricao": row_s.get("description", ""),
                        "description": row_s.get("description", ""),
                        "venda_id": row_s.get("id", "") if pd.notna(row_s.get("id")) else "",
                        "origem_match": "-"
                    })

            # Combina tudo em um único DataFrame de saída
            result_df = pd.DataFrame(matched + unmatched_extract + unmatched_sales)
            
            # Exportação estilizada premium
            metrics = {
                "total_extract": len(extract_df),
                "total_sales": len(sales_df),
                "matched": len(matched),
                "unmatched_extract": len(unmatched_extract),
                "unmatched_sales": len(unmatched_sales)
            }
            self._save_premium_excel(result_df, output_path, visual_theme, "classic", metrics)

            return {
                "success": True,
                "mode": "classic",
                "total_extract": len(extract_df),
                "total_sales": len(sales_df),
                "matched": len(matched),
                "unmatched_extract": len(unmatched_extract),
                "unmatched_sales": len(unmatched_sales),
                "output_path": output_path
            }
        except Exception as e:
            return {"success": False, "error": str(e)}

    # ========================================================================
    # MODO NF-e: XML ↔ Extrato
    # ========================================================================
    def reconcile_nfe(
        self,
        xml_folder: str,
        bank_file: str,
        output_path: str,
        tolerance: float = 0.05,
        date_tolerance_days: int = 5,
        fuzzy_threshold: int = 60,
        visual_theme: str = "classic_blue"
    ) -> Dict:
        """Concilia XMLs de NF-e com extrato bancário"""
        try:
            self._log("Iniciando conciliação de NF-e...")

            nfe_data = self._load_nfe_files(xml_folder)
            if not nfe_data:
                return {"success": False, "error": "Nenhum XML de NF-e encontrado"}

            nfe_df = pd.DataFrame(nfe_data)
            self._log(f"{len(nfe_df)} notas fiscais carregadas")

            bank_df = self._load_file(bank_file)
            if bank_df is None:
                return {"success": False, "error": "Erro ao carregar extrato bancário"}

            bank_df = self._normalize_columns(bank_df)
            bank_df['valor'] = bank_df['valor'].abs()
            self._log(f"Extrato carregado: {len(bank_df)} transações")

            # Reseta e sequencia os índices
            nfe_df = nfe_df.reset_index(drop=True)
            bank_df = bank_df.reset_index(drop=True)

            # Pre-conversão das datas
            nfe_df['date_parsed'] = pd.to_datetime(nfe_df['data'], errors='coerce')
            bank_df['date_parsed'] = pd.to_datetime(bank_df['data'], errors='coerce')

            matched = []
            matched_nfe = set()
            matched_banks = set()

            # =================================================================
            # PASSO 1: Casamento Exato (Valor Exato + Data Exata + Fuzzy se houver)
            # =================================================================
            from collections import defaultdict
            bank_exact_map = defaultdict(list)
            for idx_b, bank in bank_df.iterrows():
                val_b = float(bank.get("valor", 0))
                date_b = bank.get("date_parsed")
                dt_key = date_b if pd.notna(date_b) else None
                # Arredonda valor para evitar discrepâncias float
                bank_exact_map[(round(val_b, 4), dt_key)].append(idx_b)

            for idx_n, nfe in nfe_df.iterrows():
                val_n = float(nfe.get("valor", 0))
                date_n = nfe.get("date_parsed")
                dt_key = date_n if pd.notna(date_n) else None

                candidates = bank_exact_map.get((round(val_n, 4), dt_key), [])
                best_idx_b = None

                for idx_b in candidates:
                    if idx_b in matched_banks:
                        continue
                    bank = bank_df.iloc[idx_b]

                    # Checagem de descrição/cliente
                    desc_match = True
                    if fuzzy_threshold > 0:
                        cliente = str(nfe.get("cliente", "")).strip().lower()
                        desc_b = str(bank.get("descricao", "")).strip().lower()
                        if cliente and desc_b:
                            try:
                                from fuzzywuzzy import fuzz
                                score = fuzz.token_sort_ratio(cliente, desc_b)
                                if score < fuzzy_threshold:
                                    desc_match = False
                            except ImportError:
                                if cliente not in desc_b and desc_b not in cliente:
                                    desc_match = False

                    if desc_match:
                        best_idx_b = idx_b
                        break

                if best_idx_b is not None:
                    matched_nfe.add(idx_n)
                    matched_banks.add(best_idx_b)
                    bank = bank_df.iloc[best_idx_b]
                    matched.append({
                        "status": "CONCILIADO",
                        "nfe_numero": nfe['numero'],
                        "nfe_serie": nfe.get('serie', ''),
                        "cliente": nfe['cliente'],
                        "data_nfe": nfe['data'],
                        "valor_nfe": nfe['valor'],
                        "data_pagamento": bank['data'].strftime('%Y-%m-%d') if hasattr(bank['data'], "strftime") and pd.notna(bank['data']) else str(bank['data']),
                        "valor_pago": bank['valor'],
                        "descricao_banco": bank.get('descricao', ''),
                        "origem_match": "EXATO"
                    })

            # =================================================================
            # PASSO 2: Casamento por Tolerância (Valor Tolerância + Janela de Datas + Fuzzy)
            # =================================================================
            unmatched_bank_indices = [idx for idx in range(len(bank_df)) if idx not in matched_banks]
            if unmatched_bank_indices:
                bank_unmatched_sub = bank_df.iloc[unmatched_bank_indices].copy()
                bank_unmatched_sub['original_idx'] = unmatched_bank_indices
                bank_unmatched_sub = bank_unmatched_sub.sort_values('valor')

                bank_vals = bank_unmatched_sub['valor'].values
                bank_sub_indices = bank_unmatched_sub['original_idx'].values

                for idx_n, nfe in nfe_df.iterrows():
                    if idx_n in matched_nfe:
                        continue

                    val_n = float(nfe.get("valor", 0))
                    date_n = nfe.get("date_parsed")

                    # Busca binária O(log M)
                    left_idx = bisect.bisect_left(bank_vals, val_n - tolerance)
                    right_idx = bisect.bisect_right(bank_vals, val_n + tolerance)

                    best_idx_b = None
                    best_diff = float('inf')

                    for idx_in_sub in range(left_idx, right_idx):
                        idx_b = bank_sub_indices[idx_in_sub]
                        if idx_b in matched_banks:
                            continue

                        bank = bank_df.iloc[idx_b]
                        date_b = bank.get("date_parsed")
                        val_b = float(bank.get("valor", 0))
                        diff = abs(val_n - val_b)

                        # Checagem de data
                        date_match = True
                        if pd.notna(date_n) and pd.notna(date_b):
                            date_diff = abs((date_n - date_b).days)
                            if date_diff > date_tolerance_days:
                                date_match = False

                        if not date_match:
                            continue

                        # Checagem de descrição/cliente
                        desc_match = True
                        if fuzzy_threshold > 0:
                            cliente = str(nfe.get("cliente", "")).strip().lower()
                            desc_b = str(bank.get("descricao", "")).strip().lower()
                            if cliente and desc_b:
                                try:
                                    from fuzzywuzzy import fuzz
                                    score = fuzz.token_sort_ratio(cliente, desc_b)
                                    if score < fuzzy_threshold:
                                        desc_match = False
                                except ImportError:
                                    if cliente not in desc_b and desc_b not in cliente:
                                        desc_match = False

                        if desc_match:
                            if diff < best_diff:
                                best_idx_b = idx_b
                                best_diff = diff

                    if best_idx_b is not None:
                        matched_nfe.add(idx_n)
                        matched_banks.add(best_idx_b)
                        bank = bank_df.iloc[best_idx_b]
                        matched.append({
                            "status": "CONCILIADO",
                            "nfe_numero": nfe['numero'],
                            "nfe_serie": nfe.get('serie', ''),
                            "cliente": nfe['cliente'],
                            "data_nfe": nfe['data'],
                            "valor_nfe": nfe['valor'],
                            "data_pagamento": bank['data'].strftime('%Y-%m-%d') if hasattr(bank['data'], "strftime") and pd.notna(bank['data']) else str(bank['data']),
                            "valor_pago": bank['valor'],
                            "descricao_banco": bank.get('descricao', ''),
                            "origem_match": "TOLERANCIA"
                        })

            # =================================================================
            # PASSO 3: Fallback de Valor (Valor Tolerância apenas + Fuzzy, ignorando data)
            # =================================================================
            unmatched_bank_indices = [idx for idx in range(len(bank_df)) if idx not in matched_banks]
            if unmatched_bank_indices:
                bank_unmatched_sub = bank_df.iloc[unmatched_bank_indices].copy()
                bank_unmatched_sub['original_idx'] = unmatched_bank_indices
                bank_unmatched_sub = bank_unmatched_sub.sort_values('valor')

                bank_vals = bank_unmatched_sub['valor'].values
                bank_sub_indices = bank_unmatched_sub['original_idx'].values

                for idx_n, nfe in nfe_df.iterrows():
                    if idx_n in matched_nfe:
                        continue

                    val_n = float(nfe.get("valor", 0))
                    date_n = nfe.get("date_parsed")

                    # Busca binária O(log M)
                    left_idx = bisect.bisect_left(bank_vals, val_n - tolerance)
                    right_idx = bisect.bisect_right(bank_vals, val_n + tolerance)

                    best_idx_b = None
                    best_diff = float('inf')

                    for idx_in_sub in range(left_idx, right_idx):
                        idx_b = bank_sub_indices[idx_in_sub]
                        if idx_b in matched_banks:
                            continue

                        bank = bank_df.iloc[idx_b]
                        val_b = float(bank.get("valor", 0))
                        date_b = bank.get("date_parsed")
                        diff = abs(val_n - val_b)

                        # Se ambas as datas estão preenchidas e a diferença excede a tolerância, não casar!
                        if pd.notna(date_n) and pd.notna(date_b):
                            if abs((date_n - date_b).days) > date_tolerance_days:
                                continue

                        # Checagem de descrição/cliente
                        desc_match = True
                        if fuzzy_threshold > 0:
                            cliente = str(nfe.get("cliente", "")).strip().lower()
                            desc_b = str(bank.get("descricao", "")).strip().lower()
                            if cliente and desc_b:
                                try:
                                    from fuzzywuzzy import fuzz
                                    score = fuzz.token_sort_ratio(cliente, desc_b)
                                    if score < fuzzy_threshold:
                                        desc_match = False
                                except ImportError:
                                    if cliente not in desc_b and desc_b not in cliente:
                                        desc_match = False

                        if desc_match:
                            if diff < best_diff:
                                best_idx_b = idx_b
                                best_diff = diff

                    if best_idx_b is not None:
                        matched_nfe.add(idx_n)
                        matched_banks.add(best_idx_b)
                        bank = bank_df.iloc[best_idx_b]
                        matched.append({
                            "status": "CONCILIADO",
                            "nfe_numero": nfe['numero'],
                            "nfe_serie": nfe.get('serie', ''),
                            "cliente": nfe['cliente'],
                            "data_nfe": nfe['data'],
                            "valor_nfe": nfe['valor'],
                            "data_pagamento": bank['data'].strftime('%Y-%m-%d') if hasattr(bank['data'], "strftime") and pd.notna(bank['data']) else str(bank['data']),
                            "valor_pago": bank['valor'],
                            "descricao_banco": bank.get('descricao', ''),
                            "origem_match": "FALLBACK_VALOR"
                        })

            # =================================================================
            # CONSTRUÇÃO DOS NÃO CONCILIADOS E GERAÇÃO DO DATAFRAME DE RESULTADO
            # =================================================================
            unmatched_nfe = []
            for idx_n, nfe in nfe_df.iterrows():
                if idx_n not in matched_nfe:
                    unmatched_nfe.append({
                        "status": "PENDENTE (não pago)",
                        "nfe_numero": nfe['numero'],
                        "nfe_serie": nfe.get('serie', ''),
                        "cliente": nfe['cliente'],
                        "data_nfe": nfe['data'],
                        "valor_nfe": nfe['valor'],
                        "data_pagamento": "-",
                        "valor_pago": 0.0,
                        "descricao_banco": "Sem correspondente no extrato",
                        "origem_match": "-"
                    })

            unmatched_bank = []
            for idx_b, bank in bank_df.iterrows():
                if idx_b not in matched_banks:
                    unmatched_bank.append({
                        "status": "SEM NF CORRESPONDENTE",
                        "nfe_numero": "-",
                        "nfe_serie": "-",
                        "cliente": "-",
                        "data_nfe": "-",
                        "valor_nfe": 0.0,
                        "data_pagamento": bank['data'].strftime('%Y-%m-%d') if hasattr(bank['data'], "strftime") and pd.notna(bank['data']) else str(bank['data']),
                        "valor_pago": bank['valor'],
                        "descricao_banco": bank.get('descricao', ''),
                        "origem_match": "-"
                    })

            # Combina tudo em um único DataFrame de saída
            result_df = pd.DataFrame(matched + unmatched_nfe + unmatched_bank)
            
            # Exportação estilizada premium
            metrics = {
                "total_nfe": len(nfe_df),
                "total_bank": len(bank_df),
                "matched": len(matched),
                "unmatched_nfe": len(unmatched_nfe),
                "unmatched_bank": len(unmatched_bank)
            }
            self._save_premium_excel(result_df, output_path, visual_theme, "nfe", metrics)

            return {
                "success": True,
                "mode": "nfe",
                "total_nfe": len(nfe_df),
                "total_bank": len(bank_df),
                "matched": len(matched),
                "unmatched_nfe": len(unmatched_nfe),
                "unmatched_bank": len(unmatched_bank),
                "output_path": output_path
            }
        except Exception as e:
            return {"success": False, "error": str(e)}

    # ========================================================================
    # MODO NF-e + VENDAS: XML ↔ Planilha de Vendas (usando ExtratorNFe internamente)
    # ========================================================================
    def reconcile_nfe_vendas(
        self,
        xml_folder: str,
        sales_file: str,
        output_path: str,
        tolerance: float = 0.01,
        chave: str = "auto",
        visual_theme: str = "classic_blue"
    ) -> Dict:
        """Cruza XMLs de NF-e com planilha de vendas do marketplace (reusa motor ExtratorNFe)"""
        try:
            from src.tools.extrator_nfe.extrator_nfe_v1 import ExtratorNFe
            extrator = ExtratorNFe(log_callback=self.log_callback)
            result = extrator.cruzar_com_planilha(
                xml_source=xml_folder,
                planilha_path=sales_file,
                chave=chave,
                tolerancia_valor=tolerance,
                output_path=output_path,
            )
            if result.get("success"):
                result["mode"] = "nfe_vendas"
            return result
        except Exception as e:
            return {"success": False, "error": str(e)}

    # ========================================================================
    # COMPATIBILIDADE - Método antigo
    # ========================================================================
    def reconcile(self, extract_file: str, sales_file: str, output_path: str, tolerance: float = 0.01) -> Dict:
        try:
            return self.reconcile_classic(extract_file, sales_file, output_path, tolerance)
        except Exception as e:
            return {"success": False, "error": str(e)}

    # ========================================================================
    # HELPERS DE CARREGAMENTO & NORMALIZAÇÃO
    # ========================================================================
    def _load_nfe_files(self, path: str) -> List[Dict]:
        """Carrega XMLs de NF-e de arquivo único ou pasta"""
        data = []
        files = []

        if os.path.isdir(path):
            files = [os.path.join(path, f) for f in os.listdir(path) if f.lower().endswith('.xml')]
        elif os.path.isfile(path) and path.lower().endswith('.xml'):
            files = [path]

        for xml_file in files:
            try:
                tree = ET.parse(xml_file)
                root = tree.getroot()
                ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}

                infNFe = root.find('.//nfe:infNFe', ns)
                if infNFe is None:
                    infNFe = root.find('.//infNFe')

                if infNFe is not None:
                    ide = infNFe.find('nfe:ide', ns)
                    if ide is None:
                        ide = infNFe.find('ide')

                    n_nf = "0"
                    if ide is not None:
                        n_nf_el = ide.find('nfe:nNF', ns)
                        if n_nf_el is None:
                            n_nf_el = ide.find('nNF')
                        if n_nf_el is not None:
                            n_nf = n_nf_el.text

                    d_emi = ""
                    if ide is not None:
                        dhEmi = ide.find('nfe:dhEmi', ns)
                        if dhEmi is None:
                            dhEmi = ide.find('dhEmi')
                        if dhEmi is not None and dhEmi.text:
                            d_emi = dhEmi.text[:10]
                        else:
                            dEmi = ide.find('dEmi', ns)
                            if dEmi is None:
                                dEmi = ide.find('dEmi')
                            if dEmi is not None and dEmi.text:
                                d_emi = dEmi.text

                    total = infNFe.find('nfe:total/nfe:ICMSTot', ns)
                    if total is None:
                        total = infNFe.find('.//ICMSTot')

                    v_nf = "0"
                    if total is not None:
                        v_nf_el = total.find('nfe:vNF', ns)
                        if v_nf_el is None:
                            v_nf_el = total.find('vNF')
                        if v_nf_el is not None:
                            v_nf = v_nf_el.text

                    dest = infNFe.find('nfe:dest', ns)
                    if dest is None:
                        dest = infNFe.find('dest')

                    x_nome = "Consumidor"
                    if dest is not None:
                        x_nome_el = dest.find('nfe:xNome', ns)
                        if x_nome_el is None:
                            x_nome_el = dest.find('xNome')
                        if x_nome_el is not None and x_nome_el.text:
                            x_nome = x_nome_el.text

                    serie = ""
                    if ide is not None:
                        serie_el = ide.find('nfe:serie', ns)
                        if serie_el is None:
                            serie_el = ide.find('serie')
                        if serie_el is not None and serie_el.text:
                            serie = serie_el.text

                    data.append({
                        "numero": n_nf,
                        "serie": serie,
                        "data": d_emi,
                        "valor": float(v_nf) if v_nf else 0.0,
                        "cliente": x_nome
                    })
            except Exception as e:
                self._log(f"Erro ao ler {xml_file}: {e}")

        return data

    def _load_file(self, file_path: str) -> Optional[pd.DataFrame]:
        path = Path(file_path)
        try:
            if path.suffix.lower() == ".csv":
                for enc in ['utf-8', 'latin1', 'iso-8859-1']:
                    try:
                        return pd.read_csv(path, encoding=enc, sep=None, engine='python')
                    except Exception:
                        continue
            elif path.suffix.lower() in {".xlsx", ".xls"}:
                return pd.read_excel(path)
            elif path.suffix.lower() == ".ofx":
                return self._parse_ofx(path)
        except Exception as e:
            self._log(f"Erro ao carregar {file_path}: {e}")
        return None

    def _normalize_columns(self, df: pd.DataFrame, file_type: str = "extract") -> pd.DataFrame:
        mapping = {}
        for col in df.columns:
            c = str(col).lower()
            if any(x in c for x in ["data", "date", "dt"]):
                mapping[col] = "date"
            elif any(x in c for x in ["valor", "amount", "val", "lançamento"]):
                mapping[col] = "amount"
            elif any(x in c for x in ["desc", "hist", "memo", "origem", "description"]):
                mapping[col] = "description"
            elif any(x in c for x in ["id", "codigo", "número", "nota"]):
                mapping[col] = "id"

        df = df.rename(columns=mapping)

        if 'amount' in df.columns:
            df['amount'] = df['amount'].astype(str).str.replace(r'[^\d,.-]', '', regex=True).str.replace(',', '.').str.replace('R$', '').str.strip()
            df['amount'] = pd.to_numeric(df['amount'], errors='coerce').fillna(0.0)
            df['valor'] = df['amount']

        if 'date' in df.columns:
            df['date'] = pd.to_datetime(df['date'], errors='coerce')
            df['data'] = df['date']

        if 'description' in df.columns:
            df['descricao'] = df['description']

        target_col = 'amount' if 'amount' in df.columns else ('valor' if 'valor' in df.columns else None)
        if target_col:
            return df.dropna(subset=[target_col])
        return df

    def _normalize_columns_sales(self, df: pd.DataFrame) -> pd.DataFrame:
        df = self._normalize_columns(df, "sales")
        return df

    def _parse_ofx(self, file_path: Path) -> Optional[pd.DataFrame]:
        try:
            with open(file_path, "r", encoding="utf-8", errors='ignore') as f:
                content = f.read()

            transactions = []
            pattern = r'<STMTTRN>.*?</STMTTRN>'

            for match in re.finditer(pattern, content, re.DOTALL):
                trn = match.group(0)
                date = re.search(r'<DTPOSTED>(\d{8})', trn)
                amt = re.search(r'<TRNAMT>([-\d.]+)', trn)
                memo = re.search(r'<MEMO>([^<]+)', trn)

                if date and amt:
                    dt_obj = datetime.strptime(date.group(1), '%Y%m%d')
                    transactions.append({
                        "data": dt_obj,
                        "valor": abs(float(amt.group(1))),
                        "descricao": memo.group(1) if memo else "Transação"
                    })

            return pd.DataFrame(transactions) if transactions else None
        except Exception:
            return None

    def quick_check(self, file1: str, file2: str) -> Dict:
        """Verificação rápida da integridade dos dois arquivos"""
        try:
            df1 = self._load_file(file1)
            df2 = self._load_file(file2)
            if df1 is None or df2 is None:
                return {"success": False, "error": "Erro ao carregar arquivos"}
            return {
                "success": True,
                "file1_rows": len(df1),
                "file2_rows": len(df2)
            }
        except Exception as e:
            return {"success": False, "error": str(e)}

    # ========================================================================
    # MOTOR DE FORMATAÇÃO PREMIUM COM OPENPYXL
    # ========================================================================
    def _save_premium_excel(
        self,
        df: pd.DataFrame,
        output_path: str,
        theme_name: str,
        mode: str,
        metrics: Dict
    ):
        """Salva a conciliação aplicando formatação premium comercial e dashboard"""
        theme = self.THEMES.get(theme_name, self.THEMES["classic_blue"])
        wb = Workbook()

        # ----------------- ABA 1: PAINEL DE CONCILIAÇÃO -----------------
        ws_resumo = wb.active
        ws_resumo.title = "📊 Resumo Conciliação"
        ws_resumo.views.sheetView[0].showGridLines = True

        # Fontes e Cores
        title_font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
        section_font = Font(name="Segoe UI", size=12, bold=True, color=theme["header_fill"])
        header_font = Font(name="Segoe UI", size=10, bold=True, color=theme["header_font_color"])
        bold_font = Font(name="Segoe UI", size=10, bold=True)
        regular_font = Font(name="Segoe UI", size=10)

        header_fill = PatternFill(start_color=theme["header_fill"], end_color=theme["header_fill"], fill_type="solid")
        accent_fill = PatternFill(start_color=theme["accent_fill"], end_color=theme["accent_fill"], fill_type="solid")
        zebra_fill = PatternFill(start_color=theme["zebra_fill"], end_color=theme["zebra_fill"], fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        thin_border = Border(
            left=Side(style='thin', color=theme["border_color"]),
            right=Side(style='thin', color=theme["border_color"]),
            top=Side(style='thin', color=theme["border_color"]),
            bottom=Side(style='thin', color=theme["border_color"])
        )

        # Cabeçalho Principal
        ws_resumo.merge_cells("A1:D2")
        title_cell = ws_resumo["A1"]
        title_cell.value = "   CONCILIAÇÃO DE CONTAS - DASHBOARD PREMIUM"
        title_cell.font = title_font
        title_cell.fill = header_fill
        title_cell.alignment = Alignment(vertical="center", horizontal="left")

        # Fundo das células mescladas
        for r in range(1, 3):
            for c in range(1, 5):
                ws_resumo.cell(row=r, column=c).fill = header_fill

        # Estatísticas Gerais
        ws_resumo["A4"] = "Resumo Geral da Execução"
        ws_resumo["A4"].font = section_font

        # Calcular taxa de match
        total_matched = metrics.get("matched", 0)
        if mode == "classic":
            total_items = metrics.get("total_extract", 0)
        else:
            total_items = metrics.get("total_nfe", 0)
        
        match_rate = f"{(total_matched / total_items * 100):.1f}%" if total_items > 0 else "0.0%"

        stats = [
            ("Data da Execução", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
            ("Modo Operacional", "Clássico (Extrato ↔ Vendas)" if mode == "classic" else "Notas Fiscais (XML ↔ Extrato)"),
            ("Registros Casados", total_matched),
            ("Taxa de Correspondência", match_rate),
            ("Estilo Visual Aplicado", theme_name.replace("_", " ").title())
        ]

        row_idx = 5
        for key, val in stats:
            ws_resumo.cell(row=row_idx, column=1, value=key).font = bold_font
            ws_resumo.cell(row=row_idx, column=1).border = thin_border
            ws_resumo.cell(row=row_idx, column=1).fill = zebra_fill

            ws_resumo.cell(row=row_idx, column=2, value=val).font = regular_font
            ws_resumo.cell(row=row_idx, column=2).border = thin_border
            ws_resumo.cell(row=row_idx, column=2).fill = white_fill
            row_idx += 1

        # Tabela Detalhada de Contagem
        ws_resumo.cell(row=row_idx+1, column=1, value="Métricas Detalhadas").font = section_font
        
        headers_diag = ["Indicador", "Quantidade", "Status Comercial"]
        diag_row = row_idx + 2

        for col_idx, h in enumerate(headers_diag, start=1):
            cell = ws_resumo.cell(row=diag_row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        diag_row += 1
        
        if mode == "classic":
            counts = [
                ("Total no Extrato Bancário", metrics["total_extract"], "Origem"),
                ("Total na Planilha de Vendas", metrics["total_sales"], "Origem"),
                ("Transações Conciliadas", metrics["matched"], "Sucesso"),
                ("Pendentes (Sem Venda Correspondente)", metrics["unmatched_extract"], "Atenção"),
                ("Pendentes (Sem Recebimento)", metrics["unmatched_sales"], "Atenção")
            ]
        else:
            counts = [
                ("Total de Notas Fiscais (XML)", metrics["total_nfe"], "Origem"),
                ("Total no Extrato Bancário", metrics["total_bank"], "Origem"),
                ("Notas Conciliadas", metrics["matched"], "Sucesso"),
                ("Notas Pendentes (Sem Pagamento)", metrics["unmatched_nfe"], "Atenção"),
                ("Extratos sem Nota Fiscal vinculada", metrics["unmatched_bank"], "Atenção")
            ]

        for item_name, val, item_status in counts:
            ws_resumo.cell(row=diag_row, column=1, value=item_name).font = regular_font
            ws_resumo.cell(row=diag_row, column=1).border = thin_border

            val_cell = ws_resumo.cell(row=diag_row, column=2, value=val)
            val_cell.font = bold_font
            val_cell.alignment = Alignment(horizontal="right")
            val_cell.border = thin_border

            status_cell = ws_resumo.cell(row=diag_row, column=3, value=item_status)
            status_cell.font = bold_font
            status_cell.alignment = Alignment(horizontal="center")
            status_cell.border = thin_border

            if item_status == "Sucesso":
                status_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            elif item_status == "Atenção":
                status_cell.fill = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            diag_row += 1

        # ----------------- ABA 2: DADOS CONCILIADOS -----------------
        ws_dados = wb.create_sheet(title="Planilha Conciliada")
        ws_dados.views.sheetView[0].showGridLines = True
        ws_dados.freeze_panes = "A2"

        # Escrever Headers dos Dados
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws_dados.cell(row=1, column=col_idx, value=str(col_name).upper())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border

        # Escrever Registros
        row_idx = 2
        for r in df.values:
            row_fill = zebra_fill if row_idx % 2 == 0 else white_fill

            for col_idx, val in enumerate(r, start=1):
                cell = ws_dados.cell(row=row_idx, column=col_idx)
                
                if pd.isna(val):
                    cell.value = ""
                else:
                    cell.value = val

                cell.font = regular_font
                cell.fill = row_fill
                cell.border = thin_border

                # Formatações Específicas
                col_name = df.columns[col_idx-1].lower()
                if isinstance(val, (int, float)):
                    if any(x in col_name for x in ["valor", "amount", "pago", "nfe"]):
                        cell.number_format = "R$ #,##0.00"
                    else:
                        cell.number_format = "#,##0"
                elif isinstance(val, datetime) or hasattr(val, "strftime"):
                    cell.number_format = "yyyy-mm-dd"

            row_idx += 1

        # Ajuste Automático das Larguras das Colunas
        for ws in [ws_resumo, ws_dados]:
            for col in ws.columns:
                max_len = 0
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        wb.save(output_path)