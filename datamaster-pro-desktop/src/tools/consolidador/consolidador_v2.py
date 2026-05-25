"""
Consolidador v3.0 Pro - Otimizado para máxima eficiência, multi-formato e visual premium
Une múltiplas planilhas em estrutura única com performance extrema e estética de ponta.
Suporta XLSX, XLS, CSV, TXT (delimitado por tabulação) e JSON.
"""
import pandas as pd
import numpy as np
from typing import List, Dict, Optional, Any
from pathlib import Path
from datetime import datetime
import os

# Imports openpyxl para formatação avançada
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


class Consolidador:
    """Motor profissional de consolidação de arquivos de dados de qualquer área"""
    
    FORMATS = {".xlsx", ".xls", ".csv", ".txt", ".json"}
    
    # Paletas de cores elegantes para exportação profissional do Excel
    THEMES = {
        "classic_blue": {
            "header_fill": "1F4E79",        # Azul Escuro Clássico
            "header_font_color": "FFFFFF",
            "zebra_fill": "F2F5F8",         # Azul Muito Claro
            "border_color": "D9D9D9",
            "accent_fill": "DDEBF7",
            "summary_accent": "1F4E79"
        },
        "emerald_green": {
            "header_fill": "1E4620",        # Verde Floresta
            "header_font_color": "FFFFFF",
            "zebra_fill": "F4F9F4",         # Verde Muito Claro
            "border_color": "D9D9D9",
            "accent_fill": "E2EFDA",
            "summary_accent": "1E4620"
        },
        "modern_orange": {
            "header_fill": "262626",        # Cinza Escuro Grafite
            "header_font_color": "FFFFFF",
            "zebra_fill": "FFF2E6",         # Laranja Muito Claro
            "border_color": "D9D9D9",
            "accent_fill": "FCE4D6",
            "summary_accent": "E26B0A"      # Laranja Queimado
        },
        "slate_gray": {
            "header_fill": "404040",        # Slate Gray Escuro
            "header_font_color": "FFFFFF",
            "zebra_fill": "F2F2F2",         # Cinza Muito Claro
            "border_color": "D9D9D9",
            "accent_fill": "EAEAEA",
            "summary_accent": "595959"
        }
    }
    
    def consolidate(
        self, 
        input_files: List[str], 
        output_path: str, 
        merge_strategy: str = "concat", 
        max_rows: Optional[int] = None,
        sheet_selection: str = "first",      # 'first', 'all', ou nome específico da aba
        enable_fuzzy_mapping: bool = True,
        fuzzy_threshold: int = 80,
        join_key: Optional[str] = None,
        join_type: str = "left",             # 'left', 'inner', 'right', 'outer'
        visual_theme: str = "classic_blue",
        remove_duplicates: bool = False
    ) -> Dict:
        """Consolida múltiplos arquivos de dados em um único Excel Estilizado Premium
        
        Args:
            input_files: Lista de caminhos de arquivo (XLSX, XLS, CSV, TXT, JSON)
            output_path: Caminho do arquivo de saída Excel (.xlsx)
            merge_strategy: 'concat' (vertical), 'merge' (horizontal por índice) ou 'join' (por chave)
            max_rows: Limite de linhas (None = ilimitado)
            sheet_selection: 'first' (primeira aba), 'all' (todas) ou nome específico
            enable_fuzzy_mapping: Se True, alinha colunas com nomes parecidos
            fuzzy_threshold: Similaridade mínima (0-100) para mapeamento fuzzy
            join_key: Coluna chave para merge_strategy='join'
            join_type: Tipo de junção se merge_strategy='join'
            visual_theme: Tema estético do Excel ('classic_blue', 'emerald_green', 'modern_orange', 'slate_gray')
            remove_duplicates: Se True, elimina registros idênticos na consolidação
        
        Returns:
            {success: bool, total_rows: int, total_files: int, output_path: str, error?: str}
        """
        if not input_files:
            return {"success": False, "error": "Nenhum arquivo de entrada selecionado."}
        
        dataframes = []
        rows_added = 0
        file_diagnostics = []
        
        for file_path in input_files:
            path = Path(file_path)
            
            if not path.exists():
                file_diagnostics.append({"file": path.name, "status": "Erro", "details": "Arquivo não existe"})
                continue
            
            if max_rows and rows_added >= max_rows:
                break
            
            suffix = path.suffix.lower()
            if suffix not in self.FORMATS:
                file_diagnostics.append({"file": path.name, "status": "Ignorado", "details": f"Formato '{suffix}' não suportado"})
                continue
            
            try:
                temp_dfs = []
                loaded_sheets = []
                
                # Leitura conforme extensão
                if suffix in {".xlsx", ".xls"}:
                    xl = pd.ExcelFile(path)
                    sheet_names = xl.sheet_names
                    
                    if sheet_selection == "all":
                        loaded_sheets = sheet_names
                    elif sheet_selection == "first":
                        loaded_sheets = [sheet_names[0]]
                    else:
                        if sheet_selection in sheet_names:
                            loaded_sheets = [sheet_selection]
                        else:
                            # Tentar fuzzy match de abas
                            from fuzzywuzzy import process
                            matched_sheet, score = process.extractOne(sheet_selection, sheet_names)
                            if score >= 80:
                                loaded_sheets = [matched_sheet]
                            else:
                                loaded_sheets = [sheet_names[0]]  # Fallback primeira aba
                                
                    for sheet in loaded_sheets:
                        df = pd.read_excel(xl, sheet_name=sheet)
                        df["_source_sheet"] = sheet
                        temp_dfs.append(df)
                        
                elif suffix == ".csv":
                    # Tentar diferentes encodings
                    for encoding in ["utf-8", "latin-1", "iso-8859-1", "cp1252"]:
                        try:
                            df = pd.read_csv(path, encoding=encoding)
                            temp_dfs.append(df)
                            loaded_sheets = ["default"]
                            break
                        except Exception:
                            continue
                            
                elif suffix == ".txt":
                    # Geralmente delimitado por Tabulação (TSV)
                    for encoding in ["utf-8", "latin-1", "cp1252"]:
                        try:
                            df = pd.read_csv(path, sep="\t", encoding=encoding)
                            temp_dfs.append(df)
                            loaded_sheets = ["default"]
                            break
                        except Exception:
                            continue
                            
                elif suffix == ".json":
                    df = pd.read_json(path)
                    temp_dfs.append(df)
                    loaded_sheets = ["default"]
                
                # Processar os dataframes deste arquivo
                for df in temp_dfs:
                    if df.empty:
                        continue
                    
                    # Limpeza de colunas sem nome (Unamed)
                    df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
                    
                    # Respeitar limite de linhas
                    if max_rows:
                        remaining = max_rows - rows_added
                        if len(df) > remaining:
                            df = df.head(remaining)
                    
                    # Rastreamento
                    df["_source_file"] = path.name
                    
                    # Harmonização de tipos
                    df = self._clean_and_harmonize(df)
                    
                    dataframes.append(df)
                    rows_added += len(df)
                
                file_diagnostics.append({
                    "file": path.name, 
                    "status": "Sucesso", 
                    "details": f"Abas: {', '.join(loaded_sheets)} | Linhas: {sum(len(d) for d in temp_dfs)}"
                })
                
            except Exception as e:
                file_diagnostics.append({"file": path.name, "status": "Erro", "details": str(e)})
                continue
        
        if not dataframes:
            return {"success": False, "error": "Nenhum dado válido pôde ser lido dos arquivos fornecidos."}
        
        try:
            # 1. Alinhamento de colunas difusas (Fuzzy Mapping) se ativado
            if enable_fuzzy_mapping and len(dataframes) > 1 and merge_strategy == "concat":
                dataframes = self._align_headers(dataframes, fuzzy_threshold)
            
            # 2. Consolidação baseada na estratégia
            if merge_strategy == "concat":
                result = pd.concat(dataframes, ignore_index=True)
                
            elif merge_strategy == "merge":
                # Mescla lateral simples por índice
                result = dataframes[0]
                for next_df in dataframes[1:]:
                    # Renomear colunas sobrepostas para evitar colisão
                    overlap = set(result.columns).intersection(set(next_df.columns)) - {"_source_file", "_source_sheet"}
                    rename_dict = {col: f"{col}_{next_df['_source_file'].iloc[0]}" for col in overlap}
                    next_df = next_df.rename(columns=rename_dict)
                    result = pd.merge(result, next_df, left_index=True, right_index=True, how="outer")
                    
            elif merge_strategy == "join":
                if not join_key:
                    return {"success": False, "error": "A coluna chave (Join Key) deve ser informada para cruzar planilhas."}
                
                # Verificar se a chave existe no primeiro dataframe
                result = dataframes[0]
                
                # Normalizar coluna chave para string antes do cruzamento para evitar mismatch
                if join_key in result.columns:
                    result[join_key] = result[join_key].astype(str).str.strip().str.lower()
                else:
                    return {"success": False, "error": f"Coluna chave '{join_key}' não encontrada no primeiro arquivo."}
                
                for idx, next_df in enumerate(dataframes[1:], start=1):
                    if join_key not in next_df.columns:
                        # Tentar Fuzzy Match da chave de cruzamento no próximo dataframe
                        from fuzzywuzzy import process
                        matched_key, score = process.extractOne(join_key, next_df.columns)
                        if score >= 80:
                            actual_key = matched_key
                        else:
                            # Pular arquivo sem a chave
                            file_diagnostics[idx]["details"] += f" (Chave '{join_key}' não encontrada)"
                            continue
                    else:
                        actual_key = join_key
                    
                    next_df_cleaned = next_df.copy()
                    next_df_cleaned[join_key] = next_df_cleaned[actual_key].astype(str).str.strip().str.lower()
                    if actual_key != join_key:
                        next_df_cleaned = next_df_cleaned.drop(columns=[actual_key])
                        
                    # Evitar colisão de colunas (exceto a chave e metadados)
                    overlap = set(result.columns).intersection(set(next_df_cleaned.columns)) - {join_key, "_source_file", "_source_sheet"}
                    rename_dict = {col: f"{col}_{next_df_cleaned['_source_file'].iloc[0]}" for col in overlap}
                    next_df_cleaned = next_df_cleaned.rename(columns=rename_dict)
                    
                    result = pd.merge(result, next_df_cleaned, on=join_key, how=join_type)
            
            # Remover duplicatas se ativado
            initial_rows = len(result)
            duplicates_removed = 0
            if remove_duplicates:
                # Manter colunas de metadados fora da verificação de duplicidade real do negócio
                subset_cols = [c for c in result.columns if c not in {"_source_file", "_source_sheet"}]
                result = result.drop_duplicates(subset=subset_cols)
                duplicates_removed = initial_rows - len(result)
            
            # 3. Exportar usando openpyxl com Formatação Comercial Premium (WOW Factor)
            self._save_premium_excel(result, output_path, visual_theme, file_diagnostics, duplicates_removed)
            
            return {
                "success": True,
                "total_rows": len(result),
                "total_files": len(dataframes),
                "output_path": output_path,
                "duplicates_removed": duplicates_removed
            }
            
        except Exception as e:
            return {"success": False, "error": f"Erro na consolidação de dados: {str(e)}"}
            
    def preview(self, file_path: str, rows: int = 5) -> Optional[pd.DataFrame]:
        """Retorna visualização rápida da planilha"""
        path = Path(file_path)
        try:
            suffix = path.suffix.lower()
            if suffix in {".xlsx", ".xls"}:
                return pd.read_excel(path, nrows=rows)
            elif suffix == ".csv":
                return pd.read_csv(path, nrows=rows, encoding="utf-8")
            elif suffix == ".txt":
                return pd.read_csv(path, sep="\t", nrows=rows, encoding="utf-8")
            elif suffix == ".json":
                return pd.read_json(path).head(rows)
        except Exception:
            # Fallback com latin-1 se utf-8 falhar
            try:
                if suffix == ".csv":
                    return pd.read_csv(path, nrows=rows, encoding="latin-1")
                elif suffix == ".txt":
                    return pd.read_csv(path, sep="\t", nrows=rows, encoding="latin-1")
            except Exception:
                return None
        return None
        
    def get_preview(self, file_path: str, max_rows: int = 5) -> Optional[pd.DataFrame]:
        try:
            return self.preview(file_path, rows=max_rows)
        except Exception:
            return None

        
    def _clean_and_harmonize(self, df: pd.DataFrame) -> pd.DataFrame:
        """Limpa dados inconsistentes e harmoniza tipos comuns (moeda, data, número)"""
        df = df.copy()
        for col in df.columns:
            if col in {"_source_file", "_source_sheet"}:
                continue
                
            # Converter colunas textuais que representam dinheiro ou números com vírgula
            if df[col].dtype == 'object':
                sample = df[col].dropna().head(50).astype(str)
                if sample.empty:
                    continue
                    
                # Checar se se parece com número/moeda brasileiro (ex: 1.234,56 ou R$ 10,00)
                digit_count = sum(1 for val in sample if any(c.isdigit() for c in val))
                currency_count = sum(1 for val in sample if any(sym in val for sym in ["R$", "$", "%"]))
                
                if digit_count / len(sample) >= 0.6:
                    try:
                        # Limpeza completa de caracteres não-numéricos preservando pontos/vírgulas decimais
                        cleaned = df[col].astype(str)
                        # Remover cifrões, espaços e pontos de milhar
                        cleaned = cleaned.str.replace(r'[R\$\$\s\.]', '', regex=True)
                        # Substituir vírgula decimal por ponto
                        cleaned = cleaned.str.replace(',', '.', regex=False)
                        # Tratar porcentagens
                        is_percent = cleaned.str.contains('%', regex=False)
                        cleaned = cleaned.str.replace('%', '', regex=False)
                        
                        numeric_series = pd.to_numeric(cleaned, errors='coerce')
                        
                        # Dividir por 100 se for percentual
                        numeric_series = np.where(is_percent, numeric_series / 100.0, numeric_series)
                        
                        # Se conseguimos converter pelo menos 75% dos valores válidos, aplicamos
                        if numeric_series.notna().sum() >= df[col].notna().sum() * 0.75:
                            df[col] = numeric_series
                    except Exception:
                        pass
            
            # Converter datas
            if df[col].dtype == 'object':
                sample = df[col].dropna().head(20).astype(str)
                if not sample.empty:
                    # Checar padrão comum de data (DD/MM/AAAA ou AAAA-MM-DD)
                    date_pat = r'^\d{1,4}[-/]\d{1,2}[-/]\d{1,4}'
                    date_match_count = sum(1 for val in sample if pd.Series([val]).str.match(date_pat).iloc[0])
                    
                    if date_match_count / len(sample) >= 0.7:
                        try:
                            # Forçar conversão limpa de data
                            df[col] = pd.to_datetime(df[col], errors='coerce', dayfirst=True)
                        except Exception:
                            pass
        return df

    def _align_headers(self, dataframes: List[pd.DataFrame], threshold: int) -> List[pd.DataFrame]:
        """Mapeamento inteligente de cabeçalhos (Fuzzy mapping)"""
        from fuzzywuzzy import fuzz
        
        # O primeiro arquivo dita as colunas padrão iniciais
        standard_headers = list(dataframes[0].columns)
        aligned_dfs = [dataframes[0].copy()]
        
        for df in dataframes[1:]:
            df_copy = df.copy()
            rename_map = {}
            
            for col in df_copy.columns:
                if col in {"_source_file", "_source_sheet"}:
                    continue
                if col in standard_headers:
                    continue
                
                # Encontrar coluna padrão com maior similaridade
                best_match = None
                best_score = 0
                for std_col in standard_headers:
                    if std_col in {"_source_file", "_source_sheet"}:
                        continue
                    # Calcular métrica de similaridade (Fuzzy)
                    score = fuzz.token_sort_ratio(str(col).lower(), str(std_col).lower())
                    if score > best_score:
                        best_score = score
                        best_match = std_col
                
                # Se for maior que o limiar configurado pelo usuário, alinha
                if best_score >= threshold and best_match:
                    rename_map[col] = best_match
                else:
                    # Adiciona nova coluna mapeada ao padrão para arquivos futuros alinharem com ela também
                    standard_headers.append(col)
                    
            if rename_map:
                df_copy = df_copy.rename(columns=rename_map)
            aligned_dfs.append(df_copy)
            
        return aligned_dfs

    def _save_premium_excel(
        self, 
        df: pd.DataFrame, 
        output_path: str, 
        theme_name: str, 
        diagnostics: List[Dict],
        duplicates_removed: int
    ):
        """Salva a planilha aplicando uma estética visual de ponta e painel de diagnóstico"""
        theme = self.THEMES.get(theme_name, self.THEMES["classic_blue"])
        wb = Workbook()
        
        # ----------------- ABA 1: RESUMO DA CONSOLIDAÇÃO -----------------
        ws_resumo = wb.active
        ws_resumo.title = "📊 Resumo"
        ws_resumo.views.sheetView[0].showGridLines = True
        
        # Fontes e Cores do Resumo
        title_font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
        section_font = Font(name="Segoe UI", size=12, bold=True, color=theme["header_fill"])
        header_font = Font(name="Segoe UI", size=10, bold=True, color=theme["header_font_color"])
        bold_font = Font(name="Segoe UI", size=10, bold=True)
        regular_font = Font(name="Segoe UI", size=10)
        
        header_fill = PatternFill(start_color=theme["header_fill"], end_color=theme["header_fill"], fill_type="solid")
        accent_fill = PatternFill(start_color=theme["accent_fill"], end_color=theme["accent_fill"], fill_type="solid")
        zebra_fill = PatternFill(start_color=theme["zebra_fill"], end_color=theme["zebra_fill"], fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color=theme["border_color"]),
            right=Side(style='thin', color=theme["border_color"]),
            top=Side(style='thin', color=theme["border_color"]),
            bottom=Side(style='thin', color=theme["border_color"])
        )
        
        # Cabeçalho Principal do Resumo
        ws_resumo.merge_cells("A1:D2")
        title_cell = ws_resumo["A1"]
        title_cell.value = "  RELATÓRIO DE CONSOLIDAÇÃO PREMIUM"
        title_cell.font = title_font
        title_cell.fill = header_fill
        title_cell.alignment = Alignment(vertical="center", horizontal="left")
        
        # Preencher fundo das células mescladas para consistência de cor
        for row in range(1, 3):
            for col in range(1, 5):
                ws_resumo.cell(row=row, column=col).fill = header_fill
                
        # Sessão 1: Informações Gerais
        ws_resumo["A4"] = "Estatísticas Gerais"
        ws_resumo["A4"].font = section_font
        
        stats = [
            ("Data da Execução", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
            ("Arquivos Processados", len([d for d in diagnostics if d["status"] == "Sucesso"])),
            ("Total de Registros Consolidados", len(df)),
            ("Total de Colunas Estruturadas", len([c for c in df.columns if c not in {"_source_file", "_source_sheet"}])),
            ("Duplicatas Removidas", duplicates_removed),
            ("Estilo Visual Aplicado", theme_name.replace("_", " ").title())
        ]
        
        row_idx = 5
        for key, val in stats:
            ws_resumo.cell(row=row_idx, column=1, value=key).font = bold_font
            ws_resumo.cell(row=row_idx, column=1).border = thin_border
            ws_resumo.cell(row=row_idx, column=1).fill = zebra_fill
            
            ws_resumo.cell(row=row_idx, column=2, value=val).font = regular_font
            ws_resumo.cell(row=row_idx, column=2).border = thin_border
            ws_resumo.cell(row=row_idx, column=2).fill = white_fill
            row_idx += 1
            
        # Sessão 2: Diagnóstico por Arquivo
        ws_resumo.cell(row=row_idx+1, column=1, value="Histórico de Importação").font = section_font
        
        headers_diag = ["Nome do Arquivo", "Status", "Informação / Log"]
        diag_row = row_idx + 2
        
        for col_idx, h in enumerate(headers_diag, start=1):
            cell = ws_resumo.cell(row=diag_row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            
        ws_resumo.merge_cells(start_row=diag_row, start_column=3, end_row=diag_row, end_column=4)
        
        diag_row += 1
        for item in diagnostics:
            ws_resumo.cell(row=diag_row, column=1, value=item["file"]).font = regular_font
            ws_resumo.cell(row=diag_row, column=1).border = thin_border
            
            status_cell = ws_resumo.cell(row=diag_row, column=2, value=item["status"])
            status_cell.font = bold_font
            status_cell.alignment = Alignment(horizontal="center")
            status_cell.border = thin_border
            
            # Cores de status
            if item["status"] == "Sucesso":
                status_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Verde Claro
            else:
                status_cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Vermelho/Laranja Claro
                
            ws_resumo.cell(row=diag_row, column=3, value=item["details"]).font = regular_font
            ws_resumo.cell(row=diag_row, column=3).border = thin_border
            ws_resumo.merge_cells(start_row=diag_row, start_column=3, end_row=diag_row, end_column=4)
            ws_resumo.cell(row=diag_row, column=4).border = thin_border
            diag_row += 1

        # ----------------- ABA 2: DADOS CONSOLIDADOS -----------------
        ws_dados = wb.create_sheet(title="Planilha Consolidada")
        ws_dados.views.sheetView[0].showGridLines = True
        ws_dados.freeze_panes = "A2" # Fixar primeira linha de cabeçalho
        
        # Colunas com nomes limpos no Excel final (mover _source_file e _source_sheet para o final)
        core_cols = [c for c in df.columns if c not in {"_source_file", "_source_sheet"}]
        meta_cols = [c for c in df.columns if c in {"_source_file", "_source_sheet"}]
        ordered_cols = core_cols + meta_cols
        
        df_ordered = df[ordered_cols]
        
        # Escrever Cabeçalho dos Dados
        for col_idx, col_name in enumerate(df_ordered.columns, start=1):
            # Deixar nome das colunas internas mais amigáveis no Excel
            display_name = col_name
            if col_name == "_source_file":
                display_name = "Arquivo Origem"
            elif col_name == "_source_sheet":
                display_name = "Aba Origem"
                
            cell = ws_dados.cell(row=1, column=col_idx, value=str(display_name).upper())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border
            
        # Escrever Registros dos Dados
        row_idx = 2
        for r in df_ordered.values:
            # Alternar cor de linha para efeito zebra
            row_fill = zebra_fill if row_idx % 2 == 0 else white_fill
            
            for col_idx, val in enumerate(r, start=1):
                cell = ws_dados.cell(row=row_idx, column=col_idx)
                
                # Tratar valores vazios ou NaN
                if pd.isna(val):
                    cell.value = ""
                else:
                    cell.value = val
                    
                cell.font = regular_font
                cell.fill = row_fill
                cell.border = thin_border
                
                # Formatação de Célula Conforme Tipo de Dado
                if isinstance(val, (int, float)):
                    # Verificar se parece moeda pelo nome da coluna
                    col_name = df_ordered.columns[col_idx-1].lower()
                    if any(term in col_name for term in ["valor", "preco", "preço", "total", "custo", "faturamento", "receita", "comissao", "comissão"]):
                        cell.number_format = "R$ #,##0.00"
                    elif any(term in col_name for term in ["percent", "%", "taxa", "margem"]):
                        # Se for taxa e estiver maior que 1, dividir por 100 para representação correta
                        if val > 1.0:
                            cell.value = val / 100.0
                        cell.number_format = "0.0%"
                    else:
                        cell.number_format = "#,##0"
                elif isinstance(val, datetime) or hasattr(val, "strftime"):
                    cell.number_format = "yyyy-mm-dd"
                    
            row_idx += 1
            
        # Redimensionamento Automático das Colunas
        for ws in [ws_resumo, ws_dados]:
            for col in ws.columns:
                max_len = 0
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
        # Salvar o workbook
        wb.save(output_path)
