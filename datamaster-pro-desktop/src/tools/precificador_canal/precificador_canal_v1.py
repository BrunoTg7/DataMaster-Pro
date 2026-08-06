"""
Precificador por Canal de Venda v1.2
Motor de cálculo reverso de preços para múltiplos marketplaces.
Garante a margem líquida desejada considerando: comissão, taxa fixa,
imposto Simples Nacional e frete dinâmico por peso.

Novidades v1.2:
- Integração API Melhor Envio para frete real por CEP
- Cálculo de ICMS interestadual por UF (Convênio ICMS 23/2021)
- Simulação what-if: cálculo de múltiplos cenários de margem
- Tabelas de frete atualizáveis via JSON

As taxas e tabelas de frete são carregadas de tax_rules.json.
Consulte tax_rules.example.json para o schema completo.
"""
import logging
import json
import os
import re
import pandas as pd
import httpx
from typing import Dict, Callable, Optional, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

log = logging.getLogger(__name__)

# Tabela de ICMS interestadual (Fonte: Convênio ICMS 23/2021)
# Origem\Destino → alíquota (%)
ICMS_INTERESTADUAL = {
    "SP": {"RJ": 12, "MG": 12, "ES": 7, "PR": 12, "SC": 12, "RS": 12, "BA": 12, "CE": 12, "PE": 12, "GO": 12, "DF": 12, "MT": 12, "MS": 12, "AM": 12, "PA": 12, "MA": 12, "PI": 12, "RN": 12, "PB": 12, "AL": 12, "SE": 12, "TO": 12, "AC": 12, "RO": 12, "RR": 12, "AP": 12},
    "RJ": {"SP": 12, "MG": 12, "ES": 7, "PR": 12, "SC": 12, "RS": 12, "BA": 12, "CE": 12, "PE": 12, "GO": 12, "DF": 12, "MT": 12, "MS": 12, "AM": 12, "PA": 12, "MA": 12, "PI": 12, "RN": 12, "PB": 12, "AL": 12, "SE": 12, "TO": 12, "AC": 12, "RO": 12, "RR": 12, "AP": 12},
    "MG": {"SP": 12, "RJ": 12, "ES": 7, "PR": 12, "SC": 12, "RS": 12, "BA": 12, "CE": 12, "PE": 12, "GO": 12, "DF": 12, "MT": 12, "MS": 12, "AM": 12, "PA": 12, "MA": 12, "PI": 12, "RN": 12, "PB": 12, "AL": 12, "SE": 12, "TO": 12, "AC": 12, "RO": 12, "RR": 12, "AP": 12},
    "ES": {"SP": 12, "RJ": 12, "MG": 12, "PR": 12, "SC": 12, "RS": 12, "BA": 12, "CE": 12, "PE": 12, "GO": 12, "DF": 12, "MT": 12, "MS": 12, "AM": 12, "PA": 12, "MA": 12, "PI": 12, "RN": 12, "PB": 12, "AL": 12, "SE": 12, "TO": 12, "AC": 12, "RO": 12, "RR": 12, "AP": 12},
}


class PrecificadorCanal:
    """
    Calcula o preço de venda ideal por canal (ML, Shopee, Amazon, Magalu)
    para garantir a margem líquida desejada pelo lojista.
    """

    # Tabela de regras dos marketplaces (carregada dinamicamente via JSON)
    # Evita perdas financeiras caso as tarifas ou regras mudem.

    def __init__(self, log_callback: Callable = None, progress_callback: Callable = None):
        self.log_callback = log_callback
        self.progress_callback = progress_callback
        self.CANAIS = self._load_canais_config()

    def _load_canais_config(self) -> dict:
        """Carrega regras de taxas e tabelas de frete do JSON externo"""
        base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
        tax_file = os.path.join(base_dir, "tax_rules.json")
        example_file = os.path.join(base_dir, "tax_rules.example.json")
        
        target_file = tax_file if os.path.exists(tax_file) else example_file
        self._freight_tables = {}
        
        try:
            if os.path.exists(target_file):
                with open(target_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    fees = data.get("MARKETPLACE_FEES", {})
                    self._freight_tables = data.get("FREIGHT_TABLES", {})
                    
                    # Mapeamento nome_exibicao -> chave_json
                    mapa = {
                        "Mercado Livre": "mercadolivre",
                        "Shopee": "shopee",
                        "Amazon": "amazon",
                        "Magalu": "magalu"
                    }
                    canais_finais = {}
                    for nome_display, nome_json in mapa.items():
                        cfg = fees.get(nome_json, {})
                        canais_finais[nome_display] = {
                            "comissao_pct": cfg.get("percent", 0.15),
                            "taxa_fixa": cfg.get("fixed", 0.0),
                            "frete_gratis_min": cfg.get("frete_gratis_min", 0.0),
                            "custo_frete_medio": cfg.get("custo_frete_medio", 0.0),
                            "emoji": cfg.get("emoji", "🔗"),
                            "freight_table_key": cfg.get("freight_table_key", "mercado_envios")
                        }
                    self._log(f"Configurações de canais carregadas de: {os.path.basename(target_file)}")
                    return canais_finais
        except Exception as e:
            self._log(f"Erro ao carregar configurações de canais: {e}. Usando fallback.")
        
        # Fallback de segurança se falhar
        return {
            "Mercado Livre": {"comissao_pct": 0.16, "taxa_fixa": 6.00, "frete_gratis_min": 79.0, "custo_frete_medio": 15.0, "emoji": "🛒", "freight_table_key": "mercado_envios"},
            "Shopee": {"comissao_pct": 0.20, "taxa_fixa": 3.00, "frete_gratis_min": 0.0, "custo_frete_medio": 0.0, "emoji": "🍊", "freight_table_key": "shopee_envios"}
        }

    def _log(self, message: str):
        log.info(message)
        if self.log_callback:
            self.log_callback(message)

    def _progress(self, pct: int):
        if self.progress_callback:
            self.progress_callback(pct)

    # ------------------------------------------------------------------
    # ICMS INTERESTADUAL
    # ------------------------------------------------------------------
    def calcular_icms_interestadual(self, uf_origem: str, uf_destino: str, valor_produto: float) -> float:
        """Calcula ICMS interestadual com base na tabela Convênio ICMS 23/2021"""
        uf_origem = uf_origem.upper()
        uf_destino = uf_destino.upper()
        aliquota = ICMS_INTERESTADUAL.get(uf_origem, {}).get(uf_destino, 12)
        return valor_produto * (aliquota / 100)

    # ------------------------------------------------------------------
    # FRETE REAL VIA MELHOR ENVIO API
    # ------------------------------------------------------------------
    def calcular_frete_real(
        self,
        cep_origem: str,
        cep_destino: str,
        peso_g: float,
        altura_cm: float = 20,
        largura_cm: float = 15,
        comprimento_cm: float = 10,
        api_token: str = None
    ) -> Dict:
        """Calcula frete real via Melhor Envio API (gratuita para consultas)"""
        MELHOR_ENVIO_API = "https://www.melhorenvio.com.br/api/v2/me/shipment/calculate"
        
        token = api_token or os.environ.get("MELHOR_ENVIO_TOKEN", "")
        if not token:
            return {"error": "Token da API Melhor Envio não configurado. Configure MELHOR_ENVIO_TOKEN."}
        
        payload = {
            "from": {"postal_code": re.sub(r'\D', '', cep_origem)},
            "to": {"postal_code": re.sub(r'\D', '', cep_destino)},
            "package": {
                "height": altura_cm,
                "width": largura_cm,
                "length": comprimento_cm,
                "weight": peso_g / 1000
            }
        }
        headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
        
        try:
            resp = httpx.post(MELHOR_ENVIO_API, json=payload, headers=headers, timeout=15.0)
            if resp.status_code == 200:
                options = resp.json()
                if isinstance(options, list) and len(options) > 0:
                    cheapest = min(options, key=lambda x: float(x.get("price", 999)))
                    fastest = min(options, key=lambda x: float(x.get("delivery_time", 999)))
                    return {
                        "cheapest": {"carrier": cheapest.get("company", {}).get("name", ""), "price": float(cheapest.get("price", 0)), "days": cheapest.get("delivery_time", 0)},
                        "fastest": {"carrier": fastest.get("company", {}).get("name", ""), "price": float(fastest.get("price", 0)), "days": fastest.get("delivery_time", 0)},
                        "all_options": len(options),
                    }
            return {"error": f"Erro na API: {resp.status_code}"}
        except Exception as e:
            return {"error": f"Erro de conexão: {e}"}

    # ------------------------------------------------------------------
    # Cálculo Reverso: dado o custo + margem desejada, acha o preço de venda
    # ------------------------------------------------------------------
    def _calcular_frete_dinamico(self, canal_nome: str, peso_g: float, preco_venda_estimado: float, frete_min: float, canal_info: dict = None) -> float:
        """
        Calcula o frete embutido de forma dinâmica baseada no peso do produto.
        Utiliza tabelas de frete carregadas de tax_rules.json.
        
        PRÓXIMOS PASSOS (iteração futura):
        - Integrar API do Melhor Envio/Frenet para cálculo real por CEP
        - Adicionar seguro e peso cubagem
        """
        # Se não atinge a regra de frete grátis do canal, não embute frete
        if frete_min > 0 and preco_venda_estimado < frete_min:
            return 0.0

        # Shopee e Amazon: frete geralmente subsidiado pelo marketplace
        if canal_nome == "Shopee":
            return 0.0
        elif canal_nome == "Amazon":
            return 0.0  # FBA: custo embutido na comissão

        # Buscar tabela de frete adequada para o canal
        freight_key = "mercado_envios"
        if canal_info:
            freight_key = canal_info.get("freight_table_key", "mercado_envios")
        
        tabela = self._freight_tables.get(freight_key, [])
        
        # Procurar a faixa de peso correta na tabela
        for faixa in tabela:
            peso_max = faixa.get("peso_max_g", float("inf"))
            if peso_g <= peso_max:
                return faixa.get("valor", 0.0)
        
        # Se exceder todas as faixas, usar a última faixa (peso pesado)
        if tabela:
            return tabela[-1].get("valor", 50.0)
        
        # Fallback absoluto (tabela vazia)
        return 0.0

    def _preco_por_canal(
        self,
        canal_nome: str,
        custo: float,
        imposto_pct: float,
        margem_desejada_pct: float,
        peso_g: float,
        canal_info: dict,
    ) -> dict:
        """
        Fórmula de cálculo reverso dinâmico.
        """
        comissao = canal_info["comissao_pct"]
        taxa_fixa = canal_info["taxa_fixa"]
        
        # Inicia com o frete padrão estático para o cálculo base
        frete_embutido = canal_info.get("custo_frete_medio", 0.0)

        # Percentual total que "sai" do preço de venda
        desconto_total = comissao + (margem_desejada_pct / 100) + (imposto_pct / 100)

        if desconto_total >= 1.0:
            return {"preco": None, "margem_real": None, "erro": "Margem inviável"}

        # Preço de venda sugerido primário
        preco = (custo + taxa_fixa + frete_embutido) / (1 - desconto_total)

        # Atualiza para usar a lógica de frete dinâmico via peso
        frete_min = canal_info.get("frete_gratis_min", 0.0)
        
        # Verifica se aciona frete embutido
        if frete_min > 0 and preco >= frete_min:
            frete_embutido = self._calcular_frete_dinamico(canal_nome, peso_g, preco, frete_min, canal_info)
            preco = (custo + taxa_fixa + frete_embutido) / (1 - desconto_total)

        # Calcular margem real
        comissao_valor = preco * comissao
        imposto_valor = preco * (imposto_pct / 100)
        lucro_liquido = preco - custo - comissao_valor - taxa_fixa - frete_embutido - imposto_valor
        margem_real = (lucro_liquido / preco) * 100 if preco > 0 else 0

        return {
            "preco": round(preco, 2),
            "margem_real_pct": round(margem_real, 2),
            "lucro_liquido": round(lucro_liquido, 2),
            "comissao_valor": round(comissao_valor, 2),
            "taxa_fixa": round(taxa_fixa, 2),
            "frete_embutido": round(frete_embutido, 2),
            "imposto_valor": round(imposto_valor, 2),
            "erro": None,
        }

    # ------------------------------------------------------------------
    # Processamento em lote via planilha
    # ------------------------------------------------------------------
    def calcular_planilha(
        self,
        df: pd.DataFrame,
        margem_desejada_pct: float = 20.0,
        canais_selecionados: list = None,
        output_path: str = None,
    ) -> dict:
        """
        Processa um DataFrame com colunas:
          - produto (str) — nome do produto
          - custo (float) — custo de aquisição
          - imposto_pct (float) — % Simples Nacional (ex: 6.0 para 6%)

        Retorna dict com sucesso, DataFrame de resultados e caminho do Excel.
        """
        try:
            canais = canais_selecionados or list(self.CANAIS.keys())
            self._log(f"Iniciando cálculo para {len(df)} produto(s) em {len(canais)} canal(is)...")

            # Normalizar colunas
            df_norm = self._normalizar_df(df)
            if df_norm is None:
                return {"success": False, "error": "Colunas obrigatórias não encontradas. Use: produto, custo, imposto_pct"}

            results = []
            total = len(df_norm)
            for i, row in df_norm.iterrows():
                produto = str(row.get("produto", f"Produto {i+1}"))
                custo = float(row.get("custo", 0))
                imposto_pct = float(row.get("imposto_pct", 6.0))
                peso_g = float(row.get("peso_g", 500.0))

                result_row = {
                    "Produto": produto,
                    "Custo (R$)": custo,
                    "Peso (g)": peso_g,
                    "Imposto Simples (%)": imposto_pct,
                    "Margem Desejada (%)": margem_desejada_pct,
                }

                for canal_nome in canais:
                    info = self.CANAIS.get(canal_nome)
                    if not info:
                        continue
                    calc = self._preco_por_canal(canal_nome, custo, imposto_pct, margem_desejada_pct, peso_g, info)
                    emoji = info["emoji"]
                    if calc["erro"]:
                        result_row[f"{emoji} {canal_nome} - Preço"] = "INVIÁVEL"
                        result_row[f"{emoji} {canal_nome} - Margem Real (%)"] = "-"
                        result_row[f"{emoji} {canal_nome} - Lucro Líquido (R$)"] = "-"
                    else:
                        result_row[f"{emoji} {canal_nome} - Preço"] = calc["preco"]
                        result_row[f"{emoji} {canal_nome} - Margem Real (%)"] = calc["margem_real_pct"]
                        result_row[f"{emoji} {canal_nome} - Lucro Líquido (R$)"] = calc["lucro_liquido"]

                results.append(result_row)
                self._progress(int(((i + 1) / total) * 80))

            result_df = pd.DataFrame(results)
            self._log(f"{len(result_df)} produtos processados.")

            # Gerar Excel estilizado
            if not output_path:
                from datetime import datetime
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                import os, sys
                project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
                if project_root not in sys.path:
                    sys.path.insert(0, project_root)
                import config
                output_path = os.path.join(config.OUTPUT_DIR, f"precificacao_{ts}.xlsx")

            self._save_excel(result_df, output_path, margem_desejada_pct, canais)
            self._progress(100)

            self._log(f"Relatório salvo em: {output_path}")
            return {
                "success": True,
                "rows": len(result_df),
                "output_path": output_path,
                "dataframe": result_df,
            }

        except Exception as e:
            return {"success": False, "error": str(e)}

    def _normalizar_df(self, df: pd.DataFrame) -> Optional[pd.DataFrame]:
        """Normaliza nomes de colunas para o padrão esperado."""
        col_map = {}
        for col in df.columns:
            col_lower = col.lower().strip()
            if col_lower in ["produto", "descricao", "nome", "item", "description"]:
                col_map[col] = "produto"
            elif col_lower in ["custo", "custo_aquisicao", "preco_custo", "cost", "valor_custo"]:
                col_map[col] = "custo"
            elif col_lower in ["imposto", "imposto_pct", "simples", "simples_nacional", "aliquota", "tax_pct"]:
                col_map[col] = "imposto_pct"
            elif col_lower in ["peso", "peso_g", "peso (g)", "weight"]:
                col_map[col] = "peso_g"

        df_norm = df.rename(columns=col_map)

        if "produto" not in df_norm.columns:
            df_norm["produto"] = [f"Produto {i+1}" for i in range(len(df_norm))]
        if "custo" not in df_norm.columns:
            return None
        if "imposto_pct" not in df_norm.columns:
            df_norm["imposto_pct"] = 6.0  # Simples Nacional padrão
        if "peso_g" not in df_norm.columns:
            df_norm["peso_g"] = 500.0  # Peso padrão de 500g

        return df_norm[["produto", "custo", "imposto_pct", "peso_g"]]

    # ------------------------------------------------------------------
    # Cálculo manual (produto único)
    # ------------------------------------------------------------------
    def calcular_produto_unico(
        self,
        produto: str,
        custo: float,
        imposto_pct: float,
        margem_desejada_pct: float,
        canais_selecionados: list = None,
    ) -> dict:
        """Para cálculo rápido de um único produto sem planilha."""
        df = pd.DataFrame([{"produto": produto, "custo": custo, "imposto_pct": imposto_pct}])
        return self.calcular_planilha(df, margem_desejada_pct, canais_selecionados)

    # ------------------------------------------------------------------
    # SIMULAÇÃO WHAT-IF
    # ------------------------------------------------------------------
    def simular_cenarios(
        self,
        custo_produto: float,
        cenarios: List[Dict],
        imposto_pct: float = 6.0,
    ) -> pd.DataFrame:
        """Simula múltiplos cenários de precificação com variações de margem e canal
        
        Args:
            custo_produto: Custo de aquisição do produto
            cenarios: Lista de dicts com {nome, margem, canal, peso_g}
            imposto_pct: Alíquota Simples Nacional
            
        Returns:
            DataFrame com resultados comparativos
        """
        resultados = []
        
        for cenario in cenarios:
            margem = cenario.get("margem", 30.0)
            canal = cenario.get("canal", "Mercado Livre")
            peso = cenario.get("peso_g", 500)
            nome = cenario.get("nome", f"Margem {margem:.0f}%")
            
            info = self.CANAIS.get(canal, {})
            if not info:
                continue
            
            calc = self._preco_por_canal(canal, custo_produto, imposto_pct, margem, peso, info)
            
            if calc["erro"]:
                resultados.append({
                    "Cenário": nome,
                    "Canal": canal,
                    "Margem Desejada": f"{margem:.1f}%",
                    "Preço de Venda": "INVIÁVEL",
                    "Lucro Líquido": "-",
                    "Margem Real": "-",
                })
            else:
                resultados.append({
                    "Cenário": nome,
                    "Canal": canal,
                    "Margem Desejada": f"{margem:.1f}%",
                    "Preço de Venda": f"R$ {calc['preco']:.2f}",
                    "Lucro Líquido": f"R$ {calc['lucro_liquido']:.2f}",
                    "Margem Real": f"{calc['margem_real_pct']:.1f}%",
                })
        
        return pd.DataFrame(resultados)

    # ------------------------------------------------------------------
    # Exportação Excel Premium
    # ------------------------------------------------------------------
    def _save_excel(self, df: pd.DataFrame, output_path: str, margem: float, canais: list):
        wb = Workbook()
        ws = wb.active
        ws.title = "Precificação por Canal"

        # Paleta de cores
        HEADER_BG = "1C1C1E"
        HEADER_FG = "FFFFFF"
        SUB_BG = "2C2C2E"
        SUB_FG = "F5A623"
        ZEBRA = "F9F9F9"
        BORDER_COLOR = "D1D1D6"
        GREEN = "D4EDDA"
        ORANGE = "FFF3CD"
        RED = "F8D7DA"

        thin = Side(style="thin", color=BORDER_COLOR)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # === CABEÇALHO DA PLANILHA ===
        try:
            ws.merge_cells("A1:C1")
        except Exception:
            pass
        title_cell = ws["A1"]
        title_cell.value = f"📊 DataMaster Pro — Simulador de Precificação por Canal"
        title_cell.font = Font(name="Calibri", size=14, bold=True, color=HEADER_FG)
        title_cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        try:
            ws.merge_cells(f"D1:{get_column_letter(len(df.columns))}1")
        except Exception:
            pass
        meta_cell = ws["D1"]
        meta_cell.value = f"Margem Desejada: {margem}%  |  Canais: {', '.join(canais)}  |  Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        meta_cell.font = Font(name="Calibri", size=10, color="A1A1AA")
        meta_cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        meta_cell.alignment = Alignment(horizontal="right", vertical="center")

        # === LINHA DE CABEÇALHO DAS COLUNAS ===
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.font = Font(name="Calibri", size=10, bold=True, color=SUB_FG)
            cell.fill = PatternFill("solid", fgColor=SUB_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[2].height = 40

        # === DADOS ===
        for row_idx, (_, row) in enumerate(df.iterrows(), 3):
            bg = "FFFFFF" if row_idx % 2 == 1 else ZEBRA
            for col_idx, (col_name, value) in enumerate(row.items(), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Calibri", size=10)

                # Colorir lucro
                if "Lucro" in col_name and isinstance(value, (int, float)):
                    if value > 0:
                        cell.fill = PatternFill("solid", fgColor=GREEN)
                    elif value < 0:
                        cell.fill = PatternFill("solid", fgColor=RED)
                    else:
                        cell.fill = PatternFill("solid", fgColor=ORANGE)
                elif value == "INVIÁVEL":
                    cell.fill = PatternFill("solid", fgColor=RED)
                    cell.font = Font(name="Calibri", size=10, bold=True, color="721C24")
                else:
                    cell.fill = PatternFill("solid", fgColor=bg)

        # Ajuste de largura das colunas
        for col_idx, col_name in enumerate(df.columns, 1):
            max_len = max(len(str(col_name)), 12)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)

        # Congelar cabeçalho
        ws.freeze_panes = "A3"

        wb.save(output_path)
