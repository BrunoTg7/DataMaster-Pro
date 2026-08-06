"""
Pipeline ETL NCM/CEST Enterprise - Fonte Oficial Receita Federal
Baixa, parseia, valida e publica base NCM (TIPI) + CEST (Convênio ICMS 92/2015)
Gera: ncm_database.json (para ClassificadorNCM) + manifest.json (auditoria)
"""

import requests
import pandas as pd
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Dict, List, Optional, Any
from datetime import datetime
import logging
import json
import hashlib
from dataclasses import dataclass, asdict
import re
import os

log = logging.getLogger(__name__)


@dataclass
class NCMRecord:
    codigo: str
    descricao: str
    aliq_ipi: float
    cest: Optional[str]
    cest_descricao: str
    unidade: str
    vigencia_inicio: str
    vigencia_fim: Optional[str]
    hash_registro: str


class NCMPipeline:
    """
    Pipeline ETL completo para base NCM/CEST oficial.
    Fonte TIPI: Receita Federal (site gov.br)
    Fonte CEST: Portal SPED / Convênio ICMS 92/2015
    """
    
    # URLs base (podem mudar - usar arquivos locais versionados como fallback)
    TIPI_PAGE_URL = "https://www.gov.br/receitafederal/pt-br/assuntos/aduana-e-comercio-exterior/manuais/tipi"
    CEST_PAGE_URL = "https://www.gov.br/receitafederal/pt-br/assuntos/tributacao/convencoes-e-acordos/convencoes-icms/cest"
    
    # Arquivos locais versionados (commit no repo)
    LOCAL_TIPI_XLSX = "data/raw/tipi_vigente.xlsx"
    LOCAL_CEST_XLSX = "data/raw/cest_vigente.xlsx"
    
    def __init__(self, output_dir: str = "data/processed"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.raw_dir = Path("data/raw")
        self.raw_dir.mkdir(parents=True, exist_ok=True)
        self.records: List[NCMRecord] = []
    
    def run_full_pipeline(self) -> Dict:
        """Executa pipeline completo: download → parse → merge → validate → publish"""
        log.info("=== INICIANDO PIPELINE NCM/CEST ===")
        
        # 1. Tentar download oficial (pode falhar se URL mudou)
        tipi_df = self._download_and_parse_tipi()
        cest_df = self._download_and_parse_cest()
        
        # 2. Fallback para arquivos locais versionados
        if tipi_df is None:
            log.warning("Download TIPI falhou - usando arquivo local versionado")
            tipi_df = self._parse_local_tipi()
        if cest_df is None:
            log.warning("Download CEST falhou - usando arquivo local versionado")
            cest_df = self._parse_local_cest()
        
        if tipi_df is None:
            raise RuntimeError("TIPI não disponível (nem download nem arquivo local). "
                              "Baixe manualmente do site da RFB e salve em data/raw/tipi_vigente.xlsx")
        
        # 3. Merge TIPI + CEST
        merged = self._merge_tipi_cest(tipi_df, cest_df)
        
        # 4. Validar integridade
        validation = self._validate(merged)
        if not validation["valid"]:
            raise RuntimeError(f"Validação falhou: {validation['errors']}")
        
        # 5. Gerar registros finais
        self.records = self._build_records(merged)
        
        # 6. Publicar
        self._publish()
        
        result = {
            "success": True,
            "total_ncm": len(self.records),
            "with_cest": sum(1 for r in self.records if r.cest),
            "version": datetime.now().strftime("%Y.%m.%d"),
            "validation": validation,
            "generated_at": datetime.now().isoformat()
        }
        
        log.info(f"Pipeline concluído: {result['total_ncm']} NCMs, {result['with_cest']} com CEST")
        return result
    
    # ============================================================
    # DOWNLOAD E PARSE TIPI
    # ============================================================
    
    def _download_and_parse_tipi(self) -> Optional[pd.DataFrame]:
        """Tenta baixar TIPI do site da RFB"""
        try:
            resp = requests.get(self.TIPI_PAGE_URL, timeout=30)
            resp.raise_for_status()
            
            # Procurar link do XLSX vigente
            xlsx_links = re.findall(r'href="([^"]*\.xlsx)"', resp.text)
            tipi_link = next((l for l in xlsx_links if "tipi" in l.lower() and "vigente" in l.lower()), None)
            
            if not tipi_link:
                log.warning("Link TIPI vigente não encontrado na página")
                return None
            
            if not tipi_link.startswith("http"):
                from urllib.parse import urljoin
                tipi_link = urljoin(self.TIPI_PAGE_URL, tipi_link)
            
            # Baixar e parsear
            xlsx_resp = requests.get(tipi_link, timeout=60)
            xlsx_resp.raise_for_status()
            
            # Salvar cópia local
            local_path = self.raw_dir / "tipi_vigente.xlsx"
            local_path.write_bytes(xlsx_resp.content)
            
            return self._parse_tipi_xlsx(xlsx_resp.content)
        except Exception as e:
            log.warning(f"Download TIPI falhou: {e}")
            return None
    
    def _parse_tipi_xlsx(self, content: bytes) -> pd.DataFrame:
        """Parseia XLSX do TIPI (layout oficial Receita Federal)"""
        df = pd.read_excel(content, header=None, dtype=str)
        
        # Detectar linha de cabeçalho
        header_row = None
        for i, row in df.iterrows():
            row_str = " ".join(row.dropna().astype(str)).upper()
            if "CAPÍTULO" in row_str or "CÓDIGO" in row_str or "NBM" in row_str:
                header_row = i
                break
        
        if header_row is None:
            # Tentar linha 0 como fallback
            header_row = 0
        
        df = pd.read_excel(content, header=header_row, dtype=str)
        
        # Normalizar colunas
        col_map = {}
        for col in df.columns:
            c = str(col).upper().strip()
            if "CAPÍTULO" in c: col_map[col] = "capitulo"
            elif "POSIÇÃO" in c or "POSICAO" in c: col_map[col] = "posicao"
            elif "SUBPOSIÇÃO" in c or "SUBPOSICAO" in c: col_map[col] = "subposicao"
            elif "ITEM" in c: col_map[col] = "item"
            elif "DESCRIÇÃO" in c or "DESCRICAO" in c or "DENOMINAÇÃO" in c: col_map[col] = "descricao"
            elif "ALÍQUOTA" in c or "ALIQUOTA" in c or "IPI" in c: col_map[col] = "aliquota"
            elif "UNIDADE" in c or "UND" in c: col_map[col] = "unidade"
        
        df = df.rename(columns=col_map)
        
        # Construir código NCM 8 dígitos: CC.PP.SS.II
        def build_ncm(row):
            parts = []
            for field in ["capitulo", "posicao", "subposicao", "item"]:
                val = str(row.get(field, "")).strip()
                # Remover não-dígitos
                val = re.sub(r'\D', '', val)
                if val:
                    parts.append(val.zfill(2))
            if len(parts) >= 2:
                return f"{parts[0]}.{parts[1]}.{parts[2] if len(parts) > 2 else '00'}.{parts[3] if len(parts) > 3 else '00'}"
            return ""
        
        df["ncm_codigo"] = df.apply(build_ncm, axis=1)
        
        # Alíquota IPI
        if "aliquota" in df.columns:
            df["aliquota_ipi"] = pd.to_numeric(
                df["aliquota"].str.replace(",", ".").str.replace("%", ""),
                errors="coerce"
            ).fillna(0)
        else:
            df["aliquota_ipi"] = 0.0
        
        # Unidade
        if "unidade" not in df.columns:
            df["unidade"] = "UN"
        
        # Filtrar apenas NCMs válidos
        valid = df[df["ncm_codigo"].str.match(r"^\d{2}\.\d{2}\.\d{2}\.\d{2}$", na=False)]
        
        return valid[["ncm_codigo", "descricao", "aliquota_ipi", "unidade"]].copy()
    
    def _parse_local_tipi(self) -> Optional[pd.DataFrame]:
        path = Path(self.LOCAL_TIPI_XLSX)
        if not path.exists():
            return None
        with open(path, "rb") as f:
            return self._parse_tipi_xlsx(f.read())
    
    # ============================================================
    # DOWNLOAD E PARSE CEST
    # ============================================================
    
    def _download_and_parse_cest(self) -> Optional[pd.DataFrame]:
        """Baixa e parseia CEST oficial do Portal SPED / Receita Federal"""
        try:
            # 1. Acessar página do CEST para encontrar link do XLSX
            resp = requests.get(self.CEST_PAGE_URL, timeout=30)
            resp.raise_for_status()
            
            # Parsear HTML para encontrar link do XLSX/CSV
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(resp.text, 'html.parser')
            
            xlsx_link = None
            for link in soup.find_all('a', href=True):
                href = link['href']
                if href.endswith('.xlsx') or href.endswith('.xls') or href.endswith('.csv'):
                    if 'cest' in href.lower() or 'cest' in link.get_text().lower():
                        xlsx_link = href
                        break
            
            if xlsx_link:
                # Baixar o arquivo
                if not xlsx_link.startswith('http'):
                    from urllib.parse import urljoin
                    xlsx_link = urljoin(self.CEST_PAGE_URL, xlsx_link)
                
                self._log(f"Baixando CEST de: {xlsx_link}")
                file_resp = requests.get(xlsx_link, timeout=60)
                file_resp.raise_for_status()
                
                # Salvar localmente
                os.makedirs(os.path.dirname(self.LOCAL_CEST_XLSX), exist_ok=True)
                with open(self.LOCAL_CEST_XLSX, 'wb') as f:
                    f.write(file_resp.content)
                
                # Parsear
                return self._parse_local_cest()
            
        except Exception as e:
            self._log(f"Erro ao baixar CEST online: {e}")
        
        # Fallback: tentar arquivo local
        self._log("Tentando CEST local...")
        return self._parse_local_cest()
    
    def _parse_local_cest(self) -> Optional[pd.DataFrame]:
        path = Path(self.LOCAL_CEST_XLSX)
        if not path.exists():
            return None
        try:
            df = pd.read_excel(path, dtype=str)
            # Normalizar: CEST | NCM | Descrição
            col_map = {}
            for col in df.columns:
                c = str(col).upper()
                if "CEST" in c: col_map[col] = "cest"
                elif "NCM" in c: col_map[col] = "ncm"
                elif "DESCRI" in c or "DENOMINAÇÃO" in c: col_map[col] = "cest_descricao"
            df = df.rename(columns=col_map)
            
            if "cest" not in df.columns or "ncm" not in df.columns:
                log.warning("CEST local não tem colunas esperadas (CEST, NCM)")
                return None
            
            return df[["cest", "ncm", "cest_descricao"]].dropna().copy()
        except Exception as e:
            log.warning(f"Erro ao parsear CEST local: {e}")
            return None
    
    # ============================================================
    # MERGE E VALIDAÇÃO
    # ============================================================
    
    def _merge_tipi_cest(self, tipi_df: pd.DataFrame, cest_df: Optional[pd.DataFrame]) -> pd.DataFrame:
        if cest_df is None or cest_df.empty:
            log.warning("CEST não disponível - publicando apenas TIPI")
            tipi_df["cest"] = None
            tipi_df["cest_descricao"] = ""
            return tipi_df
        
        # Normalizar NCM no CEST (remover pontos para join)
        cest_df = cest_df.copy()
        cest_df["ncm_clean"] = cest_df["ncm"].str.replace(".", "", regex=False).str.zfill(8)
        tipi_df = tipi_df.copy()
        tipi_df["ncm_clean"] = tipi_df["ncm_codigo"].str.replace(".", "", regex=False).str.zfill(8)
        
        # Merge left join (manter todos os NCMs do TIPI)
        merged = tipi_df.merge(
            cest_df[["ncm_clean", "cest", "cest_descricao"]],
            on="ncm_clean",
            how="left"
        )
        
        cest_count = merged["cest"].notna().sum()
        if len(merged) > 0:
            log.info(f"Merge TIPI+CEST: {len(merged)} NCMs, {cest_count} com CEST ({cest_count/len(merged)*100:.1f}%)")
        else:
            log.warning("Merge TIPI+CEST: resultado vazio")
        
        return merged
    
    def _validate(self, df: pd.DataFrame) -> Dict:
        errors = []
        warnings = []
        
        # 1. NCM únicos
        dup = df[df.duplicated(subset=["ncm_codigo"], keep=False)]
        if not dup.empty:
            errors.append(f"NCMs duplicados: {dup['ncm_codigo'].unique()[:10].tolist()}")
        
        # 2. Formato NCM (8 dígitos + pontos)
        invalid = df[~df["ncm_codigo"].str.match(r"^\d{2}\.\d{2}\.\d{2}\.\d{2}$", na=False)]
        if not invalid.empty:
            errors.append(f"NCMs formato inválido: {invalid['ncm_codigo'].head(10).tolist()}")
        
        # 3. Cobertura CEST
        cest_coverage = df["cest"].notna().sum() / len(df) * 100
        if cest_coverage < 20:
            warnings.append(f"Cobertura CEST baixa: {cest_coverage:.1f}% (recomendado >30%)")
        
        # 4. Alíquotas IPI válidas (0-100%)
        if "aliquota_ipi" in df.columns:
            invalid_aliq = df[(df["aliquota_ipi"] < 0) | (df["aliquota_ipi"] > 100)]
            if not invalid_aliq.empty:
                warnings.append(f"Alíquotas IPI fora de range: {len(invalid_aliq)} registros")
        
        # 5. Descrições não vazias
        empty_desc = df[df["descricao"].isna() | (df["descricao"].str.strip() == "")]
        if not empty_desc.empty:
            warnings.append(f"NCMs sem descrição: {len(empty_desc)}")
        
        return {"valid": len(errors) == 0, "errors": errors, "warnings": warnings}
    
    # ============================================================
    # BUILD RECORDS E PUBLICAÇÃO
    # ============================================================
    
    def _build_records(self, df: pd.DataFrame) -> List[NCMRecord]:
        records = []
        for _, row in df.iterrows():
            content = f"{row['ncm_codigo']}|{row['descricao']}|{row.get('cest','')}|{row.get('aliquota_ipi',0)}"
            record_hash = hashlib.sha256(content.encode()).hexdigest()[:16]
            
            records.append(NCMRecord(
                codigo=row["ncm_codigo"],
                descricao=row["descricao"],
                aliq_ipi=float(row.get("aliquota_ipi", 0)),
                cest=row.get("cest") if pd.notna(row.get("cest")) else None,
                cest_descricao=row.get("cest_descricao", "") if pd.notna(row.get("cest_descricao")) else "",
                unidade=row.get("unidade", "UN"),
                vigencia_inicio=datetime.now().strftime("%Y-%m-%d"),
                vigencia_fim=None,
                hash_registro=record_hash
            ))
        return records
    
    def _publish(self):
        # JSON para ClassificadorNCM
        json_data = {}
        for r in self.records:
            key = r.descricao.lower().strip()
            json_data[key] = {
                "ncm": r.codigo,
                "cest": r.cest or "",
                "descricao_oficial": r.descricao,
                "aliquota_ipi": r.aliq_ipi,
                "unidade": r.unidade,
                "cest_descricao": r.cest_descricao
            }
        
        json_path = self.output_dir / "ncm_database.json"
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)
        
        # Parquet para analytics
        parquet_path = self.output_dir / "ncm_database.parquet"
        pd.DataFrame([asdict(r) for r in self.records]).to_parquet(parquet_path, compression="snappy")
        
        # Manifest com metadados de auditoria
        manifest = {
            "generated_at": datetime.now().isoformat(),
            "version": datetime.now().strftime("%Y.%m.%d"),
            "total_records": len(self.records),
            "with_cest": sum(1 for r in self.records if r.cest),
            "schema_version": "2.0",
            "source": "Receita Federal TIPI + Portal SPED CEST",
            "files": {
                "json": str(json_path),
                "parquet": str(parquet_path)
            }
        }
        with open(self.output_dir / "manifest.json", "w") as f:
            json.dump(manifest, f, indent=2)
        
        log.info(f"Publicado: {json_path} ({len(self.records)} NCMs, {manifest['with_cest']} com CEST)")


# ============================================================
# CLASSIFICADOR NCM ENTERPRISE
# ============================================================

class ClassificadorNCMEnterprise:
    """
    Classificador NCM/CEST Enterprise.
    Usa base oficial atualizada + hierarquia NCM + validação CEST.
    Thresholds: ≥90=CLASSIFICADO_ALTA, ≥80=CLASSIFICADO, ≥70=VERIFICAR, <70=REVISÃO_MANUAL
    """
    
    CONFIDENCE_THRESHOLD = 70
    HIGH_CONFIDENCE = 90
    MEDIUM_CONFIDENCE = 80
    
    def __init__(self, db_path: str = "data/processed/ncm_database.json", log_callback=None):
        self.log_callback = log_callback
        self.db: Dict = {}
        self._hierarchy_index = {"capitulos": {}, "ncm_to_key": {}}
        self._load_database(db_path)
        self._build_hierarchy_index()
    
    def _load_database(self, path: str):
        try:
            with open(path, "r", encoding="utf-8") as f:
                self.db = json.load(f)
            self._log(f"Base NCM carregada: {len(self.db)} registros")
        except FileNotFoundError:
            self._log(f"ERRO: Base NCM não encontrada em {path}. Execute o pipeline NCM primeiro.")
            self.db = {}
        except Exception as e:
            self._log(f"ERRO ao carregar base NCM: {e}")
            self.db = {}
    
    def _build_hierarchy_index(self):
        """Índice hierárquico: capítulo → posição → subposição → item"""
        self._hierarchy_index = {"capitulos": {}, "ncm_to_key": {}}
        for key, val in self.db.items():
            ncm = val["ncm"]
            parts = ncm.split(".")
            if len(parts) == 4:
                cap, pos, sub, item = parts
                self._hierarchy_index["capitulos"].setdefault(cap, {}).setdefault(pos, {}).setdefault(sub, {})[item] = key
            self._hierarchy_index["ncm_to_key"][ncm] = key
    
    def _log(self, msg: str):
        log.info(msg)
        if self.log_callback:
            self.log_callback(msg)
    
    def classify(self, description: str, top_k: int = 3) -> List[Dict]:
        """
        Classifica descrição usando fuzzy matching + hierarquia NCM.
        Retorna top_k sugestões ordenadas por confiança.
        """
        from thefuzz import process, fuzz
        import unicodedata
        
        if not description or not self.db:
            return [{"ncm": "", "cest": "", "descricao_tipi": "", "confianca_pct": 0, "status": "❌ BASE_NÃO_CARREGADA"}]
        
        # Normalizar
        norm_desc = self._normalize(description)
        
        # 1. Busca exata
        if norm_desc in self.db:
            val = self.db[norm_desc]
            return [self._format_result(val, 100, "✅ EXATO")]
        
        # 2. Fuzzy matching com token_sort_ratio
        candidates = process.extract(
            norm_desc,
            list(self.db.keys()),
            scorer=fuzz.token_sort_ratio,
            limit=top_k * 3
        )
        
        results = []
        for match_key, score in candidates:
            if score < self.CONFIDENCE_THRESHOLD:
                continue
            
            val = self.db[match_key]
            ncm = val["ncm"]
            
            # Boost hierárquico se mesmo capítulo/posição de matches anteriores
            hierarchy_boost = self._compute_hierarchy_boost(ncm, results)
            final_score = min(100, score + hierarchy_boost)
            
            status = self._determine_status(final_score, val.get("cest"))
            results.append(self._format_result(val, final_score, status))
            
            if len(results) >= top_k:
                break
        
        return results if results else [{
            "ncm": "", "cest": "", "descricao_tipi": "",
            "confianca_pct": 0, "status": "❌ SEM_MATCH"
        }]
    
    def _compute_hierarchy_boost(self, ncm: str, existing_results: List[Dict]) -> int:
        if not existing_results:
            return 0
        parts = ncm.split(".")
        if len(parts) != 4:
            return 0
        cap, pos = parts[0], parts[1]
        for r in existing_results:
            r_parts = r["ncm"].split(".")
            if len(r_parts) == 4 and r_parts[0] == cap and r_parts[1] == pos:
                return 5
            if r_parts[0] == cap:
                return 2
        return 0
    
    def _determine_status(self, score: int, cest: str) -> str:
        if score >= self.HIGH_CONFIDENCE:
            return "✅ CLASSIFICADO_ALTA"
        elif score >= self.MEDIUM_CONFIDENCE:
            return "✅ CLASSIFICADO"
        elif score >= self.CONFIDENCE_THRESHOLD:
            return "⚠️ VERIFICAR_CEST" if not cest else "⚠️ VERIFICAR"
        else:
            return "❌ REVISÃO_MANUAL"
    
    def _format_result(self, val: Dict, score: int, status: str) -> Dict:
        return {
            "ncm": val["ncm"],
            "cest": val.get("cest", ""),
            "descricao_tipi": val["descricao_oficial"],
            "confianca_pct": score,
            "status": status,
            "aliquota_ipi": val.get("aliquota_ipi", 0),
            "unidade": val.get("unidade", "UN"),
            "cest_descricao": val.get("cest_descricao", "")
        }
    
    def _normalize(self, text: str) -> str:
        if not text:
            return ""
        text = str(text).lower().strip()
        text = "".join(c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn")
        return text
    
    def classify_batch(self, descriptions: List[str], progress_callback=None) -> List[List[Dict]]:
        """Classificação em lote com callback de progresso"""
        results = []
        for i, desc in enumerate(descriptions):
            results.append(self.classify(desc))
            if progress_callback:
                progress_callback(int((i + 1) / len(descriptions) * 100))
        return results

    def classificar_planilha(
        self,
        df: pd.DataFrame,
        output_path: str = None,
    ) -> Dict:
        """
        Interface compatível com o classificador v1.
        Processa DataFrame e retorna dict com resultados + arquivo Excel.
        """
        try:
            if self.db is None or len(self.db) == 0:
                return {"success": False, "error": "Base NCM/CEST não carregada. Execute o pipeline primeiro."}
            
            # Detectar coluna de descrição
            desc_col = self._detectar_coluna_descricao(df)
            if not desc_col:
                return {
                    "success": False,
                    "error": "Coluna de descrição não encontrada. Use: descricao, produto, nome, item ou description."
                }

            self._log(f"Classificando {len(df)} produto(s)...")
            total = len(df)

            # Classificar em lote
            results = self.classify_batch(
                df[desc_col].astype(str).tolist(),
                progress_callback=progress_callback
            )

            # Processar resultados
            ncm_list = []
            cest_list = []
            tipi_list = []
            conf_list = []
            status_list = []

            for res in results:
                best = res[0] if res else {"ncm": "", "cest": "", "descricao_tipi": "", "confianca_pct": 0, "status": "❌ SEM_MATCH"}
                ncm_list.append(best["ncm"])
                cest_list.append(best["cest"])
                tipi_list.append(best["descricao_tipi"])
                conf_list.append(best["confianca_pct"])
                status_list.append(best["status"])

            df_result = df.copy()
            df_result["NCM_Sugerido"] = ncm_list
            df_result["CEST_Sugerido"] = cest_list
            df_result["Descrição_TIPI"] = tipi_list
            df_result["Confiança_%"] = conf_list
            df_result["Status_NCM"] = status_list

            # Estatísticas
            total_classificado = sum(1 for s in status_list if "CLASSIFICADO" in s)
            total_verificar = sum(1 for s in status_list if "VERIFICAR" in s)
            total_manual = sum(1 for s in status_list if "REVISÃO_MANUAL" in s)

            self._log(f"✅ Classificados: {total_classificado} | ⚠️ Verificar: {total_verificar} | 🔴 Revisão Manual: {total_manual}")

            # Gerar output
            if not output_path:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                import config as cfg
                output_path = os.path.join(cfg.OUTPUT_DIR, f"classificador_ncm_{ts}.xlsx")

            self._save_excel(df_result, output_path)
            # progress_callback(100) if progress_callback else None

            self._log(f"Relatório salvo: {output_path}")
            return {
                "success": True,
                "rows": total,
                "classificados": total_classificado,
                "verificar": total_verificar,
                "revisao_manual": total_manual,
                "output_path": output_path,
                "dataframe": df_result,
            }

        except Exception as e:
            import traceback
            return {"success": False, "error": f"{e}\n{traceback.format_exc()}"}

    def _detectar_coluna_descricao(self, df: pd.DataFrame) -> Optional[str]:
        """Detecta automaticamente a coluna de descrição do produto."""
        candidatos = [
            "descricao", "descrição", "produto", "nome", "item",
            "description", "name", "product", "titulo", "título",
        ]
        for col in df.columns:
            if col.lower().strip().replace("ã", "a").replace("ç", "c") in candidatos:
                return col
        if len(df.columns) == 1:
            return df.columns[0]
        return None

    def _save_excel(self, df: pd.DataFrame, output_path: str):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Classificação NCM"

        # Paleta
        HEADER_BG = "1C1C1E"
        HEADER_FG = "FFFFFF"
        SUB_BG = "2C2C2E"
        ACCENT = "F5A623"
        GREEN_BG = "D4EDDA"
        WARN_BG = "FFF3CD"
        RED_BG = "F8D7DA"
        ZEBRA = "F9F9F9"
        BORDER_COLOR = "D1D1D6"

        thin = Side(style="thin", color=BORDER_COLOR)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Título
        n_cols = len(df.columns)
        ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
        cell = ws["A1"]
        cell.value = f"🏷️ DataMaster Pro — Classificador NCM/CEST Enterprise | Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        cell.font = Font(name="Calibri", size=13, bold=True, color=HEADER_FG)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # Cabeçalhos
        for c_idx, col in enumerate(df.columns, 1):
            cell = ws.cell(row=2, column=c_idx, value=col)
            cell.font = Font(name="Calibri", size=10, bold=True, color=ACCENT)
            cell.fill = PatternFill("solid", fgColor=SUB_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[2].height = 35

        # Mapeamento de colunas para índice
        cols = list(df.columns)
        status_col_idx = cols.index("Status_NCM") + 1 if "Status_NCM" in cols else None
        conf_col_idx = cols.index("Confiança_%") + 1 if "Confiança_%" in cols else None

        # Dados
        for r_idx, (_, row) in enumerate(df.iterrows(), 3):
            status = str(row.get("Status_NCM", ""))
            bg = ZEBRA if r_idx % 2 == 0 else "FFFFFF"

            for c_idx, col in enumerate(cols, 1):
                val = row.get(col, "")
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

                if "CLASSIFICADO" in status and "Status_NCM" == col:
                    cell.fill = PatternFill("solid", fgColor=GREEN_BG)
                elif "VERIFICAR" in status and "Status_NCM" == col:
                    cell.fill = PatternFill("solid", fgColor=WARN_BG)
                elif "REVISÃO_MANUAL" in status and "Status_NCM" == col:
                    cell.fill = PatternFill("solid", fgColor=RED_BG)
                elif "Confiança" in col and isinstance(val, (int, float)):
                    if val >= 85:
                        cell.fill = PatternFill("solid", fgColor=GREEN_BG)
                    elif val >= self.CONFIDENCE_THRESHOLD:
                        cell.fill = PatternFill("solid", fgColor=WARN_BG)
                    else:
                        cell.fill = PatternFill("solid", fgColor=RED_BG)
                else:
                    cell.fill = PatternFill("solid", fgColor=bg)

        # Largura das colunas
        col_widths = {
            "NCM_Sugerido": 16,
            "CEST_Sugerido": 16,
            "Descrição_TIPI": 40,
            "Confiança_%": 14,
            "Status_NCM": 22,
        }
        for c_idx, col in enumerate(cols, 1):
            width = col_widths.get(col, max(len(col) + 4, 15))
            ws.column_dimensions[get_column_letter(c_idx)].width = width

        ws.freeze_panes = "A3"
        wb.save(output_path)

if __name__ == "__main__":
    import sys
    
    # Setup logging
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)]
    )
    
    # Executar pipeline
    pipeline = NCMPipeline()
    result = pipeline.run_full_pipeline()
    
    if result["success"]:
        print(f"✅ Pipeline concluído: {result['total_ncm']} NCMs, {result['with_cest']} com CEST")
        sys.exit(0)
    else:
        print(f"❌ Pipeline falhou: {result.get('error')}")
        sys.exit(1)