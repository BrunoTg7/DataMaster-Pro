"""
Classificador NCM/CEST Automático v1.0

Classifica produtos de e-commerce com os códigos NCM e CEST corretos
utilizando fuzzy matching contra um banco de dados local (TIPI).

Score > 70  → sugere o NCM com nível de confiança
Score ≤ 70  → marca como ⚠️ REVISÃO MANUAL
"""

import json
import logging
import os
import pandas as pd
from typing import Dict, List, Callable, Optional
from datetime import datetime

log = logging.getLogger(__name__)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Caminho do banco de dados local
_DB_PATH = os.path.join(os.path.dirname(__file__), "ncm_database.json")

# Limiar de confiança para sugestão automática
CONFIDENCE_THRESHOLD = 70


class ClassificadorNCM:
    """
    Motor de classificação NCM/CEST via fuzzy matching.
    Usa fuzzywuzzy (já no requirements.txt) para comparar descrições
    de produtos com as descrições oficiais da TIPI.
    """

    def __init__(self, log_callback: Callable = None, progress_callback: Callable = None):
        self.log_callback = log_callback
        self.progress_callback = progress_callback
        self._db: Dict = {}
        self._descricoes: List[str] = []
        self._load_database()

    def _log(self, msg: str):
        log.info(msg)
        if self.log_callback:
            self.log_callback(msg)

    def _progress(self, pct: int):
        if self.progress_callback:
            self.progress_callback(pct)

    def _load_database(self):
        """Carrega o banco de dados NCM na inicialização."""
        try:
            with open(_DB_PATH, "r", encoding="utf-8") as f:
                self._db = json.load(f)
            self._descricoes = list(self._db.keys())
            self._log(f"Banco NCM carregado: {len(self._descricoes)} categorias")
        except Exception as e:
            self._log(f"⚠️ Erro ao carregar banco NCM: {e}")
            self._db = {}
            self._descricoes = []

    # ──────────────────────────────────────────────────────────────────
    # CLASSIFICAÇÃO DE UM PRODUTO
    # ──────────────────────────────────────────────────────────────────
    def _classificar_um(self, descricao: str) -> Dict:
        """
        Retorna o melhor match NCM para uma descrição de produto.
        Usa token_sort_ratio para ser robusto a ordem de palavras.
        """
        if not descricao or not self._descricoes:
            return {
                "ncm": "",
                "cest": "",
                "descricao_tipi": "",
                "confianca_pct": 0,
                "status": "⚠️ REVISÃO MANUAL",
            }

        try:
            from fuzzywuzzy import process, fuzz

            # Normalizar: minúsculas, remover acentos básicos
            desc_norm = str(descricao).strip().lower()

            # Tentar match exato primeiro (case-insensitive)
            for key in self._descricoes:
                if key.lower() == desc_norm:
                    info = self._db[key]
                    return {
                        "ncm": info.get("ncm", ""),
                        "cest": info.get("cest", ""),
                        "descricao_tipi": info.get("descricao_oficial", key),
                        "confianca_pct": 100,
                        "status": "✅ CLASSIFICADO",
                    }

            # Fuzzy matching
            match, score = process.extractOne(
                desc_norm,
                [d.lower() for d in self._descricoes],
                scorer=fuzz.token_sort_ratio,
            )

            # Mapear de volta para a chave original
            match_key = None
            for key in self._descricoes:
                if key.lower() == match:
                    match_key = key
                    break
            if not match_key:
                match_key = self._descricoes[
                    [d.lower() for d in self._descricoes].index(match)
                ]

            info = self._db[match_key]

            if score >= CONFIDENCE_THRESHOLD:
                return {
                    "ncm": info.get("ncm", ""),
                    "cest": info.get("cest", ""),
                    "descricao_tipi": info.get("descricao_oficial", match_key),
                    "confianca_pct": score,
                    "status": "✅ CLASSIFICADO" if score >= 85 else "⚠️ VERIFICAR",
                }
            else:
                return {
                    "ncm": info.get("ncm", ""),   # sugestão fraca, precisa revisão
                    "cest": info.get("cest", ""),
                    "descricao_tipi": info.get("descricao_oficial", match_key),
                    "confianca_pct": score,
                    "status": "⚠️ REVISÃO MANUAL",
                }

        except ImportError:
            # Fallback sem fuzzywuzzy: busca por substring
            desc_norm = str(descricao).strip().lower()
            best_key = None
            best_score = 0
            for key in self._descricoes:
                if key.lower() in desc_norm or desc_norm in key.lower():
                    words_key = set(key.lower().split())
                    words_desc = set(desc_norm.split())
                    score = len(words_key & words_desc) / max(len(words_key | words_desc), 1) * 100
                    if score > best_score:
                        best_score = score
                        best_key = key

            if best_key and best_score >= CONFIDENCE_THRESHOLD:
                info = self._db[best_key]
                return {
                    "ncm": info.get("ncm", ""),
                    "cest": info.get("cest", ""),
                    "descricao_tipi": info.get("descricao_oficial", best_key),
                    "confianca_pct": int(best_score),
                    "status": "⚠️ VERIFICAR",
                }

            return {
                "ncm": "",
                "cest": "",
                "descricao_tipi": "",
                "confianca_pct": 0,
                "status": "⚠️ REVISÃO MANUAL",
            }

    # ──────────────────────────────────────────────────────────────────
    # CLASSIFICAÇÃO EM LOTE
    # ──────────────────────────────────────────────────────────────────
    def classificar_planilha(
        self,
        df: pd.DataFrame,
        output_path: str = None,
    ) -> Dict:
        """
        Processa DataFrame com coluna de descrição de produtos.
        Adiciona colunas: NCM_Sugerido, CEST_Sugerido, Confiança_%, Descrição_TIPI, Status_NCM.
        """
        try:
            if not self._descricoes:
                return {"success": False, "error": "Banco de dados NCM não foi carregado."}

            # Normalizar coluna de descrição
            desc_col = self._detectar_coluna_descricao(df)
            if not desc_col:
                return {
                    "success": False,
                    "error": "Coluna de descrição não encontrada. Use: descricao, produto, nome, item ou description.",
                }

            self._log(f"Classificando {len(df)} produto(s)...")
            total = len(df)

            resultados_ncm = []
            resultados_cest = []
            resultados_tipi = []
            resultados_conf = []
            resultados_status = []

            for i, (_, row) in enumerate(df.iterrows()):
                desc = str(row.get(desc_col, "")).strip()
                resultado = self._classificar_um(desc)

                resultados_ncm.append(resultado["ncm"])
                resultados_cest.append(resultado["cest"])
                resultados_tipi.append(resultado["descricao_tipi"])
                resultados_conf.append(resultado["confianca_pct"])
                resultados_status.append(resultado["status"])

                self._progress(int(((i + 1) / total) * 80))

            # Adicionar colunas ao DataFrame
            df_result = df.copy()
            df_result["NCM_Sugerido"] = resultados_ncm
            df_result["CEST_Sugerido"] = resultados_cest
            df_result["Descrição_TIPI"] = resultados_tipi
            df_result["Confiança_%"] = resultados_conf
            df_result["Status_NCM"] = resultados_status

            # Estatísticas
            total_classificado = sum(1 for s in resultados_status if "CLASSIFICADO" in s)
            total_verificar = sum(1 for s in resultados_status if "VERIFICAR" in s)
            total_manual = sum(1 for s in resultados_status if "REVISÃO MANUAL" in s)

            self._log(f"✅ Classificados: {total_classificado} | ⚠️ Verificar: {total_verificar} | 🔴 Revisão Manual: {total_manual}")

            # Gerar output
            if not output_path:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                import sys
                sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..")))
                import config as cfg
                output_path = os.path.join(cfg.OUTPUT_DIR, f"classificador_ncm_{ts}.xlsx")

            self._save_excel(df_result, output_path)
            self._progress(100)

            self._log(f"Relatório salvo: {output_path}")
            return {
                "success": True,
                "rows": total,
                "classificados": total_classificado,
                "verificar": total_verificar,
                "revisao_manual": total_manual,
                "output_path": output_path,
                "dataframe": df_result,
            }

        except Exception as e:
            import traceback
            return {"success": False, "error": f"{e}\n{traceback.format_exc()}"}

    def _detectar_coluna_descricao(self, df: pd.DataFrame) -> Optional[str]:
        """Detecta automaticamente a coluna de descrição do produto."""
        candidatos = [
            "descricao", "descrição", "produto", "nome", "item",
            "description", "name", "product", "titulo", "título",
        ]
        for col in df.columns:
            if col.lower().strip().replace("ã", "a").replace("ç", "c") in candidatos:
                return col
        # Se só há uma coluna, usar ela
        if len(df.columns) == 1:
            return df.columns[0]
        return None

    # ──────────────────────────────────────────────────────────────────
    # EXPORTAÇÃO EXCEL PREMIUM
    # ──────────────────────────────────────────────────────────────────
    def _save_excel(self, df: pd.DataFrame, output_path: str):
        wb = Workbook()
        ws = wb.active
        ws.title = "Classificação NCM"

        # Paleta
        HEADER_BG = "1C1C1E"
        HEADER_FG = "FFFFFF"
        SUB_BG = "2C2C2E"
        ACCENT = "F5A623"
        GREEN_BG = "D4EDDA"
        WARN_BG = "FFF3CD"
        RED_BG = "F8D7DA"
        ZEBRA = "F9F9F9"
        BORDER_COLOR = "D1D1D6"

        thin = Side(style="thin", color=BORDER_COLOR)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Título
        n_cols = len(df.columns)
        ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
        cell = ws["A1"]
        cell.value = f"🏷️ DataMaster Pro — Classificador NCM/CEST Automático | Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        cell.font = Font(name="Calibri", size=13, bold=True, color=HEADER_FG)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # Cabeçalhos das colunas
        for c_idx, col in enumerate(df.columns, 1):
            cell = ws.cell(row=2, column=c_idx, value=col)
            cell.font = Font(name="Calibri", size=10, bold=True, color=ACCENT)
            cell.fill = PatternFill("solid", fgColor=SUB_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[2].height = 35

        # Mapeamento de colunas para índice
        cols = list(df.columns)
        status_col_idx = cols.index("Status_NCM") + 1 if "Status_NCM" in cols else None
        conf_col_idx = cols.index("Confiança_%") + 1 if "Confiança_%" in cols else None

        # Dados
        for r_idx, (_, row) in enumerate(df.iterrows(), 3):
            status = str(row.get("Status_NCM", ""))
            bg = ZEBRA if r_idx % 2 == 0 else "FFFFFF"

            for c_idx, col in enumerate(cols, 1):
                val = row.get(col, "")
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

                # Colorir linha pelo status
                if "CLASSIFICADO" in status and "Status_NCM" in col:
                    cell.fill = PatternFill("solid", fgColor=GREEN_BG)
                elif "VERIFICAR" in status and "Status_NCM" in col:
                    cell.fill = PatternFill("solid", fgColor=WARN_BG)
                elif "REVISÃO MANUAL" in status and "Status_NCM" in col:
                    cell.fill = PatternFill("solid", fgColor=RED_BG)
                elif "Confiança" in col and isinstance(val, (int, float)):
                    if val >= 85:
                        cell.fill = PatternFill("solid", fgColor=GREEN_BG)
                    elif val >= CONFIDENCE_THRESHOLD:
                        cell.fill = PatternFill("solid", fgColor=WARN_BG)
                    else:
                        cell.fill = PatternFill("solid", fgColor=RED_BG)
                else:
                    cell.fill = PatternFill("solid", fgColor=bg)

        # Largura das colunas
        col_widths = {
            "NCM_Sugerido": 16,
            "CEST_Sugerido": 16,
            "Descrição_TIPI": 40,
            "Confiança_%": 14,
            "Status_NCM": 22,
        }
        for c_idx, col in enumerate(cols, 1):
            width = col_widths.get(col, max(len(col) + 4, 15))
            ws.column_dimensions[get_column_letter(c_idx)].width = width

        ws.freeze_panes = "A3"
        wb.save(output_path)
