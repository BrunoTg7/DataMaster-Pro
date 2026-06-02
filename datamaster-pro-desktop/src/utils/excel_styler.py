from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional, Tuple
from datetime import datetime
import pandas as pd
import logging

logger = logging.getLogger(__name__)

THEMES = {
    "classic_blue": {
        "header_fill": "1F4E79",
        "header_font_color": "FFFFFF",
        "zebra_fill": "F2F5F8",
        "border_color": "D9D9D9",
        "accent_fill": "DDEBF7",
        "summary_accent": "1F4E79"
    },
    "emerald_green": {
        "header_fill": "1E4620",
        "header_font_color": "FFFFFF",
        "zebra_fill": "F4F9F4",
        "border_color": "D9D9D9",
        "accent_fill": "E2EFDA",
        "summary_accent": "1E4620"
    },
    "modern_orange": {
        "header_fill": "262626",
        "header_font_color": "FFFFFF",
        "zebra_fill": "FFF2E6",
        "border_color": "D9D9D9",
        "accent_fill": "FCE4D6",
        "summary_accent": "E26B0A"
    },
    "slate_gray": {
        "header_fill": "404040",
        "header_font_color": "FFFFFF",
        "zebra_fill": "F2F2F2",
        "border_color": "D9D9D9",
        "accent_fill": "EAEAEA",
        "summary_accent": "595959"
    }
}

THEME_NAMES = {
    "classic_blue": "Azul Corporativo",
    "emerald_green": "Verde Esmeralda",
    "modern_orange": "Laranja Moderno",
    "slate_gray": "Cinza Minimalista",
}

THEME_NAMES_REVERSE = {v: k for k, v in THEME_NAMES.items()}


def _apply_watermark(ws, watermark_text: str = "DataMaster Pro - Versão Gratuita"):
    """Adiciona marca d'água (footer) ao worksheet - identifica versão FREE"""
    try:
        # Adiciona texto ao footer do worksheet como marca d'água visual
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.print_options.horizontalCentered = True
        
        # Adiciona marca d'água via footer (CENTER section)
        from openpyxl.worksheet.header_footer import HeaderFooter
        hf = HeaderFooter()
        hf.cFooter.text = watermark_text
        ws.header_footer = hf
        
    except Exception as e:
        logger.debug(f"Marca d'água via footer não aplicada: {e}")
        # Sem watermark, continua normalmente - não é crítico


def enforce_theme_for_plan(theme_name: str, user_plan: str) -> str:
    """
    Força tema apropriado baseado no plano do usuário
    
    Args:
        theme_name: Tema solicitado
        user_plan: Plano do usuário ('gratis', 'pro', 'enterprise')
        
    Returns:
        Nome do tema a ser usado
    """
    # Plano FREE: forçar sempre classic_blue
    if user_plan == "gratis":
        logger.info(f"Plano FREE: forçando tema classic_blue (solicitado: {theme_name})")
        return "classic_blue"
    
    # Plano PRO/ENTERPRISE: permitir todos os temas
    if theme_name in THEMES:
        return theme_name
    
    # Fallback para classic_blue se tema inválido
    return "classic_blue"


def save_premium_excel(
    df: pd.DataFrame,
    output_path: str,
    theme_name: str = "classic_blue",
    sheet_name: str = "Dados",
    title: str = "RELATÓRIO",
    stats: Optional[List[Tuple[str, str]]] = None,
    diagnostics: Optional[List[Dict]] = None,
    user_plan: str = "gratis",
):
    # Validar e forçar tema baseado no plano
    theme_name = enforce_theme_for_plan(theme_name, user_plan)
    
    theme = THEMES.get(theme_name, THEMES["classic_blue"])
    wb = Workbook()

    title_font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    section_font = Font(name="Segoe UI", size=12, bold=True, color=theme["header_fill"])
    header_font = Font(name="Segoe UI", size=10, bold=True, color=theme["header_font_color"])
    bold_font = Font(name="Segoe UI", size=10, bold=True)
    regular_font = Font(name="Segoe UI", size=10)

    header_fill = PatternFill(start_color=theme["header_fill"], end_color=theme["header_fill"], fill_type="solid")
    zebra_fill = PatternFill(start_color=theme["zebra_fill"], end_color=theme["zebra_fill"], fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color=theme["border_color"]),
        right=Side(style='thin', color=theme["border_color"]),
        top=Side(style='thin', color=theme["border_color"]),
        bottom=Side(style='thin', color=theme["border_color"])
    )

    has_summary = stats or diagnostics

    if has_summary:
        ws_resumo = wb.active
        ws_resumo.title = "📊 Resumo"
        ws_resumo.views.sheetView[0].showGridLines = True

        ws_resumo.merge_cells("A1:D2")
        title_cell = ws_resumo["A1"]
        title_cell.value = f"  {title}"
        title_cell.font = title_font
        title_cell.fill = header_fill
        title_cell.alignment = Alignment(vertical="center", horizontal="left")

        for row in range(1, 3):
            for col in range(1, 5):
                ws_resumo.cell(row=row, column=col).fill = header_fill

        row_idx = 4

        if stats:
            ws_resumo.cell(row=row_idx, column=1, value="Estatísticas Gerais").font = section_font
            row_idx += 1
            for key, val in stats:
                ws_resumo.cell(row=row_idx, column=1, value=key).font = bold_font
                ws_resumo.cell(row=row_idx, column=1).border = thin_border
                ws_resumo.cell(row=row_idx, column=1).fill = zebra_fill
                ws_resumo.cell(row=row_idx, column=2, value=val).font = regular_font
                ws_resumo.cell(row=row_idx, column=2).border = thin_border
                ws_resumo.cell(row=row_idx, column=2).fill = white_fill
                row_idx += 1

        if diagnostics:
            row_idx += 1
            ws_resumo.cell(row=row_idx, column=1, value="Histórico de Importação").font = section_font
            row_idx += 1
            diag_headers = ["Nome do Arquivo", "Status", "Informação / Log"]
            for col_idx, h in enumerate(diag_headers, start=1):
                cell = ws_resumo.cell(row=row_idx, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
            ws_resumo.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=4)
            row_idx += 1

            for item in diagnostics:
                ws_resumo.cell(row=row_idx, column=1, value=item.get("file", "")).font = regular_font
                ws_resumo.cell(row=row_idx, column=1).border = thin_border

                status_cell = ws_resumo.cell(row=row_idx, column=2, value=item.get("status", ""))
                status_cell.font = bold_font
                status_cell.alignment = Alignment(horizontal="center")
                status_cell.border = thin_border

                if item.get("status") in ("Sucesso", "sucesso", "OK"):
                    status_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                else:
                    status_cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

                ws_resumo.cell(row=row_idx, column=3, value=item.get("details", "")).font = regular_font
                ws_resumo.cell(row=row_idx, column=3).border = thin_border
                ws_resumo.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=4)
                ws_resumo.cell(row=row_idx, column=4).border = thin_border
                row_idx += 1

        ws_dados = wb.create_sheet(title=sheet_name)
    else:
        ws_dados = wb.active
        ws_dados.title = sheet_name

    ws_dados.views.sheetView[0].showGridLines = True
    ws_dados.freeze_panes = "A2"

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws_dados.cell(row=1, column=col_idx, value=str(col_name).upper())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border

    for row_idx, row_data in enumerate(df.values, start=2):
        row_fill = zebra_fill if row_idx % 2 == 0 else white_fill
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_dados.cell(row=row_idx, column=col_idx)
            if pd.isna(val):
                cell.value = ""
            else:
                cell.value = val
            cell.font = regular_font
            cell.fill = row_fill
            cell.border = thin_border

            col_name = df.columns[col_idx - 1].lower()

            if isinstance(val, (int, float)):
                currency_keywords = ["valor", "preco", "preço", "total", "custo", "faturamento", "receita", "comissao", "comissão"]
                pct_keywords = ["percent", "%", "taxa", "margem"]
                if any(kw in col_name for kw in currency_keywords):
                    cell.number_format = "R$ #,##0.00"
                elif any(kw in col_name for kw in pct_keywords):
                    if val > 1.0:
                        cell.value = val / 100.0
                    cell.number_format = "0.0%"
                else:
                    cell.number_format = "#,##0"
            elif hasattr(val, "strftime"):
                cell.number_format = "yyyy-mm-dd"

    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
        # Aplicar marca d'água se plano FREE
        if user_plan == "gratis":
            _apply_watermark(ws, "DataMaster Pro - Versão Gratuita")

    wb.save(output_path)
